"""Phase demos as tests (§11).

Phase 0: config validation (a 1-approver config is rejected — four eyes),
         clean empty run, audit chain.
Phase 1: mixed folder files + mocked mailbox, everything registered exactly
         once, rerun produces zero duplicates.
Phase 2: golden set — 100% structured XML, >=95% on text PDFs, review-queue
         routing on low confidence / hard failures.
Phase 3: anomaly rules incl. the CRITICAL IBAN fraud flag, alerts, park file
         with control totals, shadow-mode gate.
Phase 4: 1.500/30.000/500.000 route to the correct chains; four-eyes
         distinctness enforced; chain snapshot survives config changes.
"""

import copy
import glob
import json
import os
import shutil
import sys
import tempfile
import unittest

APP_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.insert(0, APP_DIR)

from src.analyze.history import History  # noqa: E402
from src.approvals import engine  # noqa: E402
from src.audit import log as audit  # noqa: E402
from src.common import db  # noqa: E402
from src.common.config import ConfigError, load_settings, validate_settings  # noqa: E402
from src.extract import normalize  # noqa: E402
from src.pipeline import process_file, run_once  # noqa: E402
from src.validate import master_data  # noqa: E402
from tests.helpers import make_text_pdf, make_ubl  # noqa: E402

CLERK = "a.clerk@okypy.org.cy"
ACCOUNTANT = "b.acc@okypy.org.cy"
CHIEF = "c.chief@okypy.org.cy"
CFO = "d.cfo@okypy.org.cy"
GCFO = "e.gcfo@okypy.org.cy"
CEO = "f.ceo@okypy.org.cy"

VENDOR = ("A. KIRMITSIS LTD", "CY10131613M", "CY19002001300000001112585300")
VENDOR2 = ("A.J. VOUROS LTD", "CY10114836F", "CY50002001940000001100015200")

from datetime import date


def _recent_date():
    return date.today().replace(day=1).isoformat()


def make_env():
    tmp = tempfile.mkdtemp(prefix="invagent_test_")
    settings = load_settings(APP_DIR)
    settings["base_dir"] = tmp
    settings["db_path"] = os.path.join(tmp, "db", "test.db")
    settings["folders"] = {"watch": os.path.join(tmp, "watch"),
                           "archive_root": os.path.join(tmp, "archive"),
                           "output_root": os.path.join(tmp, "output")}
    settings["vendor_master"] = {
        "vendors_csv": os.path.join(APP_DIR, "..", "vendor-cleanup", "input", "vendors_agent.csv"),
        "mapping_csv": os.path.join(tmp, "vendor_mapping.csv"),
    }
    settings["shadow_mode"] = False
    os.makedirs(settings["folders"]["watch"])
    return tmp, settings


def drop(settings, name, data):
    path = os.path.join(settings["folders"]["watch"], name)
    with open(path, "wb") as f:
        f.write(data)
    return path


class MockGraph:
    """Stand-in for GraphClient: 5 messages with one attachment each."""

    def __init__(self, attachments):
        self.messages = [{"id": f"msg{i}"} for i in range(len(attachments))]
        self.attachments = {f"msg{i}": [a] for i, a in enumerate(attachments)}
        self.moved = {}

    def list_unread(self):
        return [m for m in self.messages if m["id"] not in self.moved]

    def get_attachments(self, message_id):
        return self.attachments[message_id]

    def move_message(self, message_id, folder):
        self.moved[message_id] = folder


class TestPhase0Config(unittest.TestCase):
    def test_repo_config_is_valid(self):
        load_settings(APP_DIR)

    def test_one_approver_config_rejected(self):
        s = load_settings(APP_DIR)
        bad = copy.deepcopy(s)
        bad["approval"]["four_eyes_minimum"] = 1
        with self.assertRaises(ConfigError):
            validate_settings(bad)

    def test_single_person_chain_rejected(self):
        s = copy.deepcopy(load_settings(APP_DIR))
        one = "only.person@okypy.org.cy"
        s["approval"]["users"] = {r: [one] for r in s["approval"]["users"]}
        with self.assertRaises(ConfigError):
            validate_settings(s)

    def test_retention_minimum(self):
        s = copy.deepcopy(load_settings(APP_DIR))
        s["audit"]["retention_years"] = 3
        with self.assertRaises(ConfigError):
            validate_settings(s)

    def test_clean_empty_run(self):
        tmp, settings = make_env()
        try:
            conn = db.connect(settings)
            stats = run_once(conn, settings, graph_client=None, send_digest=False)
            self.assertEqual(stats["files"], 0)
            ok, _ = audit.verify_chain(conn)
            self.assertTrue(ok)
            conn.close()
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


class TestPhase1Ingest(unittest.TestCase):
    def test_mixed_ingest_and_rerun_dedupe(self):
        tmp, settings = make_env()
        try:
            conn = db.connect(settings)
            # 10 mixed files in the watch folder
            for i in range(5):
                drop(settings, f"ubl_{i}.xml", make_ubl(*VENDOR[:2], f"WF-{i}", _recent_date(),
                                                        100.0 + i, 19, VENDOR[2],
                                                        [("MAINTENANCE WORK", 1, 100.0 + i)]))
            for i in range(5):
                drop(settings, f"scan_{i}.pdf", make_text_pdf(*VENDOR2[:2], f"WP-{i}", _recent_date(),
                                                              200.0 + i, 19, VENDOR2[2]))
            # 5 emails with one attachment each
            mails = MockGraph([(f"mail_{i}.xml", make_ubl(*VENDOR[:2], f"EM-{i}", _recent_date(),
                                                          300.0 + i, 19, VENDOR[2],
                                                          [("CLEANING SERVICES", 1, 300.0 + i)]))
                               for i in range(5)])
            stats = run_once(conn, settings, graph_client=mails, send_digest=False)
            self.assertEqual(stats["files"], 15)
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM files").fetchone()[0], 15)
            self.assertEqual(set(mails.moved.values()), {"Processed"})

            # rerun: zero new registrations
            mails2 = MockGraph([(f"mail_{i}.xml", make_ubl(*VENDOR[:2], f"EM-{i}", _recent_date(),
                                                           300.0 + i, 19, VENDOR[2],
                                                           [("CLEANING SERVICES", 1, 300.0 + i)]))
                                for i in range(5)])
            stats2 = run_once(conn, settings, graph_client=mails2, send_digest=False)
            self.assertEqual(stats2["files"], 0)
            self.assertEqual(conn.execute("SELECT COUNT(*) FROM files").fetchone()[0], 15)
            conn.close()
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


class TestPhase2Extract(unittest.TestCase):
    def test_golden_set(self):
        xml_total = xml_ok = pdf_fields = pdf_fields_ok = 0
        for exp_path in sorted(glob.glob(os.path.join(APP_DIR, "tests/golden/*.expected.json"))):
            src = exp_path[: -len(".expected.json")]
            with open(exp_path, encoding="utf-8") as f:
                expected = json.load(f)
            record, reason = normalize.extract_file(src, os.path.basename(src))
            self.assertIsNotNone(record, f"{src}: {reason}")
            is_xml = expected["source"] == "xml"
            for k, v in expected.items():
                got = record.get(k)
                match = abs((got or 0) - v) < 0.01 if isinstance(v, float) else got == v
                if is_xml:
                    xml_total += 1
                    xml_ok += match
                else:
                    pdf_fields += 1
                    pdf_fields_ok += match
        self.assertEqual(xml_ok, xml_total, "structured XML must be 100%")
        self.assertGreaterEqual(pdf_fields_ok / pdf_fields, 0.95, "text PDFs must be >=95%")

    def test_hard_failures_route_to_review(self):
        tmp, settings = make_env()
        try:
            conn = db.connect(settings)
            by_vat, _ = master_data.load_vendor_master(settings)
            master_data.sync_vendors_table(conn, settings)
            hist = History(conn, {})
            cases = {
                "bad_arithmetic.xml": make_ubl(*VENDOR[:2], "BAD-1", _recent_date(), 100.0, 19,
                                               VENDOR[2], [("X", 1, 55.0)]),  # lines != net
                "unknown_vendor.xml": make_ubl("GHOST LTD", "CY99999999A", "BAD-2",
                                               _recent_date(), 100.0, 19, VENDOR[2],
                                               [("X", 1, 100.0)]),
                "bad_rate.xml": make_ubl(*VENDOR[:2], "BAD-3", _recent_date(), 100.0, 21,
                                         VENDOR[2], [("X", 1, 100.0)]),
                "future_date.xml": make_ubl(*VENDOR[:2], "BAD-4", "2031-01-01", 100.0, 19,
                                            VENDOR[2], [("X", 1, 100.0)]),
            }
            from src.ingest import dedupe, folder_watch
            for name, data in cases.items():
                drop(settings, name, data)
            for fid in folder_watch.scan(conn, settings):
                iid = process_file(conn, settings, fid, by_vat, hist)
                status = conn.execute("SELECT status, review_reason FROM invoices WHERE id=?",
                                      (iid,)).fetchone()
                self.assertEqual(status["status"], "needs_review", dict(status))
            conn.close()
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    def test_duplicate_invoice_number_blocked(self):
        tmp, settings = make_env()
        try:
            conn = db.connect(settings)
            by_vat, _ = master_data.load_vendor_master(settings)
            master_data.sync_vendors_table(conn, settings)
            hist = History(conn, {})
            from src.ingest import folder_watch
            drop(settings, "a.xml", make_ubl(*VENDOR[:2], "DUP-1", _recent_date(), 100.0, 19,
                                             VENDOR[2], [("MAINTENANCE", 1, 100.0)]))
            # same invoice number, different file content
            drop(settings, "b.xml", make_ubl(*VENDOR[:2], "DUP-1", _recent_date(), 999.0, 19,
                                             VENDOR[2], [("MAINTENANCE", 1, 999.0)]))
            statuses = []
            for fid in folder_watch.scan(conn, settings):
                iid = process_file(conn, settings, fid, by_vat, hist)
                statuses.append(conn.execute(
                    "SELECT status FROM invoices WHERE id=?", (iid,)).fetchone()["status"])
            self.assertIn("needs_review", statuses[1])
            conn.close()
        finally:
            shutil.rmtree(tmp, ignore_errors=True)


class TestPhase3Analyze(unittest.TestCase):
    def setUp(self):
        self.tmp, self.settings = make_env()
        self.conn = db.connect(self.settings)
        master_data.sync_vendors_table(self.conn, self.settings)
        self.by_vat, _ = master_data.load_vendor_master(self.settings)
        self.hist = History(self.conn, {})

    def tearDown(self):
        self.conn.close()
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _ingest(self, name, data):
        from src.ingest import folder_watch
        drop(self.settings, name, data)
        fids = folder_watch.scan(self.conn, self.settings)
        return process_file(self.conn, self.settings, fids[-1], self.by_vat, self.hist)

    def test_iban_mismatch_is_critical_and_holds(self):
        wrong_iban = "CY17002001280000001200527600"
        iid = self._ingest("fraud.xml", make_ubl(VENDOR[0], VENDOR[1], "FR-1", _recent_date(),
                                                 500.0, 19, wrong_iban,
                                                 [("MAINTENANCE", 1, 500.0)]))
        inv = self.conn.execute("SELECT * FROM invoices WHERE id=?", (iid,)).fetchone()
        self.assertEqual(inv["status"], "on_hold")
        f = self.conn.execute(
            "SELECT * FROM findings WHERE invoice_id=? AND rule='iban_mismatch'", (iid,)).fetchone()
        self.assertEqual(f["severity"], "CRITICAL")
        # CRITICAL alert email landed in the outbox
        outbox = glob.glob(os.path.join(self.settings["folders"]["output_root"], "outbox", "*.eml"))
        self.assertTrue(any(b"CRITICAL" in open(p, "rb").read() for p in outbox))

    def test_near_duplicate_and_new_vendor(self):
        iid1 = self._ingest("n1.xml", make_ubl(*VENDOR[:2], "ND-100", _recent_date(), 750.0, 19,
                                               VENDOR[2], [("MAINTENANCE", 1, 750.0)]))
        self.assertTrue(self.conn.execute(
            "SELECT 1 FROM findings WHERE invoice_id=? AND rule='new_vendor'", (iid1,)).fetchone())
        iid2 = self._ingest("n2.xml", make_ubl(*VENDOR[:2], "ND-101", _recent_date(), 750.0, 19,
                                               VENDOR[2], [("MAINTENANCE", 1, 750.0)]))
        self.assertTrue(self.conn.execute(
            "SELECT 1 FROM findings WHERE invoice_id=? AND rule='near_duplicate'", (iid2,)).fetchone())

    def test_price_change_warning(self):
        self._ingest("p1.xml", make_ubl(*VENDOR[:2], "PC-1", _recent_date(), 100.0, 19,
                                        VENDOR[2], [("CLEANING SERVICES MONTHLY", 1, 100.0)]))
        iid = self._ingest("p2.xml", make_ubl(*VENDOR[:2], "PC-2", _recent_date(), 150.0, 19,
                                              VENDOR[2], [("CLEANING SERVICES MONTHLY", 1, 150.0)]))
        f = self.conn.execute(
            "SELECT * FROM findings WHERE invoice_id=? AND rule='price_change'", (iid,)).fetchone()
        self.assertIsNotNone(f)
        self.assertEqual(f["severity"], "WARNING")

    def test_park_file_and_shadow_mode(self):
        from src.sap import park_file
        self._ingest("pk1.xml", make_ubl(*VENDOR[:2], "PK-1", _recent_date(), 400.0, 19,
                                         VENDOR[2], [("MAINTENANCE WORKS", 1, 400.0)]))
        # shadow mode on -> no park file
        self.settings["shadow_mode"] = True
        self.assertEqual(park_file.write_park_files(self.conn, self.settings), {})
        # shadow mode off -> file with control totals; invoice becomes parked
        self.settings["shadow_mode"] = False
        written = park_file.write_park_files(self.conn, self.settings)
        self.assertIn("1000", written)
        from openpyxl import load_workbook
        ws = load_workbook(written["1000"]).active
        rows = list(ws.iter_rows(values_only=True))
        self.assertIn("CONTROL TOTALS", rows[0][0])
        self.assertEqual(rows[1][:3], ("CompanyCode", "VendorAccount", "InvoiceNumber"))
        self.assertEqual(self.conn.execute(
            "SELECT status FROM invoices WHERE invoice_number='PK-1'").fetchone()["status"], "parked")

    def test_gl_unmapped_goes_to_review(self):
        iid = self._ingest("gl.xml", make_ubl(*VENDOR[:2], "GL-1", _recent_date(), 100.0, 19,
                                              VENDOR[2], [("UNMAPPABLE EXOTIC ITEM", 1, 100.0)]))
        inv = self.conn.execute("SELECT * FROM invoices WHERE id=?", (iid,)).fetchone()
        self.assertEqual(inv["status"], "needs_review")
        self.assertIn("no GL rule", inv["review_reason"])


class TestPhase4Approvals(unittest.TestCase):
    def setUp(self):
        self.tmp, self.settings = make_env()
        self.conn = db.connect(self.settings)
        master_data.sync_vendors_table(self.conn, self.settings)
        self.by_vat, _ = master_data.load_vendor_master(self.settings)
        self.hist = History(self.conn, {})

    def tearDown(self):
        self.conn.close()
        shutil.rmtree(self.tmp, ignore_errors=True)

    def _invoice_of(self, gross, number):
        net = round(gross / 1.19, 2)
        from src.ingest import folder_watch
        drop(self.settings, f"{number}.xml",
             make_ubl(*VENDOR[:2], number, _recent_date(), net, 19, VENDOR[2],
                      [("MAINTENANCE CONTRACT", 1, net)]))
        fids = folder_watch.scan(self.conn, self.settings)
        return process_file(self.conn, self.settings, fids[-1], self.by_vat, self.hist)

    def _chain_roles(self, iid):
        return [r["role"] for r in self.conn.execute(
            "SELECT role FROM approvals WHERE invoice_id=? ORDER BY step", (iid,))]

    def test_routing_1500(self):
        iid = self._invoice_of(1500, "V-1500")
        self.assertEqual(self._chain_roles(iid), ["clerk", "accountant"])

    def test_routing_30000(self):
        iid = self._invoice_of(30000, "V-30000")
        self.assertEqual(self._chain_roles(iid),
                         ["clerk", "accountant", "chief_accountant", "cfo"])

    def test_routing_500000(self):
        iid = self._invoice_of(500000, "V-500000")
        self.assertEqual(self._chain_roles(iid),
                         ["clerk", "accountant", "chief_accountant", "cfo", "group_cfo", "ceo"])

    def test_full_approval_and_four_eyes(self):
        iid = self._invoice_of(30000, "V-4EYES")
        # clerk (preparer) cannot act again
        with self.assertRaises(engine.ApprovalError):
            engine.act(self.conn, iid, CLERK, "approve")
        engine.act(self.conn, iid, ACCOUNTANT, "approve")
        # accountant cannot sign the chief step too
        with self.assertRaises(engine.ApprovalError):
            engine.act(self.conn, iid, ACCOUNTANT, "approve")
        engine.act(self.conn, iid, CHIEF, "approve")
        engine.act(self.conn, iid, CFO, "approve")
        self.assertEqual(self.conn.execute(
            "SELECT status FROM invoices WHERE id=?", (iid,)).fetchone()["status"], "approved")

    def test_reject_requires_reason_and_returns_to_review(self):
        iid = self._invoice_of(1500, "V-REJ")
        with self.assertRaises(engine.ApprovalError):
            engine.act(self.conn, iid, ACCOUNTANT, "reject", note="")
        engine.act(self.conn, iid, ACCOUNTANT, "reject", note="wrong cost center")
        self.assertEqual(self.conn.execute(
            "SELECT status FROM invoices WHERE id=?", (iid,)).fetchone()["status"], "rejected")

    def test_chain_snapshot_survives_config_change(self):
        iid = self._invoice_of(30000, "V-SNAP")
        # limits change AFTER routing: active chain finishes under old rules
        self.settings["approval"]["chain"][1]["approve_up_to"] = 100000
        engine.act(self.conn, iid, ACCOUNTANT, "approve")
        engine.act(self.conn, iid, CHIEF, "approve")
        engine.act(self.conn, iid, CFO, "approve")
        self.assertEqual(self.conn.execute(
            "SELECT status FROM invoices WHERE id=?", (iid,)).fetchone()["status"], "approved")
        self.assertEqual(self._chain_roles(iid),
                         ["clerk", "accountant", "chief_accountant", "cfo"])

    def test_delegate(self):
        iid = self._invoice_of(1500, "V-DELEG")
        engine.act(self.conn, iid, ACCOUNTANT, "delegate", delegate_to=CHIEF, note="on leave")
        engine.act(self.conn, iid, CHIEF, "approve")
        self.assertEqual(self.conn.execute(
            "SELECT status FROM invoices WHERE id=?", (iid,)).fetchone()["status"], "approved")

    def test_vendor_mapping_folds_history(self):
        # map an old account onto VENDOR's account; history must combine
        with open(self.settings["vendor_master"]["mapping_csv"], "w", encoding="utf-8") as f:
            f.write("old_supplier,surviving_supplier,effective_date\n100340,100000,2026-07-14\n")
        mapping = master_data.load_mapping(self.settings)
        hist = History(self.conn, mapping)
        self.assertEqual(hist.canonical("100340"), "100000")
        self.assertIn("100340", hist._family("100000"))


if __name__ == "__main__":
    unittest.main(verbosity=2)
