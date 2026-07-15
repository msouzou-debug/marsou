"""Monthly summary — spend per vendor and per entity, finding counts, written
as an xlsx for the finance team."""

import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font

from ..audit import log as audit


def write_summary(conn, settings, month=None, actor="system"):
    month = month or date.today().strftime("%Y-%m")
    wb = Workbook()

    ws = wb.active
    ws.title = "Spend by vendor"
    ws.append(["Vendor", "Name", "Invoices", "Gross"])
    ws["A1"].font = Font(bold=True)
    for r in conn.execute(
            """SELECT vendor_account, vendor_name, COUNT(*) n, SUM(gross_total) g
               FROM invoices WHERE substr(invoice_date,1,7)=? AND status NOT IN ('rejected','needs_review')
               GROUP BY vendor_account ORDER BY g DESC""", (month,)):
        ws.append([r["vendor_account"], r["vendor_name"], r["n"], round(r["g"], 2)])

    ws2 = wb.create_sheet("Findings")
    ws2.append(["Rule", "Severity", "Count"])
    for r in conn.execute(
            """SELECT f.rule, f.severity, COUNT(*) n FROM findings f
               JOIN invoices i ON i.id=f.invoice_id
               WHERE substr(i.invoice_date,1,7)=? GROUP BY f.rule, f.severity ORDER BY n DESC""",
            (month,)):
        ws2.append([r["rule"], r["severity"], r["n"]])

    out = os.path.join(settings["folders"]["output_root"], f"monthly_summary_{month}.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    audit.log(conn, actor, "monthly_summary", f"file={os.path.basename(out)}")
    return out
