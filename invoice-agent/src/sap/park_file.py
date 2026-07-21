"""Stage 6 — park-ready Excel batch, one file per entity per day.

Column layout follows FV60 entry order (header data then line splits) — to be
finalized with the chief accountant in Phase 3. One row per invoice line item,
control-totals header on top. The clerk reviews and enters/uploads; file name
and hash go into the audit log.
"""

import hashlib
import json
import os
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font

from ..audit import log as audit

COLUMNS = ["CompanyCode", "VendorAccount", "InvoiceNumber", "DocumentDate", "PostingDate",
           "Reference", "Currency", "GrossAmount", "TaxCode", "LineNo", "GLAccount",
           "CostCenter", "LineAmount", "LineText"]

# Cyprus VAT rate -> SAP input tax code (confirm with the chief accountant)
TAX_CODES = {19.0: "V9", 9.0: "V5", 5.0: "V3", 3.0: "V2", 0.0: "V0"}


def gl_for_line(line, vendor_account, settings, vendor_name=""):
    """vendor default -> keyword rules (line description OR vendor name) ->
    None (never guess silently)."""
    override = (settings.get("vendors_overrides") or {}).get(vendor_account) or {}
    if override.get("default_gl"):
        return override["default_gl"], override.get("cost_center", "")
    haystack = f"{line.get('description') or ''} {vendor_name or ''}".upper()
    for rule in settings.get("gl_rules", []):
        if any(kw.upper() in haystack for kw in rule.get("keywords", [])):
            return str(rule["gl"]), rule.get("cost_center", "")
    return None, None


def assign_gl(conn, invoice_id, record, vendor_account, settings, master_name=""):
    """Assign GL/cost center to every line. Returns list of unmapped line nos."""
    unmapped = []
    haystack_name = f"{record.get('vendor_name', '')} {master_name}"
    for ln in record.get("lines") or [_whole_invoice_line(record)]:
        gl, cc = gl_for_line(ln, vendor_account, settings, vendor_name=haystack_name)
        if gl is None:
            unmapped.append(ln["line_no"])
        conn.execute(
            "UPDATE invoice_lines SET gl_account=?, cost_center=? WHERE invoice_id=? AND line_no=?",
            (gl or "", cc or "", invoice_id, ln["line_no"]),
        )
    conn.commit()
    return unmapped


def _whole_invoice_line(record):
    rate = next(iter(record.get("net_by_rate") or {19.0: 0}), 19.0)
    return {"line_no": 1, "description": "Invoice total", "quantity": 1,
            "unit_price": record["net_total"], "line_total": record["net_total"],
            "vat_rate": float(rate)}


def write_park_files(conn, settings, actor="system"):
    """One xlsx per entity for all park_ready invoices not yet in a batch."""
    if settings.get("shadow_mode"):
        audit.log(conn, actor, "park_file_skipped", "shadow_mode=true (vendor master gate)")
        return {}
    out_dir = os.path.join(settings["folders"]["output_root"], "park")
    os.makedirs(out_dir, exist_ok=True)
    today = date.today().isoformat()
    written = {}

    for entity in settings["entities"]:
        code = entity["code"]
        invoices = conn.execute(
            "SELECT * FROM invoices WHERE status='park_ready' AND entity=? ORDER BY id",
            (code,)).fetchall()
        if not invoices:
            continue
        wb = Workbook()
        ws = wb.active
        ws.title = f"PARK {code} {today}"

        gross_sum = sum(i["gross_total"] for i in invoices)
        ws.append(["CONTROL TOTALS", "", f"Invoices: {len(invoices)}", "",
                   f"Gross: {gross_sum:.2f}", "", f"Entity: {code}", f"Date: {today}"])
        ws["A1"].font = Font(bold=True)
        ws.append(COLUMNS)
        for cell in ws[2]:
            cell.font = Font(bold=True)

        for inv in invoices:
            lines = conn.execute(
                "SELECT * FROM invoice_lines WHERE invoice_id=? ORDER BY line_no",
                (inv["id"],)).fetchall()
            vat_by_rate = json.loads(inv["vat_by_rate"] or "{}")
            main_rate = max(vat_by_rate, key=lambda r: vat_by_rate[r], default="19.0")
            for ln in lines:
                rate = ln["vat_rate"] if ln["vat_rate"] else float(main_rate)
                ws.append([code, inv["vendor_account"], inv["invoice_number"],
                           inv["invoice_date"], today, inv["invoice_number"],
                           inv["currency"], inv["gross_total"],
                           TAX_CODES.get(float(rate), "V9"), ln["line_no"],
                           ln["gl_account"], ln["cost_center"], ln["line_total"],
                           (ln["description"] or "")[:50]])
            conn.execute("UPDATE invoices SET status='parked' WHERE id=?", (inv["id"],))

        path = os.path.join(out_dir, f"{code}_{today}.xlsx")
        wb.save(path)
        with open(path, "rb") as f:
            digest = hashlib.sha256(f.read()).hexdigest()
        conn.commit()
        audit.log(conn, actor, "park_file",
                  f"file={os.path.basename(path)} sha256={digest} invoices={len(invoices)} gross={gross_sum:.2f}")
        written[code] = path
    return written
