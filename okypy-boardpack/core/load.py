# -*- coding: utf-8 -*-
"""
core/load.py — workbook → tidy per-category frames for 2026 and 2025.

Both ``DATA 2026`` and ``DATA 2025`` carry one column per posted month
(January = column H), a running total, a period budget (2026 only), a category
(col C) and a hospital/unit dimension (col E «Νοσηλ.»). We read the monthly
columns directly so month, YTD and the monthly trend all come from one source
and generalise to any MM. EBITDA-basis exclusions (spec §5.1) are applied here;
the excluded pharma pass-through and D&A lines are captured separately for the
analytical P&L tab.

No cloud calls; pure openpyxl.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl

import config


@dataclass
class Category:
    """One P&L line, merged across years, with the monthly breakdown."""
    name: str
    section: str
    months2026: list = field(default_factory=list)   # per-month, Jan..MM
    months2025: list = field(default_factory=list)
    budget_period: float = 0.0                        # col N (cumulative to date)

    # convenience scalars derived after load
    @property
    def m2026(self) -> float:
        return self.months2026[-1] if self.months2026 else 0.0

    @property
    def ytd2026(self) -> float:
        return sum(self.months2026)

    @property
    def m2025(self) -> float:
        return self.months2025[-1] if self.months2025 else 0.0

    @property
    def ytd2025(self) -> float:
        return sum(self.months2025)


@dataclass
class Aside:
    """An excluded line kept aside for the P&L tab (pharma / D&A)."""
    ytd2026: float = 0.0
    ytd2025: float = 0.0
    budget: float = 0.0


@dataclass
class HospAgg:
    """Per-unit aggregation for the Ανά Νοσηλευτήριο tab (EBITDA basis)."""
    r: float = 0.0      # revenue 2026
    e: float = 0.0      # opex 2026 (ex D&A)
    rb: float = 0.0     # revenue budget
    eb: float = 0.0     # opex budget
    d: float = 0.0      # depreciation & provisions 2026
    rnoph: float = 0.0  # revenue ex Public Health
    r25: float = 0.0    # revenue 2025
    e25: float = 0.0    # opex 2025 (ex D&A)


@dataclass
class LoadResult:
    categories: list = field(default_factory=list)
    warnings: list = field(default_factory=list)
    mm: int = 0
    pharma_rev: Aside = field(default_factory=Aside)
    pharma_exp: Aside = field(default_factory=Aside)
    dep: Aside = field(default_factory=Aside)
    hospitals: dict = field(default_factory=dict)   # code → HospAgg
    alles_sub: dict = field(default_factory=dict)   # Άλλες Λειτ. description → [ytd, budget]
    oay_by_hosp: dict = field(default_factory=dict) # INP/EXO category → {hosp code: [ytd, budget]}
    alla_sub26: dict = field(default_factory=dict)   # Άλλα έσοδα: description → [ytd, budget] (2026)
    alla_sub25: dict = field(default_factory=dict)   # Άλλα έσοδα: description → ytd (2025)
    alla_by_hosp: dict = field(default_factory=dict) # Άλλα έσοδα: hosp code → [ytd, budget] (2026)
    tameio_m25: list = field(default_factory=list)   # Ταμείο Ιατροφ. prior-year recovery, 2025 monthly
    payroll: dict = field(default_factory=dict)      # payroll detail sheets

    def revenue(self):
        return [c for c in self.categories if c.section == config.FLAG_REVENUE]

    def expense(self):
        return [c for c in self.categories if c.section == config.FLAG_EXPENSE]


# ── filename month detection ─────────────────────────────────────────────────
_MM_RE = re.compile(r"(\d{1,2})")


def detect_month_from_filename(filename: str) -> int | None:
    stem = filename.rsplit("/", 1)[-1]
    months = [int(x) for x in _MM_RE.findall(stem) if 1 <= int(x) <= 12]
    if len(months) >= 2:
        return months[1]      # skip the "01-" prefix
    return months[0] if months else None


def _num(v) -> float:
    if hasattr(v, "value"):
        v = v.value
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(" ", "").replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _text(row, idx: int) -> str:
    if idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


_LAT2GR = str.maketrans("ABEZHIKMNOPTXY", "ΑΒΕΖΗΙΚΜΝΟΡΤΧΥ")


def _norm_gr(s: str) -> str:
    """Uppercase + Latin→Greek homoglyphs + strip accents, for token matching."""
    import unicodedata
    s = (s or "").translate(_LAT2GR).upper()
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def load_workbook(path: str, mm: int) -> LoadResult:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    warnings: list = []
    for sh in (config.SHEET_2026, config.SHEET_2025):
        if sh not in wb.sheetnames:
            raise ValueError(f"Λείπει το φύλλο «{sh}».")

    jan = config.COL_2025_JAN  # column H (0-based 7) — same layout both years
    rec: dict = {}
    pharma_rev, pharma_exp, dep = Aside(), Aside(), Aside()
    alles_sub: dict = {}
    oay_by_hosp: dict = {config.INP_CATEGORY: {}, config.EXO_CATEGORY: {}}
    alla_sub26: dict = {}
    alla_sub25: dict = {}
    alla_by_hosp: dict = {}
    tameio_m25 = [0.0] * mm

    def month_cols(row):
        return [_num(row[jan + k]) if jan + k < len(row) else 0.0 for k in range(mm)]

    # ── DATA 2026 ────────────────────────────────────────────────────────────
    for row in wb[config.SHEET_2026].iter_rows(min_row=2, values_only=True):
        section = _text(row, config.COL_SECTION)
        if section not in (config.FLAG_REVENUE, config.FLAG_EXPENSE):
            continue
        name = _text(row, config.COL_CATEGORY)
        months = month_cols(row)
        budget = _num(row[config.COL_2026_BUDGET]) if config.COL_2026_BUDGET < len(row) else 0.0
        aside = _classify_aside(section, name)
        if aside == "pharma_rev":
            pharma_rev.ytd2026 += sum(months); pharma_rev.budget += budget
        elif aside == "pharma_exp":
            pharma_exp.ytd2026 += sum(months); pharma_exp.budget += budget
        elif aside == "dep":
            dep.ytd2026 += sum(months); dep.budget += budget
        elif not _is_excluded(section, name, row, 2026):
            c = rec.get((section, name)) or Category(name=name, section=section,
                                                     months2026=[0.0] * mm, months2025=[0.0] * mm)
            for k in range(mm):
                c.months2026[k] += months[k]
            c.budget_period += budget
            rec[(section, name)] = c
            if name == config.ALLES_CATEGORY:      # sub-analysis by description (col G)
                desc = _text(row, config.DESC_COL) or "(κενό)"
                slot = alles_sub.setdefault(desc, [0.0, 0.0])
                slot[0] += sum(months); slot[1] += budget
            if name in oay_by_hosp:                 # inpatient/outpatient by unit (col E)
                code = _text(row, config.HOSPITAL_DIM_COL)
                if code:
                    slot = oay_by_hosp[name].setdefault(code, [0.0, 0.0])
                    slot[0] += sum(months); slot[1] += budget
            if name == config.ALLA_CATEGORY:        # Άλλα έσοδα by description + unit
                desc = _text(row, config.DESC_COL) or "(κενό)"
                s1 = alla_sub26.setdefault(desc, [0.0, 0.0]); s1[0] += sum(months); s1[1] += budget
                code = _text(row, config.HOSPITAL_DIM_COL)
                if code:
                    s2 = alla_by_hosp.setdefault(code, [0.0, 0.0]); s2[0] += sum(months); s2[1] += budget

    # ── DATA 2025 ────────────────────────────────────────────────────────────
    for row in wb[config.SHEET_2025].iter_rows(min_row=2, values_only=True):
        section = _text(row, config.COL_SECTION)
        if section not in (config.FLAG_REVENUE, config.FLAG_EXPENSE):
            continue
        name = _text(row, config.COL_CATEGORY)
        months = month_cols(row)
        aside = _classify_aside(section, name)
        if aside == "pharma_rev":
            pharma_rev.ytd2025 += sum(months)
        elif aside == "pharma_exp":
            pharma_exp.ytd2025 += sum(months)
        elif aside == "dep":
            dep.ytd2025 += sum(months)
        elif not _is_excluded(section, name, row, 2025):
            c = rec.get((section, name)) or Category(name=name, section=section,
                                                     months2026=[0.0] * mm, months2025=[0.0] * mm)
            for k in range(mm):
                c.months2025[k] += months[k]
            rec[(section, name)] = c
            if name == config.ALLA_CATEGORY:
                desc = _text(row, config.DESC_COL) or "(κενό)"
                alla_sub25[desc] = alla_sub25.get(desc, 0.0) + sum(months)
                # Ταμείο prior-year recovery: capture its 2025 monthly so the
                # monthly trend can be shown net of it (deck convention)
                if config.TAMEIO_RECOVERY_TOKEN in _norm_gr(desc):
                    for k in range(mm):
                        tameio_m25[k] += months[k]

    wb.close()
    # Drop blank-name lines (unallocated / HO-allocation noise) — they can't be
    # mapped to a P&L group. The 01870 special fund is already excluded above.
    cats = [c for c in rec.values() if c.name]
    if not cats:
        warnings.append("Δεν βρέθηκαν κατηγορίες — ελέγξτε τους δείκτες στηλών στο config.py.")

    hospitals = _load_hospitals(path, mm)
    from core import payroll as payrollmod
    payroll = payrollmod.load_details(path, mm)
    return LoadResult(categories=cats, warnings=warnings, mm=mm,
                      pharma_rev=pharma_rev, pharma_exp=pharma_exp, dep=dep,
                      hospitals=hospitals, alles_sub=alles_sub, oay_by_hosp=oay_by_hosp,
                      alla_sub26=alla_sub26, alla_sub25=alla_sub25, alla_by_hosp=alla_by_hosp,
                      tameio_m25=tameio_m25, payroll=payroll)


def _load_hospitals(path: str, mm: int) -> dict:
    """Aggregate EBITDA-basis figures per unit (col E), for the hospitals tab."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    jan = config.COL_2025_JAN
    hc = config.HOSPITAL_DIM_COL
    agg: dict = {}

    def get(code):
        if code not in agg:
            agg[code] = HospAgg()
        return agg[code]

    def scan(sheet, year):
        for row in wb[sheet].iter_rows(min_row=2, values_only=True):
            section = _text(row, config.COL_SECTION)
            if section not in (config.FLAG_REVENUE, config.FLAG_EXPENSE):
                continue
            name = _text(row, config.COL_CATEGORY)
            if _classify_aside(section, name) in ("pharma_rev", "pharma_exp"):
                continue  # pharma pass-through excluded both sides
            code = _text(row, hc)
            if not code:
                continue
            ytd = sum(_num(row[jan + k]) if jan + k < len(row) else 0.0 for k in range(mm))
            h = get(code)
            if section == config.FLAG_REVENUE:
                if _is_excluded(section, name, row, year) or not name:
                    continue  # 01870 special fund / blank
                if year == 2026:
                    h.r += ytd
                    h.rb += _num(row[config.COL_2026_BUDGET]) if config.COL_2026_BUDGET < len(row) else 0.0
                    if name != config.PUBLIC_HEALTH_CATEGORY:
                        h.rnoph += ytd
                else:
                    h.r25 += ytd
            else:  # expense
                if name == config.DEP_CATEGORY:
                    if year == 2026:
                        h.d += ytd
                    continue  # D&A excluded from EBITDA opex
                if year == 2026:
                    h.e += ytd
                    h.eb += _num(row[config.COL_2026_BUDGET]) if config.COL_2026_BUDGET < len(row) else 0.0
                else:
                    h.e25 += ytd

    scan(config.SHEET_2026, 2026)
    scan(config.SHEET_2025, 2025)
    wb.close()
    return agg


def _classify_aside(section: str, name: str) -> str | None:
    """Excluded lines we still want for the P&L tab."""
    if section == config.FLAG_REVENUE and name in config.EXCLUDE_REVENUE:
        return "pharma_rev"
    if section == config.FLAG_EXPENSE:
        if name in ("ΑΝΑΛΩΣΗ ΦΑΡΜΑΚΩΝ Β ΦΑΣΗΣ",):
            return "pharma_exp"
        if name == "Αποσβέσεις και προβλέψεις":
            return "dep"
    return None


def _is_excluded(section: str, name: str, row, year: int) -> bool:
    """EBITDA-basis exclusions of spec §5.1 (pharma/D&A handled via _classify_aside)."""
    if section == config.FLAG_REVENUE:
        if name in config.EXCLUDE_REVENUE:
            return True
        if year == 2025:
            if name in config.EXCLUDE_REVENUE_2025_ONLY:
                return True
            if not name:
                art = _text(row, config.ARTICLE_COL)
                if art in config.EXCLUDE_REVENUE_2025_ARTICLES:
                    return True
    elif section == config.FLAG_EXPENSE:
        if name in config.EXCLUDE_EXPENSE:
            return True
    return False
