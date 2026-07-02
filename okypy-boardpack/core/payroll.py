# -*- coding: utf-8 -*-
"""
core/payroll.py — read the payroll detail sheets for the Μισθοδοσία tab.

Sources (all in the monthly workbook, but irregular / not the tidy DATA format):
  * ΟΚΥΠΥ ΜΙΣΘΟΔ DATA  — Σύμβαση ΟΚΥπΥ salary components (monthly + budget + 2025)
  * ΩΡΟΜΙΣΘΙΟ DATA     — Ωρομίσθιο components (actual + budget) → waterfall buckets
  * ΥΠΕΡΩΡΙΕΣ ΕΠΙΔΟΜΑΤΑ 2026/2025 — employee-level allowances/overtime → Top-N
  * HEADCOUNT          — Αποσπασμένοι / Ωρομίσθιοι monthly (2025 & 2026)

No budget-by-unit/type exists in the workbook, so the Top-N tables carry actuals
(2026 & 2025) only; their Π/Υ column is shown as «—».
"""
from __future__ import annotations

import collections

import openpyxl

import config


def _num(v) -> float:
    return float(v) if isinstance(v, (int, float)) else 0.0


def _txt(row, i) -> str:
    if i >= len(row) or row[i] is None:
        return ""
    return str(row[i]).strip()


def _strip_prefix(s: str) -> str:
    """'5-Γ.Ν. ΛΕΥΚΩΣΙ' → 'Γ.Ν. ΛΕΥΚΩΣΙ' (drop leading 'N-' code)."""
    if "-" in s and s.split("-", 1)[0].strip().isdigit():
        return s.split("-", 1)[1].strip()
    return s


def load_details(path: str, mm: int) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = {
        "symv": _symv(wb, mm),
        "hora": _hora(wb),
        "headcount": _headcount(wb),
        "overtime": _overtime(wb, mm),
    }
    wb.close()
    return out


def _symv(wb, mm: int) -> list:
    if config.SYMV_SHEET not in wb.sheetnames:
        return []
    ws = wb[config.SYMV_SHEET]
    agg = collections.defaultdict(lambda: {"m": [0.0] * mm, "bud": 0.0, "y25": 0.0})
    for row in ws.iter_rows(min_row=3, values_only=True):
        cat = _txt(row, config.SYMV_CAT_COL)
        if not cat:
            continue
        d = agg[cat]
        for k in range(mm):
            d["m"][k] += _num(row[config.SYMV_M0 + k]) if config.SYMV_M0 + k < len(row) else 0.0
        d["bud"] += _num(row[config.SYMV_BUD]) if config.SYMV_BUD < len(row) else 0.0
        d["y25"] += _num(row[config.SYMV_Y25]) if config.SYMV_Y25 < len(row) else 0.0
    rows = [{"label": config.SYMV_LABELS.get(c, c.capitalize()),
             "m": d["m"], "ytd": sum(d["m"]), "bud": d["bud"], "y25": d["y25"]}
            for c, d in agg.items()]
    rows.sort(key=lambda r: -r["ytd"])
    return rows


def _hora(wb) -> list:
    if config.HORA_SHEET not in wb.sheetnames:
        return []
    ws = wb[config.HORA_SHEET]
    vals = {}
    tot26 = tot_bud = 0.0
    for row in ws.iter_rows(min_row=4, values_only=True):
        nm = _txt(row, config.HORA_NAME)
        v26 = _num(row[config.HORA_VAL]) if config.HORA_VAL < len(row) else 0.0
        bud = _num(row[config.HORA_BUD]) if config.HORA_BUD < len(row) else 0.0
        if not nm or (v26 == 0 and bud == 0):
            continue
        vals[nm] = (vals.get(nm, (0.0, 0.0))[0] + v26, vals.get(nm, (0.0, 0.0))[1] + bud)
        tot26 += v26
        tot_bud += bud
    buckets = []
    named26 = named_bud = 0.0
    for raw, label in config.HORA_BUCKETS:
        v26, bud = vals.get(raw, (0.0, 0.0))
        buckets.append({"label": label, "v2026": v26, "bud": bud})
        named26 += v26
        named_bud += bud
    buckets.insert(3, {"label": config.HORA_RESIDUAL_LABEL,
                       "v2026": tot26 - named26, "bud": tot_bud - named_bud})
    return buckets


def _headcount(wb) -> dict:
    out = {"apos": {"m26": [], "m25": []}, "hora": {"m26": [], "m25": []}}
    if config.HEADCOUNT_SHEET not in wb.sheetnames:
        return out
    ws = wb[config.HEADCOUNT_SHEET]
    for row in ws.iter_rows(values_only=True):
        label = _txt(row, 0)
        key = "apos" if label.startswith("Αποσπ") else ("hora" if label.startswith("Ωρομ") else None)
        if not key:
            continue
        out[key]["m26"] = [_num(row[config.HC_2026_LO + k]) for k in range(5)
                           if config.HC_2026_LO + k < len(row)]
        out[key]["m25"] = [_num(row[config.HC_2025_LO + k]) for k in range(5)
                           if config.HC_2025_LO + k < len(row)]
    return out


def _overtime(wb, mm: int) -> dict:
    def scan(sheet, svc_col, allow_lo, allow_hi, ot_col):
        by_hosp_epid = collections.defaultdict(float)
        by_hosp_yper = collections.defaultdict(float)
        by_type = collections.defaultdict(float)
        if sheet not in wb.sheetnames:
            return by_hosp_epid, by_hosp_yper, by_type
        ws = wb[sheet]
        header = None
        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if ri == 1:  # row 2 = header with type names
                header = row
                continue
            if header is None:
                continue
            svc = _strip_prefix(_txt(row, svc_col))
            if not svc:
                continue
            for i in range(allow_lo, allow_hi + 1):
                v = _num(row[i]) if i < len(row) else 0.0
                if v:
                    by_hosp_epid[svc] += v
                    tname = _txt(header, i) or f"col{i}"
                    by_type[tname] += v
            ov = _num(row[ot_col]) if ot_col < len(row) else 0.0
            if ov:
                by_hosp_yper[svc] += ov
        return by_hosp_epid, by_hosp_yper, by_type

    e26, y26, t26 = scan(config.OVERTIME_SHEET_2026, config.OT_SERVICE_COL,
                         config.OT_ALLOW_LO, config.OT_ALLOW_HI, config.OT_OVERTIME_COL)
    e25, y25, t25 = scan(config.OVERTIME_SHEET_2025, config.OT_2025_SERVICE_COL,
                         config.OT_2025_ALLOW_LO, config.OT_2025_ALLOW_HI, config.OT_2025_OVERTIME_COL)
    return {"epid_hosp26": e26, "epid_hosp25": e25,
            "yper_hosp26": y26, "yper_hosp25": y25,
            "type26": t26, "type25": t25}
