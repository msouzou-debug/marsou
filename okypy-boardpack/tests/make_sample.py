# -*- coding: utf-8 -*-
"""Generate a synthetic 01-05_ΓΙΑ_CLAUDE.xlsx that reproduces the approved deck's
May-2026 figures, so the whole pipeline can be exercised without the real file.

Category YTD / budget / 2025 values are taken from the approved deck's own data
object (D.plRevCats / D.plExpCats). Monthly figures are apportioned so the month
totals tie to the deck (rev €46,6 / exp €50,6). Excluded rows are added to prove
the EBITDA-basis exclusions and the 2025-revenue tie.
"""
import os
import openpyxl

MM = 5
JAN_COL = 8  # Excel column H (1-based) — DATA 2025 January

# (name, ytd2026, budget_period, ytd2025)
REV = [
    ("Ενδονοσοκομειακή Φροντίδα ΟΑΥ", 94911563, 100416667, 97123136),
    ("Εξωνοσοκομειακή Φροντίδα ΟΑΥ",  10715732, 12083333,  10965018),
    ("Ημερήσιες Νοσηλείες ΟΑΥ",       11315931, 11483414,  11425517),
    ("ΤΑΕΠ ΟΑΥ",                      11032340, 10120921,  9922472),
    ("ΥΓΟΣ ΟΑΥ",                      16165561, 16165561,  15848589),
    ("ΥΠΑΣ ΟΑΥ",                      7291949,  7291949,   7148970),
    ("Εγγραφές ΟΑΥ",                  3865657,  3910996,   3393594),
    ("Δημόσια Υγεία / Αρμόδιες Αρχές", 52651722, 48227310, 48129870),
    ("Άλλα έσοδα",                    11055604, 14614968,  14382956),
]
EXP = [
    ("Αποσπασμένο Προσωπικό", 123520104, 120399263, 122369536),
    ("Συμβόλαιο ΟΚΥπΥ",       47105702,  46028509,  43067321),
    ("Ωρομίσθιο Προσωπικό",   23289094,  20833333,  22711263),
    ("Ανάλωση Προμηθειών",    20709823,  20625000,  20625000),
    ("Συντηρήσεις",           8942463,   9679013,   6913009),
    ("Ηλεκτρισμός",           8601726,   9166667,   7057715),
    ("Άλλες Λειτουργικές",    8157333,   6958333,   6964175),
    ("Αγορά Υπηρεσιών",       2965223,   3125000,   2734130),
]

# Month totals (deck May) for apportioning per-category month figures.
REV_M26, REV_Y26 = 46610836, 219006059
REV_M25, REV_Y25 = 43708142, 218340122
EXP_M26, EXP_Y26 = 50640570, 243291468
EXP_M25, EXP_Y25 = 46677578, 232442149


def build(path):
    wb = openpyxl.Workbook()
    ws26 = wb.active
    ws26.title = "DATA 2026"
    ws25 = wb.create_sheet("DATA 2025")

    # headers (row 1) — content irrelevant, just occupy the row
    for ws in (ws26, ws25):
        ws.cell(row=1, column=1, value="SECTION")
        ws.cell(row=1, column=3, value="CATEGORY")

    def write26(section, rows, m_tot, y_tot):
        for name, ytd, bud, _y25 in rows:
            r = ws26.max_row + 1
            ws26.cell(row=r, column=1, value=section)          # A
            ws26.cell(row=r, column=3, value=name)             # C
            ws26.cell(row=r, column=12, value=round(ytd * m_tot / y_tot))  # L month
            ws26.cell(row=r, column=13, value=ytd)             # M YTD
            ws26.cell(row=r, column=14, value=bud)             # N budget

    write26("ΕΣΟΔΑ", REV, REV_M26, REV_Y26)
    write26("EXPENSES", EXP, EXP_M26, EXP_Y26)
    # excluded 2026 rows (pharma both sides + D&A) — must be dropped
    for name, ytd, bud in [("ΑΠΟΖ. ΦΑΡΜΑΚΩΝ Β ΦΑΣΗΣ", 59600000, 50000000)]:
        r = ws26.max_row + 1
        ws26.cell(row=r, column=1, value="ΕΣΟΔΑ")
        ws26.cell(row=r, column=3, value=name)
        ws26.cell(row=r, column=12, value=12000000)
        ws26.cell(row=r, column=13, value=ytd)
        ws26.cell(row=r, column=14, value=bud)
    for name, ytd, bud in [("ΑΝΑΛΩΣΗ ΦΑΡΜΑΚΩΝ Β ΦΑΣΗΣ", 59600000, 50000000),
                           ("Αποσβέσεις και προβλέψεις", 6300000, 6300000)]:
        r = ws26.max_row + 1
        ws26.cell(row=r, column=1, value="EXPENSES")
        ws26.cell(row=r, column=3, value=name)
        ws26.cell(row=r, column=12, value=1260000)
        ws26.cell(row=r, column=13, value=ytd)
        ws26.cell(row=r, column=14, value=bud)

    def write25(section, rows, m_tot, y_tot):
        for name, _ytd26, _bud, y25 in rows:
            r = ws25.max_row + 1
            ws25.cell(row=r, column=1, value=section)  # A
            ws25.cell(row=r, column=3, value=name)     # C
            # spread YTD across Jan..May (H..L); last month carries the May figure
            month = round(y25 * m_tot / y_tot)
            prior = y25 - month
            per = prior / (MM - 1)
            for k in range(MM - 1):
                ws25.cell(row=r, column=JAN_COL + k, value=round(per))
            ws25.cell(row=r, column=JAN_COL + MM - 1, value=y25 - round(per) * (MM - 1))

    write25("ΕΣΟΔΑ", REV, REV_M25, REV_Y25)
    write25("EXPENSES", EXP, EXP_M25, EXP_Y25)
    # 2025 revenue-only exclusion: blank category, article 01870 in col B (~€5.1M)
    r = ws25.max_row + 1
    ws25.cell(row=r, column=1, value="ΕΣΟΔΑ")
    ws25.cell(row=r, column=2, value="01870")   # B article
    # category (C) left blank on purpose
    for k in range(MM):
        ws25.cell(row=r, column=JAN_COL + k, value=1020000)  # 5 × 1.02M ≈ 5.1M

    wb.save(path)
    print("wrote", path)


if __name__ == "__main__":
    out = os.path.join(os.path.dirname(__file__), "01-05_ΓΙΑ_CLAUDE.xlsx")
    build(out)
