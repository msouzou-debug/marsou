"""Δημιουργία workbook νέας χρονιάς από το περσινό.

Χρήση:  python -m daspani.new_year <περσινό.xlsx> <νέο.xlsx> [--timarithmos 12,67%]

Κρατά τη δομή και τις φόρμουλες (ΣΥΝΟΠΤΙΚΟ με SUMIFS, κεφαλίδες «Ανάλυση») και:
- αδειάζει όλες τις γραμμές δεδομένων του «Ανάλυση»,
- αφαιρεί τις περσινές χειροκίνητες προσαρμογές ±0,01 από τις φόρμουλες του ΣΥΝΟΠΤΙΚΟΥ,
- καθαρίζει τους περσινούς αριθμούς «ΚΑΤΑΧΩΡΗΣΗ ΔΑΠΑΝΗΣ» (συμπληρώνονται χειροκίνητα),
- προαιρετικά ενημερώνει το ποσοστό τιμαρίθμου στην κεφαλίδα (στήλη J).
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

from .excel_write import HEADER_ROW, get_analysis_sheet
from .models import normalize_label

SYNOPTIKO_SHEET = "ΣΥΝΟΠΤΙΚΟ"
ADJUSTMENT_RE = re.compile(r"([+-]0[.,]0\d+)\s*$")  # π.χ. =SUMIFS(...)+0.01


def create_new_year(source: str, target: str, timarithmos: str | None = None) -> list[str]:
    notes: list[str] = []
    wb = load_workbook(source)
    ws = get_analysis_sheet(wb)

    if ws.max_row > HEADER_ROW:
        ws.delete_rows(HEADER_ROW + 1, ws.max_row - HEADER_ROW)
        notes.append("Αδειάστηκαν οι γραμμές δεδομένων του «Ανάλυση» (κρατήθηκαν οι κεφαλίδες).")

    if timarithmos:
        cell = ws.cell(row=HEADER_ROW, column=10)  # J: ΤΙΜΑΡΙΘΜΟΣ
        cell.value = f"ΤΙΜΑΡΙΘΜΟΣ\n{timarithmos}"
        notes.append(f"Κεφαλίδα τιμαρίθμου: {timarithmos}")

    for name in wb.sheetnames:
        if normalize_label(name) != normalize_label(SYNOPTIKO_SHEET):
            continue
        syn = wb[name]
        cleaned = 0
        for row in syn.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    new = ADJUSTMENT_RE.sub("", cell.value)
                    if new != cell.value:
                        cell.value = new
                        cleaned += 1
        if cleaned:
            notes.append(f"Αφαιρέθηκαν {cleaned} περσινές προσαρμογές ±0,01 από φόρμουλες του ΣΥΝΟΠΤΙΚΟΥ.")
        # Καθαρισμός περσινών αριθμών καταχώρησης (στήλη «ΚΑΤΑΧΩΡΗΣΗ ΔΑΠΑΝΗΣ» και
        # γραμμές «καταχώρηση - 1000:» / old-new / CR) — μένουν οι ετικέτες.
        cleared = 0
        for row in syn.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                label = normalize_label(cell.value)
                if "ΚΑΤΑΧΩΡΗΣΗ" in label and "ΔΑΠΑΝΗΣ" in label:
                    col = cell.column
                    for r in range(cell.row + 1, syn.max_row + 1):
                        if syn.cell(row=r, column=col).value is not None:
                            syn.cell(row=r, column=col).value = None
                            cleared += 1
                elif label.startswith("ΚΑΤΑΧΩΡΗΣΗ -") or label == "CR":
                    for c in range(cell.column + 1, syn.max_column + 1):
                        for dr in (0, 1):  # και η γραμμή old/new από κάτω
                            v = syn.cell(row=cell.row + dr, column=c).value
                            if v is not None:
                                syn.cell(row=cell.row + dr, column=c).value = None
                                cleared += 1
        if cleared:
            notes.append(f"Καθαρίστηκαν {cleared} περσινές τιμές καταχώρησης/CR στο ΣΥΝΟΠΤΙΚΟ.")

    wb.save(target)
    notes.append(f"Αποθηκεύτηκε: {target}")
    return notes


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        prog="daspani.new_year",
        description="Δημιουργία workbook νέας χρονιάς από το περσινό (άδειο «Ανάλυση», ίδιες φόρμουλες).",
    )
    parser.add_argument("source", help="Το περσινό workbook")
    parser.add_argument("target", help="Το νέο αρχείο που θα δημιουργηθεί")
    parser.add_argument("--timarithmos", help="Νέο ποσοστό τιμαρίθμου για την κεφαλίδα, π.χ. «12,67%%»")
    args = parser.parse_args(argv)

    if not Path(args.source).is_file():
        print(f"Σφάλμα: δεν βρέθηκε το αρχείο {args.source}", file=sys.stderr)
        return 2
    if Path(args.target).exists():
        print(f"Σφάλμα: το {args.target} υπάρχει ήδη — δεν αντικαθίσταται.", file=sys.stderr)
        return 2

    for note in create_new_year(args.source, args.target, args.timarithmos):
        print(f"• {note}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
