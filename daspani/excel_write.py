"""Εγγραφή στο workbook: προσθήκη μπλοκ μήνα στο φύλλο «Ανάλυση».

Δομή workbook (όπως το πραγματικό αρχείο ΧΡΕΩΣΕΙΣ ΕΙΔΙΚΟΙ ΑΣΤΥΦΥΛΑΚΕΣ):

- Φύλλο «Ανάλυση»: ενιαίος πίνακας όλων των μηνών, μία γραμμή ανά άτομο ανά
  μήνα, με κενή γραμμή-διαχωριστικό ανάμεσα στους μήνες.
    A ΜΗΝΑΣ | B Α/Α | C ΑΚΑ | D ΑΔΤ | E Ε/ΑΣΤ | F ΟΝΟΜΑΤΕΠΩΝΥΜΟ
    G ΗΜ. ΤΟΠΟΘΕΤΗΣΗΣ | H ΝΟΣΗΛΕΥΤΗΡΙΟ | I ΒΑΣΙΚΟΣ | J ΤΙΜΑΡΙΘΜΟΣ
    K–N παλιές αυξήσεις (κενές) | O ΑΥΞΗΣΗ | P ΒΑΡΔΙΑ | Q ΚΥΡΙΑΚΗ/ΑΡΓΙΑ
    R =SUM(I:Q) | S =R*24,25% | T =R+S | U =T*10% | V =T+U
- Φύλλο «ΣΥΝΟΠΤΙΚΟ»: SUMIFS σε ολόκληρες στήλες του «Ανάλυση» — οι νέες
  γραμμές πιάνονται αυτόματα, δεν το αγγίζουμε ποτέ.

Το ΑΔΤ και το νοσηλευτήριο (αν λείπει από τις παρατηρήσεις) συμπληρώνονται
από την πιο πρόσφατη προηγούμενη γραμμή του ίδιου ΑΚΑ στο φύλλο.
"""

from __future__ import annotations

import copy
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

from .models import TableRow, hospital_from_remarks, normalize_label, parse_date

ANALYSIS_SHEET = "Ανάλυση"
HEADER_ROW = 1

# Στήλες του φύλλου «Ανάλυση» (1-βάσης).
C_MINAS, C_AA, C_AKA, C_ADT, C_EAST, C_NAME, C_DATE, C_NOSIL = 1, 2, 3, 4, 5, 6, 7, 8
C_BASIC, C_TIM = 9, 10
C_AUXISI, C_BARDIA, C_KYRIAKI = 15, 16, 17
AMOUNT_COLS = {"basic": C_BASIC, "tim": C_TIM, "auxisi": C_AUXISI,
               "bardia": C_BARDIA, "kyriaki": C_KYRIAKI}
FORMULA_TEMPLATES = {
    18: "=SUM(I{r}:Q{r})",   # R: Σύνολο μισθών και επιδομάτων
    19: "=R{r}*24.25%",      # S: Εισφορά
    20: "=R{r}+S{r}",        # T: Συνολικό κόστος μισθοδοσίας
    21: "=T{r}*10%",         # U: Διοικητικά έξοδα
    22: "=T{r}+U{r}",        # V: Συνολικό κόστος
}
LAST_COL = 22

DATE_FORMAT = "dd/mm/yyyy"
AMOUNT_FORMAT = "#,##0.00"


class ExcelWriteError(RuntimeError):
    pass


@dataclass
class WriteResult:
    written: int = 0
    replaced_existing: bool = False
    filled_from_history: int = 0
    missing_adt: list[str] = field(default_factory=list)
    missing_hospital: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    output_path: str = ""
    first_row: int = 0
    last_row: int = 0


def _norm_aka(value) -> str:
    return re.sub(r"\D", "", str(value or "")).lstrip("0")


def get_analysis_sheet(wb):
    for name in wb.sheetnames:
        if normalize_label(name) == normalize_label(ANALYSIS_SHEET):
            return wb[name]
    raise ExcelWriteError(
        f"Δεν βρέθηκε φύλλο «{ANALYSIS_SHEET}» στο workbook. "
        f"Διαθέσιμα φύλλα: {', '.join(wb.sheetnames)}"
    )


def last_data_row(ws) -> int:
    """Τελευταία γραμμή με τιμή στη στήλη A (ΜΗΝΑΣ). Η κεφαλίδα αν είναι άδειο."""
    for r in range(ws.max_row, HEADER_ROW, -1):
        if ws.cell(row=r, column=C_MINAS).value not in (None, ""):
            return r
    return HEADER_ROW


def month_block(ws, month: str) -> tuple[int, int] | None:
    """(πρώτη, τελευταία) γραμμή του μπλοκ ενός μήνα, ή None αν δεν υπάρχει."""
    target = normalize_label(month)
    first = last = None
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=C_MINAS).value
        if v and normalize_label(v) == target:
            if first is None:
                first = r
            last = r
    return (first, last) if first is not None else None


def existing_months(ws) -> list[str]:
    seen: list[str] = []
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        v = ws.cell(row=r, column=C_MINAS).value
        if v and str(v) not in seen:
            seen.append(str(v))
    return seen


def build_history(ws, month: str) -> dict[str, dict]:
    """ΑΚΑ -> {adt, nosil} από την πιο πρόσφατη γραμμή ΠΡΙΝ από τον μήνα `month`.

    Οι γραμμές του ίδιου του μήνα (αν υπάρχει ήδη και θα αντικατασταθεί)
    παραλείπονται, ώστε μια επανεγγραφή να μην «κληρονομεί» από τον εαυτό της.
    """
    target = normalize_label(month)
    history: dict[str, dict] = {}
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        m = ws.cell(row=r, column=C_MINAS).value
        if not m or normalize_label(m) == target:
            continue
        aka = _norm_aka(ws.cell(row=r, column=C_AKA).value)
        if not aka:
            continue
        entry = history.setdefault(aka, {})
        adt = ws.cell(row=r, column=C_ADT).value
        nosil = ws.cell(row=r, column=C_NOSIL).value
        if adt not in (None, ""):
            entry["adt"] = adt
        if nosil not in (None, ""):
            entry["nosil"] = nosil
    return history


def build_history_from_workbook(path: str) -> dict[str, dict]:
    """ΑΚΑ -> {adt, nosil} από άλλο workbook (π.χ. το περσινό, για τους πρώτους
    μήνες της νέας χρονιάς που το τρέχον αρχείο δεν έχει ακόμη ιστορικό)."""
    wb = load_workbook(path, read_only=True)
    ws = get_analysis_sheet(wb)
    history: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, max_col=C_NOSIL, values_only=True):
        aka = _norm_aka(row[C_AKA - 1] if len(row) >= C_AKA else None)
        if not aka:
            continue
        entry = history.setdefault(aka, {})
        adt = row[C_ADT - 1] if len(row) >= C_ADT else None
        nosil = row[C_NOSIL - 1] if len(row) >= C_NOSIL else None
        if adt not in (None, ""):
            entry["adt"] = adt
        if nosil not in (None, ""):
            entry["nosil"] = nosil
    wb.close()
    return history


def delete_month_block(ws, month: str) -> bool:
    """Διαγράφει το μπλοκ ενός μήνα (και το διαχωριστικό κενό του)."""
    block = month_block(ws, month)
    if block is None:
        return False
    first, last = block
    # Μαζί και η κενή γραμμή-διαχωριστικό ακριβώς μετά (ή πριν) το μπλοκ.
    count = last - first + 1
    if last + 1 <= ws.max_row and ws.cell(row=last + 1, column=C_MINAS).value in (None, ""):
        count += 1
    elif first > HEADER_ROW + 1 and ws.cell(row=first - 1, column=C_MINAS).value in (None, ""):
        first -= 1
        count += 1
    ws.delete_rows(first, count)
    return True


def refresh_row_formulas(ws) -> None:
    """Ξαναγράφει τις φόρμουλες R–V κάθε γραμμής δεδομένων ώστε να δείχνουν
    στη δική τους γραμμή — απαραίτητο μετά από διαγραφή ενδιάμεσου μπλοκ,
    γιατί το openpyxl δεν αναπροσαρμόζει τις φόρμουλες κατά το delete_rows."""
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        if ws.cell(row=r, column=C_MINAS).value in (None, ""):
            continue
        if ws.cell(row=r, column=18).value in (None, ""):
            continue  # γραμμή χωρίς φόρμουλες — δεν προσθέτουμε
        for col, formula in FORMULA_TEMPLATES.items():
            ws.cell(row=r, column=col).value = formula.format(r=r)


def _copy_style(ws, src_row: int, dst_row: int) -> None:
    for col in range(1, LAST_COL + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst._style = copy.copy(src._style)


def _style_template_row(ws) -> int | None:
    """Μια υπάρχουσα γραμμή δεδομένων για αντιγραφή μορφοποίησης (η τελευταία)."""
    last = last_data_row(ws)
    return last if last > HEADER_ROW else None


def append_month_block(
    ws, month: str, rows: list[TableRow], extra_history: dict[str, dict] | None = None
) -> WriteResult:
    """Προσθέτει το μπλοκ του μήνα στο τέλος του «Ανάλυση».

    Το extra_history (π.χ. από το περσινό workbook) χρησιμοποιείται όπου το
    τρέχον αρχείο δεν έχει δικό του ιστορικό για ένα ΑΚΑ.
    """
    result = WriteResult()

    history = dict(extra_history or {})
    for aka, entry in build_history(ws, month).items():
        merged = {**history.get(aka, {}), **entry}
        history[aka] = merged
    template = _style_template_row(ws)

    last = last_data_row(ws)
    start = HEADER_ROW + 1 if last == HEADER_ROW else last + 2  # 1 κενή γραμμή ανάμεσα στους μήνες

    for i, row in enumerate(rows):
        r = start + i
        if template:
            _copy_style(ws, template, r)

        past = history.get(row.aka_norm, {})

        ws.cell(row=r, column=C_MINAS).value = month
        if row.aa:
            try:
                ws.cell(row=r, column=C_AA).value = int(re.sub(r"\D", "", row.aa))
            except ValueError:
                ws.cell(row=r, column=C_AA).value = row.aa
        if row.aka_norm:
            ws.cell(row=r, column=C_AKA).value = int(row.aka_norm)

        adt = row.adt or past.get("adt")
        if adt not in (None, ""):
            ws.cell(row=r, column=C_ADT).value = adt
            if not row.adt:
                result.filled_from_history += 1
        elif row.aka_norm:
            result.missing_adt.append(f"{row.aka} {row.name}".strip())

        if row.east:
            try:
                ws.cell(row=r, column=C_EAST).value = int(re.sub(r"\D", "", row.east))
            except ValueError:
                ws.cell(row=r, column=C_EAST).value = row.east

        ws.cell(row=r, column=C_NAME).value = row.name

        date_value = parse_date(row.date)
        date_cell = ws.cell(row=r, column=C_DATE)
        if "ΑΠΟ" in normalize_label(row.date):
            date_cell.value = row.date  # αναλογία «ΑΠΟ ...» — μένει ως κείμενο
        elif date_value is not None:
            date_cell.value = date_value
            date_cell.number_format = DATE_FORMAT if not template else date_cell.number_format
        elif row.date:
            date_cell.value = row.date

        hospital = row.hospital or hospital_from_remarks(row.remarks) or past.get("nosil")
        if hospital:
            ws.cell(row=r, column=C_NOSIL).value = int(hospital)
        else:
            result.missing_hospital.append(f"{row.aka} {row.name}".strip())
            row.flag("Δεν προσδιορίστηκε νοσηλευτήριο — συμπληρώστε τη στήλη H χειροκίνητα")

        for fld, col in AMOUNT_COLS.items():
            cell = ws.cell(row=r, column=col)
            cell.value = round(getattr(row, fld), 2)
            if not template:
                cell.number_format = AMOUNT_FORMAT
        for col, formula in FORMULA_TEMPLATES.items():
            ws.cell(row=r, column=col).value = formula.format(r=r)

        result.written += 1

    result.first_row, result.last_row = start, start + len(rows) - 1
    if result.missing_adt:
        result.warnings.append(
            "Χωρίς ΑΔΤ (νέα πρόσληψη; συμπληρώστε τη στήλη D): "
            + "; ".join(result.missing_adt)
        )
    if result.missing_hospital:
        result.warnings.append(
            "Χωρίς νοσηλευτήριο (στήλη H): " + "; ".join(result.missing_hospital)
        )
    return result


def output_path_for(workbook_path: str, month: str):
    from pathlib import Path

    p = Path(workbook_path)
    return p.with_name(f"{p.stem}_{month}{p.suffix}")


def write_workbook(
    workbook_path: str,
    month: str,
    rows: list[TableRow],
    replace_existing: bool = False,
    save_path: str | None = None,
    history_workbook: str | None = None,
) -> WriteResult:
    """Ανοίγει το workbook, προσθέτει το μπλοκ του μήνα και αποθηκεύει.

    Αν ο μήνας υπάρχει ήδη: με replace_existing=True το παλιό μπλοκ
    διαγράφεται και ξαναγράφεται (idempotency), αλλιώς σφάλμα.
    """
    wb = load_workbook(workbook_path)  # data_only=False: διατηρούνται οι φόρμουλες
    ws = get_analysis_sheet(wb)

    replaced = was_middle = False
    block = month_block(ws, month)
    if block is not None:
        if not replace_existing:
            raise ExcelWriteError(
                f"Ο μήνας {month} υπάρχει ήδη στο φύλλο «{ws.title}». "
                "Χρησιμοποιήστε αντικατάσταση για να ξαναγραφτεί χωρίς διπλοεγγραφή."
            )
        was_middle = block[1] < last_data_row(ws)
        delete_month_block(ws, month)
        if was_middle:
            refresh_row_formulas(ws)
        replaced = True

    extra_history = build_history_from_workbook(history_workbook) if history_workbook else None
    result = append_month_block(ws, month, rows, extra_history=extra_history)
    result.replaced_existing = replaced
    if was_middle:
        result.warnings.append(
            f"Ο μήνας {month} δεν ήταν ο τελευταίος: το μπλοκ του ξαναγράφτηκε στο ΤΕΛΟΣ "
            "του φύλλου και οι φόρμουλες R–V των υπόλοιπων γραμμών ανανεώθηκαν. "
            "Αν η στήλη Α/Α άλλων μηνών χρησιμοποιεί φόρμουλες (=1+B...), ελέγξτε τις."
        )

    out = save_path or str(output_path_for(workbook_path, month))
    wb.save(out)
    result.output_path = str(out)
    return result
