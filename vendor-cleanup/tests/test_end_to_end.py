"""End-to-end tests over the real extract in input/.

Covers the brief's acceptance gates:
- Phase 1: detection finds the known §2 patterns (ELYSEE/ELYSSE, the
  100340/101762 VAT pair, ~290 shared-IBAN vendors, worklist counts).
- Phase 2/3 demo: reviewer resolves 10 groups, approver confirms, all four
  outputs are produced and consistent; four-eyes is enforced in code.
- Phase 4: re-import keeps resolutions; ticked SAP actions close; unmerge
  reverses cleanly; audit chain verifies.
"""

import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from vendor_cleanup import audit, db, detect, exports, load, workflow  # noqa: E402
from vendor_cleanup.config import load_settings  # noqa: E402
from vendor_cleanup.normalize import normalize_name, transliterate_greek  # noqa: E402

APP_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
REVIEWER = "reviewer@okypy.example"
APPROVER = "chief.accountant@okypy.example"


def fresh_env():
    tmp = tempfile.mkdtemp(prefix="vcleanup_test_")
    os.makedirs(os.path.join(tmp, "input"))
    os.makedirs(os.path.join(tmp, "config"))
    shutil.copy(os.path.join(APP_DIR, "input", "vendors_agent.csv"),
                os.path.join(tmp, "input", "vendors_agent.csv"))
    shutil.copy(os.path.join(APP_DIR, "config", "settings.yaml"),
                os.path.join(tmp, "config", "settings.yaml"))
    return tmp, load_settings(tmp)


class TestNormalize(unittest.TestCase):
    def test_transliteration(self):
        self.assertEqual(transliterate_greek("ΕΛΛΗΝΑΣ"), "ELLINAS")

    def test_name_norm_matches_across_scripts(self):
        a = normalize_name("ΣΠΥΡΟΣ ΣΤΑΥΡΙΝΙΔΗΣ ΛΤΔ")
        b = normalize_name("SPYROS STAVRINIDES LTD")
        from rapidfuzz import fuzz
        self.assertGreaterEqual(fuzz.token_sort_ratio(a, b), 85)

    def test_token_sort_handles_name_order_swap(self):
        self.assertEqual(normalize_name("GEORGIOS STYLIANOU"),
                         normalize_name("STYLIANOU GEORGIOS"))


class TestFullCycle(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.tmp, cls.settings = fresh_env()
        cls.conn = db.connect(cls.settings)
        cls.import_id, cls.vendors, cls.source = load.run_import(cls.conn, cls.settings)
        cls.groups, cls.stats = detect.persist_groups(
            cls.conn, cls.import_id, cls.vendors, cls.settings)
        cls.wl = detect.refresh_worklists(cls.conn, cls.import_id, cls.vendors)

    @classmethod
    def tearDownClass(cls):
        cls.conn.close()
        shutil.rmtree(cls.tmp, ignore_errors=True)

    def group_of(self, supplier):
        row = self.conn.execute(
            """SELECT g.id FROM groups g JOIN group_members m ON m.group_id=g.id
               WHERE m.supplier=? AND g.status != 'stale'""", (supplier,)).fetchone()
        return row["id"] if row else None

    # ---- Phase 1 acceptance -------------------------------------------------
    def test_known_state_counts(self):
        self.assertEqual(len(self.vendors), 6148)
        self.assertEqual(self.wl["no_iban"], 155)
        self.assertEqual(self.wl["no_vat"], 219)
        self.assertEqual(self.wl["blocked_active"], 65)
        # ~290 vendors share an IBAN (brief: 290 in 133 groups, ± union-find)
        iban_vendors = {s for g in self.groups.values() for a, b, rule, *_ in g["edges"]
                        if rule == "same_iban" for s in (a, b)}
        self.assertAlmostEqual(len(iban_vendors), 290, delta=10)

    def test_finds_elysee_elysse(self):
        gid = self.group_of("100298")
        self.assertIsNotNone(gid)
        members = {r["supplier"] for r in self.conn.execute(
            "SELECT supplier FROM group_members WHERE group_id=?", (gid,))}
        self.assertIn("105269", members)  # ELYSSE

    def test_finds_confirmed_vat_pair(self):
        gid = self.group_of("100340")
        members = {r["supplier"] for r in self.conn.execute(
            "SELECT supplier FROM group_members WHERE group_id=?", (gid,))}
        self.assertIn("101762", members)

    def test_all_ibans_pass_mod97(self):
        self.assertEqual(load.invalid_ibans(self.vendors), [])

    # ---- Phase 2/3: demo cycle ------------------------------------------------
    def test_full_review_cycle(self):
        conn = self.conn
        vmap = {v["supplier"]: v for v in self.vendors}
        open_groups = conn.execute(
            "SELECT id FROM groups WHERE status='open' ORDER BY score DESC LIMIT 10").fetchall()
        self.assertEqual(len(open_groups), 10)

        decision_ids = []
        for i, g in enumerate(open_groups):
            members = [r["supplier"] for r in conn.execute(
                "SELECT supplier FROM group_members WHERE group_id=?", (g["id"],))]
            if i < 8:  # 8 merges
                survivor = detect.propose_survivor(members, vmap)
                did = workflow.record_decision(conn, g["id"], "merge", REVIEWER, survivor=survivor)
            else:      # 2 not-duplicates
                did = workflow.record_decision(conn, g["id"], "not_duplicate", REVIEWER,
                                               reason="government collection account")
            decision_ids.append(did)

        # four-eyes: reviewer cannot approve their own decision
        with self.assertRaises(workflow.WorkflowError):
            workflow.approve_decision(conn, decision_ids[0], REVIEWER)

        for did in decision_ids:
            workflow.approve_decision(conn, did, APPROVER)

        resolved = conn.execute(
            "SELECT COUNT(*) c FROM groups WHERE status='resolved'").fetchone()["c"]
        self.assertEqual(resolved, 10)

        # all four outputs
        paths = exports.write_all(conn, self.settings, actor=APPROVER)
        for p in paths.values():
            self.assertTrue(os.path.exists(p), p)

        # mapping rows = losers of the 8 merges; clean file excludes exactly them
        mapping = exports.merge_mapping(conn)
        self.assertGreater(len(mapping), 0)
        import csv
        with open(paths["clean_vendors"], encoding="utf-8") as f:
            clean = {r["Supplier"] for r in csv.DictReader(f)}
        for m in mapping:
            self.assertNotIn(m["old_supplier"], clean)
            self.assertIn(m["surviving_supplier"], clean)
        self.assertEqual(len(clean), 6148 - len(mapping))

        # not_duplicate fed the pair whitelist
        wl = conn.execute("SELECT COUNT(*) c FROM whitelist WHERE kind='pair'").fetchone()["c"]
        self.assertGreater(wl, 0)

        # SAP action list contains one BLOCK per merged-away account
        from openpyxl import load_workbook
        wb = load_workbook(paths["sap_actions"])
        rows = list(wb.active.iter_rows(values_only=True))
        blocks = [r for r in rows[1:] if r[1] == "BLOCK"]
        self.assertEqual(len(blocks), len(mapping))
        self.assertIn("Duplicate of", blocks[0][4])

        # ---- ticked re-import closes the loop
        ws = wb.active
        ws.cell(row=2, column=8, value="X")
        ticked = os.path.join(self.tmp, "ticked.xlsx")
        wb.save(ticked)
        closed = exports.close_ticked_actions(conn, ticked, actor=APPROVER)
        self.assertEqual(closed, 1)

        # ---- unmerge reverses: mapping shrinks, actions cancelled
        merged_group = conn.execute(
            "SELECT id FROM groups WHERE status='resolved' AND resolution='merge' LIMIT 1").fetchone()
        before = len(exports.merge_mapping(conn))
        workflow.unmerge(conn, merged_group["id"], APPROVER, reason="test reversal")
        self.assertLess(len(exports.merge_mapping(conn)), before)
        self.assertEqual(conn.execute(
            "SELECT status FROM groups WHERE id=?", (merged_group["id"],)).fetchone()["status"], "open")

    # ---- Phase 4: re-import keeps resolutions ---------------------------------
    def test_reimport_preserves_resolutions(self):
        conn = self.conn
        resolved_before = conn.execute(
            "SELECT COUNT(*) c FROM groups WHERE status='resolved'").fetchone()["c"]
        import_id, vendors, _ = load.run_import(conn, self.settings)
        _, stats = detect.persist_groups(conn, import_id, vendors, self.settings)
        resolved_after = conn.execute(
            "SELECT COUNT(*) c FROM groups WHERE status='resolved'").fetchone()["c"]
        self.assertEqual(resolved_before, resolved_after)
        self.assertEqual(stats["new"], 0)      # identical data -> no new groups
        self.assertEqual(stats["reopened"], 0)

    def test_audit_chain_verifies(self):
        ok, bad = audit.verify_chain(self.conn)
        self.assertTrue(ok, f"chain broken at {bad}")


if __name__ == "__main__":
    unittest.main(verbosity=2)
