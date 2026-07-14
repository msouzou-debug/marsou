"""Load vendor data into a new import snapshot.

Two input modes, auto-detected from the files present in input/:

- The four SE16N extracts (CSV or XLSX): LFA1, LFB1, LFBK, TIBAN — file names
  must contain the table name (case-insensitive).
- A combined ``vendors_agent.csv`` (Supplier,Name,VAT,TIN,CompanyCodes,IBANs,
  blocked) — the same format the app itself exports; used when the raw
  extracts are not at hand.

Optionally an FBL1N open-items export (file name containing ``fbl1n``):
Supplier, item count / open amount columns.
"""

import csv
import glob
import hashlib
import json
import os
from datetime import datetime, timezone

from . import audit
from .normalize import (
    iban_valid,
    normalize_id,
    normalize_name,
    normalize_phone,
    split_ibans,
    vat_core,
)


def _sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _read_rows(path):
    """Read a CSV or XLSX file into a list of dicts (headers from row 1)."""
    if path.lower().endswith((".xlsx", ".xlsm")):
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
        out = []
        for r in rows:
            out.append({h: ("" if v is None else str(v).strip()) for h, v in zip(headers, r)})
        wb.close()
        return out
    with open(path, encoding="utf-8-sig", newline="") as f:
        return [{k: (v or "").strip() for k, v in row.items()} for row in csv.DictReader(f)]


def _col(row, *candidates):
    """Fetch a value by tolerant header match (case-insensitive substring)."""
    keys = {k.lower().strip(): k for k in row}
    for cand in candidates:
        c = cand.lower()
        if c in keys:
            return row[keys[c]]
    for cand in candidates:
        c = cand.lower()
        for lk, orig in keys.items():
            if c in lk:
                return row[orig]
    return ""


def _find_file(input_dir, token):
    hits = []
    for path in glob.glob(os.path.join(input_dir, "*")):
        if token.lower() in os.path.basename(path).lower() and path.lower().endswith(
            (".csv", ".xlsx", ".xlsm")
        ):
            hits.append(path)
    return sorted(hits)[0] if hits else None


def _pad_supplier(value):
    v = (value or "").strip()
    return v.lstrip("0") or v  # SAP pads supplier numbers with zeros


def build_vendor_record(supplier, name, vat, tin, phone, company_codes, ibans, blocked, cc_blocked=""):
    return {
        "supplier": supplier,
        "name": name,
        "name_norm": normalize_name(name),
        "vat": vat,
        "vat_norm": vat_core(vat),
        "tin": tin,
        "tin_norm": vat_core(tin),
        "telephone": phone,
        "phone_norm": normalize_phone(phone),
        "company_codes": company_codes,
        "ibans": "; ".join(split_ibans(ibans)),
        "blocked": blocked,
        "cc_blocked": cc_blocked,
    }


def load_combined(path):
    vendors = []
    for row in _read_rows(path):
        supplier = _pad_supplier(_col(row, "supplier"))
        if not supplier:
            continue
        vendors.append(
            build_vendor_record(
                supplier=supplier,
                name=_col(row, "name"),
                vat=_col(row, "vat"),
                tin=_col(row, "tin"),
                phone=_col(row, "telephone", "phone"),
                company_codes=_col(row, "companycodes", "company codes"),
                ibans=_col(row, "ibans", "iban"),
                blocked=_col(row, "blocked"),
            )
        )
    return vendors


def load_extracts(lfa1, lfb1, lfbk, tiban):
    # LFA1 — master records
    masters = {}
    for row in _read_rows(lfa1):
        supplier = _pad_supplier(_col(row, "supplier", "lifnr"))
        if not supplier:
            continue
        flags = []
        for flag, label in [
            ("central deletion", "DEL"),
            ("deletion flag", "DEL"),
            ("posting block", "POST"),
            ("central posting", "POST"),
            ("purchasing block", "PURCH"),
            ("central purchasing", "PURCH"),
            ("payment block", "PAY"),
            ("central payment", "PAY"),
        ]:
            v = _col(row, flag)
            if v and v.upper() not in ("", "0", "N", "NO", "FALSE"):
                if label not in flags:
                    flags.append(label)
        masters[supplier] = {
            "name": _col(row, "name 1", "name"),
            "vat": _col(row, "vat registration", "vat"),
            "tin": _col(row, "tax number 1", "tax number", "tin"),
            "phone": _col(row, "telephone 1", "telephone"),
            "blocked": "Y" if flags else "",
        }

    # LFB1 — company codes (+ per-company-code blocks)
    ccodes, cc_blocked = {}, {}
    for row in _read_rows(lfb1):
        supplier = _pad_supplier(_col(row, "supplier", "lifnr"))
        cc = _col(row, "company code", "bukrs")
        if not supplier or not cc:
            continue
        ccodes.setdefault(supplier, []).append(cc)
        pay = _col(row, "payment block")
        dele = _col(row, "deletion", "deletion flag")
        if (pay and pay not in ("", "0")) or (dele and dele.upper() in ("X", "Y", "TRUE")):
            cc_blocked.setdefault(supplier, []).append(cc)

    # TIBAN — bank key -> IBAN (join: country + bank key + account)
    iban_by_key = {}
    for row in _read_rows(tiban):
        key = (
            _col(row, "bank country", "banks"),
            _col(row, "bank key", "bankl"),
            _col(row, "bank account", "bankn"),
        )
        iban = _col(row, "iban")
        if iban:
            iban_by_key[key] = iban

    # LFBK — vendor bank rows, mapped through TIBAN
    ibans = {}
    for row in _read_rows(lfbk):
        supplier = _pad_supplier(_col(row, "supplier", "lifnr"))
        key = (
            _col(row, "bank country", "banks"),
            _col(row, "bank key", "bankl"),
            _col(row, "bank account", "bankn"),
        )
        iban = iban_by_key.get(key)
        if supplier and iban:
            ibans.setdefault(supplier, set()).add(iban)

    vendors = []
    for supplier, m in sorted(masters.items()):
        vendors.append(
            build_vendor_record(
                supplier=supplier,
                name=m["name"],
                vat=m["vat"],
                tin=m["tin"],
                phone=m["phone"],
                company_codes=", ".join(sorted(ccodes.get(supplier, []))),
                ibans="; ".join(sorted(ibans.get(supplier, set()))),
                blocked=m["blocked"],
                cc_blocked=", ".join(sorted(cc_blocked.get(supplier, []))),
            )
        )
    return vendors


def load_fbl1n(conn, path, actor="system"):
    """Load an FBL1N open-items export tied to its own import row."""
    counts = {}
    amounts = {}
    for row in _read_rows(path):
        supplier = _pad_supplier(_col(row, "supplier", "account", "vendor"))
        if not supplier:
            continue
        counts[supplier] = counts.get(supplier, 0) + 1
        raw = _col(row, "amount in local currency", "local currency amount", "amount")
        try:
            amounts[supplier] = amounts.get(supplier, 0.0) + float(
                raw.replace(".", "").replace(",", ".") if raw.count(",") == 1 else raw.replace(",", "")
            )
        except (ValueError, AttributeError):
            pass
    cur = conn.execute(
        "INSERT INTO imports (imported_at, source, files_json, vendor_count) VALUES (?,?,?,?)",
        (
            datetime.now(timezone.utc).isoformat(),
            "fbl1n",
            json.dumps([{"name": os.path.basename(path), "sha256": _sha256(path), "rows": sum(counts.values())}]),
            len(counts),
        ),
    )
    import_id = cur.lastrowid
    conn.executemany(
        "INSERT INTO open_items (import_id, supplier, item_count, open_amount) VALUES (?,?,?,?)",
        [(import_id, s, counts[s], amounts.get(s, 0.0)) for s in counts],
    )
    conn.commit()
    audit.log(conn, actor, "import_fbl1n", f"import_id={import_id} vendors={len(counts)}")
    return import_id


def run_import(conn, settings, actor="system"):
    """Detect input mode, snapshot vendors, return (import_id, vendors, source)."""
    input_dir = settings["paths"]["input_dir"]
    files = {t: _find_file(input_dir, t) for t in ("lfa1", "lfb1", "lfbk", "tiban")}
    combined = _find_file(input_dir, "vendors_agent") or _find_file(input_dir, "vendors")

    if all(files.values()):
        source = "extracts"
        used = [files["lfa1"], files["lfb1"], files["lfbk"], files["tiban"]]
        vendors = load_extracts(*used)
    elif combined:
        source = "combined"
        used = [combined]
        vendors = load_combined(combined)
    else:
        missing = [t.upper() for t, p in files.items() if not p]
        raise FileNotFoundError(
            f"input/ has neither the four extracts (missing: {', '.join(missing)}) "
            "nor a combined vendors_agent.csv"
        )

    files_json = json.dumps(
        [{"name": os.path.basename(p), "sha256": _sha256(p)} for p in used]
    )
    cur = conn.execute(
        "INSERT INTO imports (imported_at, source, files_json, vendor_count) VALUES (?,?,?,?)",
        (datetime.now(timezone.utc).isoformat(), source, files_json, len(vendors)),
    )
    import_id = cur.lastrowid
    conn.executemany(
        """INSERT INTO vendors (import_id, supplier, name, name_norm, vat, vat_norm, tin, tin_norm,
                                telephone, phone_norm, company_codes, ibans, blocked, cc_blocked)
           VALUES (:import_id, :supplier, :name, :name_norm, :vat, :vat_norm, :tin, :tin_norm,
                   :telephone, :phone_norm, :company_codes, :ibans, :blocked, :cc_blocked)""",
        [dict(v, import_id=import_id) for v in vendors],
    )
    conn.commit()

    fbl1n = _find_file(input_dir, "fbl1n")
    if fbl1n:
        load_fbl1n(conn, fbl1n, actor)

    audit.log(conn, actor, "import", f"import_id={import_id} source={source} vendors={len(vendors)}")
    return import_id, vendors, source


def open_items_for(conn, supplier):
    imp = conn.execute(
        "SELECT id FROM imports WHERE source='fbl1n' ORDER BY id DESC LIMIT 1"
    ).fetchone()
    if not imp:
        return None  # no FBL1N data loaded at all
    row = conn.execute(
        "SELECT item_count, open_amount FROM open_items WHERE import_id=? AND supplier=?",
        (imp["id"], supplier),
    ).fetchone()
    return {"item_count": row["item_count"], "open_amount": row["open_amount"]} if row else {
        "item_count": 0,
        "open_amount": 0.0,
    }


def invalid_ibans(vendors):
    bad = []
    for v in vendors:
        for iban in split_ibans(v["ibans"]):
            if not iban_valid(iban):
                bad.append((v["supplier"], iban))
    return bad
