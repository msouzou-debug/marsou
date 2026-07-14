"""The four outputs (§6). The app's exports are instructions for humans —
never transactions against SAP."""

import csv
import html
import os
from datetime import date, datetime, timezone

from openpyxl import Workbook, load_workbook

from . import audit, db
from .load import open_items_for


def _now():
    return datetime.now(timezone.utc).isoformat()


def _out(settings, name):
    out_dir = settings["paths"]["output_dir"]
    os.makedirs(out_dir, exist_ok=True)
    return os.path.join(out_dir, name)


def _release_held_actions(conn, actor):
    """§8: held BLOCKs are released once a later FBL1N shows the items cleared."""
    released = 0
    for row in conn.execute("SELECT * FROM sap_actions WHERE status='held'").fetchall():
        items = open_items_for(conn, row["supplier"])
        if items is not None and items["item_count"] == 0:
            conn.execute("UPDATE sap_actions SET status='pending', hold_reason='' WHERE id=?", (row["id"],))
            released += 1
    if released:
        conn.commit()
        audit.log(conn, actor, "actions_released", f"count={released}")
    return released


def _approved_update_actions(conn):
    """UPDATE rows for missing IBAN/VAT fixes the team confirmed as 'fixed'."""
    for item in conn.execute(
        "SELECT * FROM worklist_items WHERE status='fixed' AND kind IN ('no_iban','no_vat')"
    ).fetchall():
        field = "IBAN" if item["kind"] == "no_iban" else "VAT number"
        note = f"Add missing {field} — vendor cleanup ({item['reason'] or 'confirmed by ' + item['reviewer']})"
        conn.execute(
            """INSERT INTO sap_actions (action, supplier, survivor, note, source, status, created_at)
               VALUES ('UPDATE',?,'',?,?, 'pending',?)
               ON CONFLICT(action, supplier, survivor) DO NOTHING""",
            (item["supplier"], note, f"worklist:{item['id']}", _now()),
        )
    conn.commit()


def write_sap_actions(conn, settings, actor="system"):
    _release_held_actions(conn, actor)
    _approved_update_actions(conn)
    today = date.today().isoformat()
    path = _out(settings, f"sap_actions_{today}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "SAP actions"
    headers = ["ActionID", "Action", "Supplier", "SurvivorAccount", "Note", "Status", "HoldReason", "Applied"]
    ws.append(headers)
    rows = conn.execute(
        "SELECT * FROM sap_actions WHERE status IN ('pending','exported','held') ORDER BY action, supplier"
    ).fetchall()
    for r in rows:
        ws.append([r["id"], r["action"], r["supplier"], r["survivor"], r["note"],
                   r["status"], r["hold_reason"], ""])
    for col, width in zip("ABCDEFGH", (9, 9, 12, 15, 60, 10, 28, 9)):
        ws.column_dimensions[col].width = width
    wb.save(path)

    conn.execute("UPDATE sap_actions SET status='exported' WHERE status='pending'")
    conn.commit()
    audit.log(conn, actor, "export_sap_actions", f"file={os.path.basename(path)} rows={len(rows)}")
    return path


def close_ticked_actions(conn, path, actor="system"):
    """Re-import the ticked action list: rows with a mark in 'Applied' are closed."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "").strip() for h in next(rows)]
    idx = {h: i for i, h in enumerate(headers)}
    closed = 0
    for r in rows:
        applied = str(r[idx.get("Applied", -1)] or "").strip().upper()
        if applied in ("X", "Y", "YES", "TRUE", "1", "OK", "DONE"):
            action_id = r[idx["ActionID"]]
            cur = conn.execute(
                "UPDATE sap_actions SET status='applied', applied_at=? WHERE id=? AND status IN ('exported','pending')",
                (_now(), action_id),
            )
            closed += cur.rowcount
    wb.close()
    conn.commit()
    audit.log(conn, actor, "actions_applied", f"file={os.path.basename(path)} closed={closed}")
    return closed


def merge_mapping(conn):
    """old_supplier -> surviving_supplier for all approved, unrevoked merges."""
    rows = conn.execute(
        """SELECT d.id, d.group_id, d.survivor, d.approved_at
           FROM decisions d JOIN groups g ON g.id = d.group_id
           WHERE d.action='merge' AND d.status='approved' AND g.resolution='merge'"""
    ).fetchall()
    mapping = []
    for d in rows:
        members = [r["supplier"] for r in conn.execute(
            "SELECT supplier FROM group_members WHERE group_id=?", (d["group_id"],))]
        for m in members:
            if m != d["survivor"]:
                mapping.append({"old_supplier": m, "surviving_supplier": d["survivor"],
                                "effective_date": (d["approved_at"] or "")[:10]})
    return sorted(mapping, key=lambda r: r["old_supplier"])


def write_mapping(conn, settings, actor="system"):
    path = _out(settings, "vendor_mapping.csv")
    mapping = merge_mapping(conn)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=["old_supplier", "surviving_supplier", "effective_date"])
        w.writeheader()
        w.writerows(mapping)
    audit.log(conn, actor, "export_mapping", f"rows={len(mapping)}")
    return path


def write_clean_vendors(conn, settings, actor="system"):
    """The contract with the invoice agent — same columns as today, survivors only."""
    path = _out(settings, "vendors_agent.csv")
    merged_away = {m["old_supplier"] for m in merge_mapping(conn)}
    vendors = db.current_vendors(conn)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Supplier", "Name", "VAT", "TIN", "CompanyCodes", "IBANs", "blocked"])
        n = 0
        for v in vendors:
            if v["supplier"] in merged_away:
                continue
            w.writerow([v["supplier"], v["name"], v["vat"], v["tin"],
                        v["company_codes"], v["ibans"], v["blocked"]])
            n += 1
    audit.log(conn, actor, "export_clean_vendors", f"rows={n} excluded={len(merged_away)}")
    return path


def progress_stats(conn):
    s = {}
    s["groups_total"] = conn.execute(
        "SELECT COUNT(*) c FROM groups WHERE status != 'stale'").fetchone()["c"]
    s["groups_resolved"] = conn.execute(
        "SELECT COUNT(*) c FROM groups WHERE status='resolved'").fetchone()["c"]
    s["groups_pending"] = conn.execute(
        "SELECT COUNT(*) c FROM groups WHERE status='pending_approval'").fetchone()["c"]
    s["groups_open"] = conn.execute(
        "SELECT COUNT(*) c FROM groups WHERE status='open'").fetchone()["c"]
    s["merges"] = conn.execute(
        "SELECT COUNT(*) c FROM groups WHERE status='resolved' AND resolution='merge'").fetchone()["c"]
    s["not_duplicates"] = conn.execute(
        "SELECT COUNT(*) c FROM groups WHERE status='resolved' AND resolution='not_duplicate'").fetchone()["c"]
    s["worklists"] = {
        r["kind"]: {"open": r["open_c"], "done": r["done_c"]}
        for r in conn.execute(
            """SELECT kind,
                      SUM(CASE WHEN status IN ('open','pending_approval') THEN 1 ELSE 0 END) open_c,
                      SUM(CASE WHEN status IN ('fixed','accepted','gone') THEN 1 ELSE 0 END) done_c
               FROM worklist_items GROUP BY kind""")
    }
    s["actions"] = {
        r["status"]: r["c"]
        for r in conn.execute("SELECT status, COUNT(*) c FROM sap_actions GROUP BY status")
    }
    s["imports"] = [dict(r) for r in conn.execute(
        "SELECT id, imported_at, source, vendor_count FROM imports WHERE source IN ('extracts','combined') ORDER BY id")]
    return s


def write_progress_report(conn, settings, actor="system"):
    path = _out(settings, "progress_report.html")
    s = progress_stats(conn)
    resolved_pct = 100.0 * s["groups_resolved"] / s["groups_total"] if s["groups_total"] else 0.0
    wl_labels = {"no_iban": "No IBAN", "no_vat": "No VAT number", "blocked_active": "Blocked but active"}
    wl_rows = "".join(
        f"<tr><td>{wl_labels.get(k, k)}</td><td>{v['open']}</td><td>{v['done']}</td></tr>"
        for k, v in sorted(s["worklists"].items())
    )
    action_rows = "".join(
        f"<tr><td>{html.escape(k)}</td><td>{v}</td></tr>" for k, v in sorted(s["actions"].items())
    ) or "<tr><td colspan=2>none yet</td></tr>"
    trend_rows = "".join(
        f"<tr><td>#{i['id']}</td><td>{i['imported_at'][:10]}</td><td>{i['source']}</td><td>{i['vendor_count']}</td></tr>"
        for i in s["imports"]
    )
    body = f"""<!doctype html><html><head><meta charset="utf-8">
<title>Vendor cleanup — progress</title>
<style>body{{font-family:system-ui,sans-serif;max-width:52rem;margin:2rem auto;color:#1a202c}}
h1{{font-size:1.4rem}} table{{border-collapse:collapse;margin:0.6rem 0 1.4rem}}
td,th{{border:1px solid #cbd5e0;padding:0.3rem 0.7rem;text-align:left}}
.big{{font-size:2rem;font-weight:700}} .bar{{background:#e2e8f0;border-radius:4px;height:14px;width:20rem}}
.fill{{background:#2f855a;height:14px;border-radius:4px}}</style></head><body>
<h1>OKYpY vendor master cleanup — progress</h1>
<p>Generated {date.today().isoformat()}</p>
<p><span class="big">{s['groups_resolved']} / {s['groups_total']}</span> duplicate groups resolved
({resolved_pct:.0f}%) — {s['merges']} merges, {s['not_duplicates']} confirmed not duplicates,
{s['groups_pending']} awaiting approval, {s['groups_open']} open.</p>
<div class="bar"><div class="fill" style="width:{resolved_pct:.0f}%"></div></div>
<h2>Worklists</h2>
<table><tr><th>Worklist</th><th>Open</th><th>Done</th></tr>{wl_rows}</table>
<h2>SAP actions</h2>
<table><tr><th>Status</th><th>Count</th></tr>{action_rows}</table>
<h2>Imports</h2>
<table><tr><th>#</th><th>Date</th><th>Source</th><th>Vendors</th></tr>{trend_rows}</table>
</body></html>"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(body)
    audit.log(conn, actor, "export_progress_report", os.path.basename(path))
    return path


def write_all(conn, settings, actor="system"):
    return {
        "sap_actions": write_sap_actions(conn, settings, actor),
        "mapping": write_mapping(conn, settings, actor),
        "clean_vendors": write_clean_vendors(conn, settings, actor),
        "progress": write_progress_report(conn, settings, actor),
    }
