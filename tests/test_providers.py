"""Non-hospital ΟΑΥ providers: a mental-health month, several units at once.

ΟΑΥ pays each mental-health unit on its own cheque and ships the whole month
as one folder — five SRAs, five paid-claims files, five activity exports, no
DRG summary and no pharmacy anywhere.  The batch must therefore be grouped by
PROVIDER (from file content only) and reconciled per provider, then presented
in one workbook with a consolidated summary.
"""
import io

import pytest
from openpyxl import load_workbook

from recon.build_xlsx import _Evaluator, build_provider_workbook, verify_workbook
from recon.checks import (group_by_provider, is_provider_batch,
                          run_provider_batches, validate_batch,
                          validate_provider_batches)
from recon.extract import extract_activity_table, parse_sra_supplier, parse_sra_text
from recon.identify import identify
from recon.models import IdentifiedFile, ReportType

import synth

UNITS = [
    # (code, cheque, OS, NM, AP, adjustment)
    ("F1070", "266444", 7_585.84, 0.0, 0.0, 148.36),
    ("F1088", "266474", 510.55, 0.0, 0.0, 8.45),
    ("F1089", "266458", 42_358.08, 71_252.16, 30_917.80, 896.21),
    ("F1090", "266457", 12_441.95, 4_362.21, 39_130.39, 0.0),
    ("F1097", "266475", 0.0, 0.0, 1_394.20, 0.0),
]


def _sra_file(code, cheque, os_amt, nm, ap, adj):
    text = synth.provider_sra_text(code=code, cheque=cheque, os_amt=os_amt,
                                   nm_amt=nm, ap_amt=ap, adj=adj)
    f = IdentifiedFile(f"{code}.pdf", b"", report_type=ReportType.SRA,
                       year=2026, month=5)
    f.raw_text = text
    f.provider_code, f.provider_label, cq = parse_sra_supplier(text)
    f.cheques = [cq]
    return f


def _batch():
    files = []
    for code, cheque, os_amt, nm, ap, adj in UNITS:
        files.append(_sra_file(code, cheque, os_amt, nm, ap, adj))
        files.append(identify("claims.xlsx", synth.provider_claims_xlsx(
            cheque=cheque, segments={"Outpatient Specialists": os_amt,
                                     "Nurses-Midwives": nm,
                                     "Allied Health": ap})))
        files.append(identify("export.xlsx", synth.activity_table_xlsx(
            provider=code, name=f"ΜΟΝΑΔΑ {code} (ΟΚΥπΥ)",
            rows=[("A1", cheque, round(os_amt + nm + ap, 2))])))
    return files


def test_sra_supplier_read_from_a_wrapped_header():
    """The payee wraps: «...INCOME-MENTAL / HEALTH-F1089 Payment Currency»."""
    code, _label, cheque = parse_sra_supplier(synth.provider_sra_text())
    assert code == "F1089" and cheque == "266458"


def test_activity_export_as_a_spreadsheet_matches_the_xml_path():
    """ΟΑΥ ships these units' activity export as .xlsx inside an «XMLS»
    folder — same normalized output as the real XML."""
    x = extract_activity_table(synth.activity_table_xlsx())
    assert x.total == 146_020.77
    assert x.by_payment == {"266458": 144_528.04, "263370": 1_492.73}
    assert x.n_claims == 3
    assert x.date_from == "2026-05-01" and x.date_to == "2026-06-01"
    f = identify("whatever.xlsx", synth.activity_table_xlsx())
    assert f.report_type == ReportType.XML_ACTIVITY
    assert f.provider_code == "F1089" and f.provider_label.startswith("ΚΟΙΝΟΤΙΚΑ")


def test_hospital_batch_is_not_treated_as_a_provider_batch():
    from test_checks import _identified_batch
    assert not is_provider_batch(_identified_batch())


def test_files_are_grouped_by_provider_from_content_only():
    files = _batch()
    assert is_provider_batch(files)
    batches, leftovers = group_by_provider(files)
    assert leftovers == []
    assert [b.code for b in batches] == ["F1070", "F1090", "F1089", "F1088", "F1097"]
    for b in batches:
        # SRA + claims + activity, attributed by F-code / PAYMENT NO. only
        assert sorted(f.report_type.name for f in b.files) == [
            "CLAIMS_ALL", "SRA", "XML_ACTIVITY"]


def test_claims_file_follows_its_cheque_not_its_filename():
    """Every unit's claims file is called the same thing in ΟΑΥ's folders —
    only PAYMENT NO. says which provider it belongs to."""
    files = _batch()
    for f in files:
        f.filename = "PAID CLAIMS.xlsx"       # identical names everywhere
    batches, leftovers = group_by_provider(files)
    assert leftovers == []
    for b in batches:
        claims = next(f for f in b.files if f.report_type == ReportType.CLAIMS_ALL)
        assert claims.cheques == b.cheques


def test_gates_pass_and_every_provider_reconciles():
    batches, leftovers = group_by_provider(_batch())
    gates, period, _notes = validate_provider_batches(batches, leftovers)
    assert all(g.passed for g in gates), [g.message for g in gates if not g.passed]
    assert period == (2026, 5)
    entries = run_provider_batches(batches, period)
    assert len(entries) == 5
    for (code, _label, res), unit in zip(entries, sorted(UNITS, key=lambda u: u[1])):
        expected = round(sum(unit[2:]), 2)
        assert res.bundle.sra.stated_total == expected
        assert round(sum(res.buckets.values()), 2) == expected


def test_missing_claims_file_for_one_provider_stops_with_its_name():
    files = [f for f in _batch()
             if not (f.report_type == ReportType.CLAIMS_ALL and f.cheques == ["266475"])]
    batches, leftovers = group_by_provider(files)
    gates, _period, _notes = validate_provider_batches(batches, leftovers)
    bad = [g for g in gates if not g.passed]
    assert bad and bad[0].number == 3
    assert "F1097" in bad[0].message


def test_provider_workbook_ties_every_cheque_and_verifies():
    batches, leftovers = group_by_provider(_batch())
    _gates, period, _notes = validate_provider_batches(batches, leftovers)
    entries = run_provider_batches(batches, period)
    data = build_provider_workbook(entries)
    assert verify_workbook(data) == []          # gate 5 over every provider
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames[0] == "Σύνοψη_παρόχων"
    assert [s for s in wb.sheetnames if s.startswith("SRA_")] == [
        "SRA_266444", "SRA_266457", "SRA_266458", "SRA_266474", "SRA_266475"]
    assert wb.sheetnames[-5:] == ["Ανάλυση_ελέγχων", "Ανά_μονάδα_ιατρό",
                                  "Ανά_κλινική", "JOURNAL ENTRIES", "Legend"]
    ws = wb["Σύνοψη_παρόχων"]
    ev = _Evaluator(wb)
    total_row = next(r for r in range(1, ws.max_row + 1)
                     if str(ws.cell(row=r, column=1).value or "").startswith("ΣΥΝΟΛΟ"))
    grand = round(ev.evaluate(ws.cell(row=total_row, column=8).value, "Σύνοψη_παρόχων"), 2)
    assert grand == round(sum(round(sum(u[2:]), 2) for u in UNITS), 2)
    # the per-stream split is LIVE off each provider's own SRA tab
    os_total = round(ev.evaluate(ws.cell(row=total_row, column=4).value, "Σύνοψη_παρόχων"), 2)
    assert os_total == round(sum(u[2] for u in UNITS), 2)


def test_crosscheck_and_audit_tabs_are_sectioned_per_provider():
    batches, leftovers = group_by_provider(_batch())
    _gates, period, _notes = validate_provider_batches(batches, leftovers)
    data = build_provider_workbook(run_provider_batches(batches, period))
    wb = load_workbook(io.BytesIO(data))
    labels = [str(ws.cell(row=r, column=1).value or "")
              for ws in (wb["Source_crosscheck"],) for r in range(1, ws.max_row + 1)]
    for code in ("F1070", "F1088", "F1089", "F1090", "F1097"):
        assert any(code in l for l in labels), code
    # each audit block is prefixed with its provider, and its tie-backs hold
    ws = wb["Ανάλυση_ελέγχων"]
    ev = _Evaluator(wb)
    ties = [r for r in range(1, ws.max_row + 1)
            if str(ws.cell(row=r, column=1).value or "").startswith("Έλεγχος: Σύνολο")]
    assert ties
    for r in ties:
        assert round(ev.evaluate(ws.cell(row=r, column=2).value, "Ανάλυση_ελέγχων"), 2) == 0.0


def test_by_unit_and_doctor_tab_splits_each_cheque_and_ties_to_it():
    """The posting sheet: every unit's cheque split by speciality and
    professional, bridged to the cheque in live formulas — the claims-vs-SRA
    gap is its OWN row, never spread across the professionals."""
    batches, leftovers = group_by_provider(_batch())
    _gates, period, _notes = validate_provider_batches(batches, leftovers)
    entries = run_provider_batches(batches, period)
    wb = load_workbook(io.BytesIO(build_provider_workbook(entries)))
    ws = wb["Ανά_μονάδα_ιατρό"]
    ev = _Evaluator(wb)
    labels = {r: str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)}
    # one block per unit, each ending in a zero-check that recomputes to 0
    checks = [r for r, l in labels.items() if l.startswith("Zero-check")]
    assert len(checks) == len(entries)
    for r in checks:
        assert round(ev.evaluate(ws.cell(row=r, column=5).value, "Ανά_μονάδα_ιατρό"), 2) == 0.0
    # professionals are listed under their speciality, with live subtotals
    specs = [r for r in range(1, ws.max_row + 1)
             if str(ws.cell(row=r, column=3).value or "").startswith("Υποσύνολο —")]
    assert specs
    for r in specs:
        assert str(ws.cell(row=r, column=5).value).startswith("=SUM(")
    docs = {str(ws.cell(row=r, column=4).value) for r in range(1, ws.max_row + 1)
            if ws.cell(row=r, column=4).value}
    assert any("SKORDI" in d or "ΣΚΟΡΔΗ" in d for d in docs)
    # the two grand totals: allocated by professional, and the cheques
    cheques_row = next(r for r, l in labels.items() if l.startswith("ΓΕΝΙΚΟ ΣΥΝΟΛΟ — επιταγές"))
    grand = round(ev.evaluate(ws.cell(row=cheques_row, column=5).value, "Ανά_μονάδα_ιατρό"), 2)
    assert grand == round(sum(round(sum(u[2:]), 2) for u in UNITS), 2)
