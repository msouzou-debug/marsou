"""OKYπY's own SAP master data driving the journal: company code per hospital,
the HIO revenue account per stream, and the clinic's cost centre picked by
flavour — ward for DRG, ημερήσια φροντίδα for daily treatments, εξωτερικά
ιατρεία for the outpatient specialists."""
import io

from openpyxl import Workbook, load_workbook

import synth
from recon.build_xlsx import build_workbook, verify_workbook
from recon.checks import run_reconciliation
from recon.extract import extract_inpatient_summary
from recon.identify import identify
from recon.models import ReportType
from recon.sapmaster import (SapMaster, company_for, extract_sap_master,
                             looks_like_sap_master)
from test_workbook import _build


def master_xlsx() -> bytes:
    """The shape SAP exports: company codes, the cost-centre master and the
    chart of accounts, one sheet each."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Company Codes"
    ws.append(["Comp. Code", "Περιγραφή"])
    for code, name in (("1003", "ΔΥΨΥ"), ("1040", "ΓΝ Λάρνακας"),
                       ("1041", "ΓΝ Αμμοχώστου")):
        ws.append([code, name])
    cc = wb.create_sheet("Cost centers")
    cc.append(["Company Code", "Cost Center", "Name"])
    for company, code, name in (
            ("1041", "1064102205", "ΚΑΡΔΙΟΛΟΓΙΚΗ-ΘΑΛ Α"),
            ("1041", "1064102203", "ΚΑΡΔΙΟΛΟΓΙΚΗ-ΗΦ"),
            ("1041", "1064102202", "ΚΑΡΔΙΟΛΟΓΙΚΗ-ΕΙ"),
            ("1041", "1064102200", "ΚΑΡΔΙΟΛΟΓΙΚΗ-ΓΕΝΙΚΑ"),
            ("1041", "1064105001", "ΤΑΕΠ"),
            ("1041", "1064105002", "ΚΩΔΙΚΟΠΟΙΗΣΗ ΤΑΕΠ"),
            ("1041", "1064105600", "ΦΑΡΜΑΚΕΙΟ"),
            # another hospital's centres must never be picked for F1049
            ("1040", "1064002205", "ΚΑΡΔΙΟΛΟΓΙΚΗ-ΘΑΛ Α")):
        cc.append([company, code, name])
    coa = wb.create_sheet("Chart of accounts")
    coa.append(["G/L Account", "G/L Acct Long Text"])
    for code, text in (("412000", "HIO - Capitation Fees"),
                       ("412001", "HIO In-Patient Fees"),
                       ("412002", "HIO Out-Patient Fees"),
                       ("412003", "HIO TAEP Fees"),
                       ("412005", "HIO Day Care Fees"),
                       ("412006", "HIO Drugs Phase B"),
                       ("412007", "HIO Catalogue Z Items")):
        coa.append([code, text])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_master_is_identified_by_content_not_by_name():
    f = identify("anything.xlsx", master_xlsx())
    assert f.report_type == ReportType.SAP_MASTER
    assert looks_like_sap_master(["Cost centers", "Chart of accounts"])
    m = extract_sap_master(master_xlsx())
    assert m.companies["1041"] == "ΓΝ Αμμοχώστου"
    assert len(m.cost_centres) == 8 and len(m.accounts) == 7


def test_a_stream_picks_its_own_account_and_flavour_of_cost_centre():
    m = extract_sap_master(master_xlsx())
    assert company_for("F1049") == "1041"
    assert company_for("F1070") == "1003"          # mental health
    assert m.account("inpatient_drg") == ("412001", "HIO In-Patient Fees")
    assert m.account("inpatient_z") == ("412007", "HIO Catalogue Z Items")
    ward = m.find_centre("1041", "CARDIOLOGY", "ward")
    day = m.find_centre("1041", "CARDIOLOGY", "daycare")
    clinic = m.find_centre("1041", "CARDIOLOGY", "clinic")
    assert (ward.code, day.code, clinic.code) == \
        ("1064102205", "1064102203", "1064102202")
    # «ΤΑΕΠ» is that stream's own centre, not «ΚΩΔΙΚΟΠΟΙΗΣΗ ΤΑΕΠ»
    assert m.find_centre("1041", "A&E").code == "1064105001"
    # a speciality the dictionary does not cover is NOT guessed
    assert m.find_centre("1041", "SOMETHING ELSE") is None
    # and one hospital's centre never answers for another
    assert m.find_centre("1040", "A&E") is None


def test_an_account_the_chart_does_not_carry_is_not_written():
    thin = SapMaster(accounts={"412001": "HIO In-Patient Fees"})
    assert thin.account("inpatient_drg") == ("412001", "HIO In-Patient Fees")
    assert thin.account("inpatient_z") == ("", "")


def test_the_journal_posts_each_stream_to_its_own_account_and_centre():
    _data, res = _build(with_optional=True)
    res.bundle.inpatient = extract_inpatient_summary(
        synth.inpatient_summary_xlsx(with_procedure_detail=True))
    res.bundle.sap = extract_sap_master(master_xlsx())
    res = run_reconciliation(res.bundle)
    data = build_workbook(res)
    assert verify_workbook(data) == []
    ws = load_workbook(io.BytesIO(data))["JOURNAL ENTRIES"]
    credits = [r for r in range(4, ws.max_row + 1)
               if ws.cell(row=r, column=9).value == "50"
               and isinstance(ws.cell(row=r, column=12).value, (int, float))]
    # the hospital's own company code, not the mental-health default
    assert ws.cell(row=4, column=4).value == "1041"
    # one inpatient clinic row became three lines, one per revenue account
    by_account: dict[str, float] = {}
    for r in credits:
        acct = str(ws.cell(row=r, column=10).value)
        by_account[acct] = round(by_account.get(acct, 0.0)
                                 + ws.cell(row=r, column=12).value, 2)
    assert by_account["412001"] == 700_000.00       # DRG
    assert by_account["412005"] == 336_000.00       # daily treatments
    assert by_account["412007"] == 25_728.70        # catalogue Z
    assert by_account["412003"] == 131_284.66       # ΤΑΕΠ
    # and the whole document still ties to the cheque
    assert round(sum(ws.cell(row=r, column=12).value for r in credits), 2) == \
        1_936_528.19
    cardio = [r for r in credits
              if ws.cell(row=r, column=14).value == "1064102205"]
    assert cardio and ws.cell(row=cardio[0], column=10).value == "412001"


def test_without_the_master_the_journal_falls_back_and_invents_nothing():
    _data, res = _build(with_optional=True)
    ws = load_workbook(io.BytesIO(build_workbook(res)))["JOURNAL ENTRIES"]
    credits = [r for r in range(4, ws.max_row + 1)
               if ws.cell(row=r, column=9).value == "50"
               and isinstance(ws.cell(row=r, column=12).value, (int, float))]
    assert credits
    assert all(ws.cell(row=r, column=10).value == "412002" for r in credits)
    assert all(not ws.cell(row=r, column=14).value for r in credits)


def test_a_hospital_posts_no_internal_order():
    """The internal order (11-16) is the mental-health professional category.
    A hospital has none, so that column stays empty even when a lookup offers
    one."""
    from recon.mapping import extract_cost_centres
    _data, res = _build(with_optional=True)
    res.bundle.sap = extract_sap_master(master_xlsx())
    res.bundle.cost_centres = extract_cost_centres(synth.cost_centre_map_xlsx(
        rows=[("Inpatient", "26001", "77", "INPATIENT", "", "F1049")],
        with_hospital=True))
    ws = load_workbook(io.BytesIO(build_workbook(res)))["JOURNAL ENTRIES"]
    credits = [r for r in range(4, ws.max_row + 1)
               if ws.cell(row=r, column=9).value == "50"
               and isinstance(ws.cell(row=r, column=12).value, (int, float))]
    assert credits
    assert all(not ws.cell(row=r, column=15).value for r in credits)


def test_nurses_and_allied_health_post_to_the_outpatient_clinics():
    """ΟΑΥ pays both segments as one number with no speciality of their own,
    so both post to «ΕΞ.ΙΑΤΡΕΙΑ-ΓΕΝΙΚΑ» — and a hospital whose master has no
    such centre still gets a blank rather than a guess."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost centers"
    ws.append(["Company Code", "Cost Center", "Name"])
    ws.append(["1041", "1064110701", "ΕΞ.ΙΑΤΡΕΙΑ-ΓΕΝΙΚΑ"])
    ws.append(["1033", "1053305600", "ΦΑΡΜΑΚΕΙΟ"])          # Polis has none
    coa = wb.create_sheet("Chart of accounts")
    coa.append(["G/L Account", "G/L Acct Long Text"])
    coa.append(["412002", "HIO Out-Patient Fees"])
    buf = io.BytesIO()
    wb.save(buf)
    m = extract_sap_master(buf.getvalue())
    for label in ("Νοσηλευτές/Μαίες (Nurses-Midwives)",
                  "Άλλοι Επαγγελματίες Υγείας (Allied Health)"):
        assert m.find_centre("1041", label).code == "1064110701"
        assert m.find_centre("1033", label) is None


def _centres(*rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cost centers"
    ws.append(["Company Code", "Cost Center", "Name"])
    for r in rows:
        ws.append(list(r))
    coa = wb.create_sheet("Chart of accounts")
    coa.append(["G/L Account", "G/L Acct Long Text"])
    coa.append(["412002", "HIO Out-Patient Fees"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_the_flavour_is_read_after_the_speciality_not_inside_it():
    """«ΟΦΘΑΛ» contains «ΘΑΛ» and «ΧΕΙΡΟΥΡΓΙΚΗ» contains «ΕΙ». Testing the whole
    name made every ophthalmology centre look like a ward and every surgery
    centre like an outpatient clinic, so the match was thrown out as ambiguous
    and the line came out uncoded — the real F1048 symptom."""
    m = extract_sap_master(_centres(
        ("1040", "1064003900", "ΟΦΘΑΛΜΟΛΟΓΙΚΗ-ΓΕΝΙΚΑ"),
        ("1040", "1064003901", "ΟΦΘΑΛΜΟΛΟΓΙΚΗ-ΕΙ"),
        ("1040", "1064003902", "ΟΦΘΑΛΜΟΛΟΓΙΚΗ-ΘΑΛ"),
        ("1040", "1064003903", "ΟΦΘΑΛΜΟΛΟΓΙΚΗ Η.Φ."),
        ("1040", "1064000601", "ΧΕΙΡΟΥΡΓΙΚΗ-ΕΙ"),
        ("1040", "1064000602", "ΧΕΙΡΟΥΡΓΙΚΗ-ΘΑΛ Α"),
        ("1040", "1064000603", "ΧΕΙΡΟΥΡΓΙΚΗ Η.Φ.")))
    got = {v: m.find_centre("1040", "OPHTHALMOLOGY", v)
           for v in ("ward", "daycare", "clinic", "general")}
    assert {v: c.code for v, c in got.items()} == {
        "ward": "1064003902", "daycare": "1064003903",
        "clinic": "1064003901", "general": "1064003900"}
    assert m.find_centre("1040", "GENERAL SURGERY", "clinic").code == "1064000601"
    assert m.find_centre("1040", "GENERAL SURGERY", "ward").code == "1064000602"


def test_a_hyphenated_speciality_still_finds_its_stem():
    """norm_label turns «DERMATO-VENEREOLOGY» into «DERMATO VENEREOLOGY», so a
    hyphenated dictionary key never matched the speciality it was written for."""
    m = extract_sap_master(_centres(
        ("1040", "1064001001", "ΔΕΡΜΑΤΟΛΟΓΙΚΗ-ΕΙ"),
        ("1040", "1064001002", "ΔΕΡΜΑΤΟΛΟΓΙΚΗ-ΘΑΛ")))
    assert m.find_centre("1040", "DERMATO-VENEREOLOGY", "clinic").code == "1064001001"
    # and the label as By_Clinic_Split writes it
    assert m.find_centre("1040", "Ειδικοί Ιατροί — DERMATO-VENEREOLOGY (OS)",
                         "clinic").code == "1064001001"


def test_nicosia_books_the_alpha_ward_when_a_clinic_has_two():
    m = extract_sap_master(_centres(
        ("1020", "1042002200", "ΚΑΡΔΙΟΛΟΓΙΚΗ-ΓΕΝΙΚΑ"),
        ("1020", "1042002205", "ΚΑΡΔΙΟΛΟΓΙΚΗ-ΘΑΛ Α"),
        ("1020", "1042002206", "ΚΑΡΔΙΟΛΟΓΙΚΗ-ΘΑΛ Β"),
        ("1020", "1042000602", "ΧΕΙΡΟΥΡΓΙΚΗ-ΘΑΛ Α"),
        ("1020", "1042000603", "ΧΕΙΡΟΥΡΓΙΚΗ-ΘΑΛ Β"),
        # not general surgery, and must not be picked as one
        ("1020", "1042006102", "ΝΕΥΡΟΧΕΙΡΟΥΡΓΙΚΗ-ΘΑΛ")))
    assert m.find_centre("1020", "CARDIOLOGY", "ward").code == "1042002205"
    assert m.find_centre("1020", "GENERAL SURGERY", "ward").code == "1042000602"
    assert m.find_centre("1020", "NEUROLOGICAL SURGERY", "ward").code == "1042006102"


def test_the_alert_names_only_lines_that_carry_money():
    """«Alert and ignore if they don't have amounts allocated to them»: a line
    with no cost centre AND no amount is not a finding."""
    from recon.build_xlsx import _missing_note
    info = {"missing": {"RENAL DISEASES": 12_345.67, "PLASTIC SURGERY": 0.0,
                        "UROLOGY": -250.00}, "master_seen": True}
    note = _missing_note(info, True)
    assert "RENAL DISEASES" in note and "UROLOGY" in note
    assert "PLASTIC SURGERY" not in note
    assert note.index("RENAL") < note.index("UROLOGY")     # biggest first
    assert _missing_note({"missing": {"X": 0.0}}, True) is None
    # and when the master was never uploaded, say so first
    assert "Chart_of_Accounts" in _missing_note(info, False)
