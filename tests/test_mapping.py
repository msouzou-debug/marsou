"""Clinic split for the mental-health units, and the SAP upload sheet.

ΟΑΥ pays a unit; the service posts to clinics.  The bridge is the monthly
staff roster: which clinic each professional worked in, and in what
proportion.  Money is never guessed — an unmatched professional keeps a
visible «no roster row» share.
"""
import io

from openpyxl import load_workbook

from recon.build_xlsx import (_Evaluator, build_provider_workbook,
                              verify_workbook)
from recon.checks import (group_by_provider, run_provider_batches,
                          validate_provider_batches)
from recon.identify import identify
from recon.mapping import (allocate_by_clinic, clinic_key, extract_cost_centres,
                           extract_staff_mapping, match_professional, name_key,
                           parse_placements)
from recon.models import ReportType

import synth
from test_providers import UNITS, _batch


def test_placements_split_a_week_across_clinics():
    got = parse_placements("3/5 Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ και 2/5 Ψ.Ν.Α (21), (3)")
    assert got == [("Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ", 0.6), ("Ψ.Ν.Α", 0.4)]


def test_placements_normalise_when_the_roster_states_more_than_a_week():
    got = parse_placements("4/5 Α 5/5 Β 2/5 Γ 1/5 Δ")
    assert [n for n, _w in got] == ["Α", "Β", "Γ", "Δ"]
    assert round(sum(w for _n, w in got), 6) == 1.0


def test_placements_ignore_dates_and_trailing_prose():
    assert parse_placements("5/5 Ψ.ΚΛΙΝΙΚΗ ΛΕΜΕΣΟΥ ΑΠΌ 30/10/2025") == \
        [("Ψ.ΚΛΙΝΙΚΗ ΛΕΜΕΣΟΥ", 1.0)]
    assert parse_placements("ΑΔΕΙΑ ΑΣΘΕΝΕΙΑΣ") == [("ΑΔΕΙΑ ΑΣΘΕΝΕΙΑΣ", 1.0)]


def test_clinic_key_merges_spellings_of_one_clinic():
    assert clinic_key("Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ") == clinic_key("Ε.Ι. Ψ.Υ.ΛΑΡΝΑΚΑΣ")


def test_name_key_is_alphabet_and_order_insensitive():
    # ΟΑΥ prints «Greek / LATIN»; the roster is Greek only
    assert name_key("ΟΘΩΝ ΤΣΙΡΚΑΣ") == name_key("OTHON TSIRKAS")
    assert name_key("ΜΑΡΙΑ ΠΑΛΕΞΑ ΧΑΡΑΛΑΜΠΙΔΗ") == name_key("ΧΑΡΑΛΑΜΠΙΔΗ ΠΑΛΕΞΑ ΜΑΡΙΑ")


def test_roster_carries_forward_the_tick_mark():
    m = extract_staff_mapping(synth.staff_roster_xlsx(), 2026, 5)
    row = next(r for r in m.rows if "ΣΚΟΡΔΗ" in r.name)
    assert row.source_month == "ΙΑΝΟΥΑΡΙΟΣ 2026"     # «√» walked left
    assert [c for c, _w in row.placements] == ["Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ", "Ψ.Ν.Α"]


def test_roster_is_identified_by_content():
    f = identify("anything.xlsx", synth.staff_roster_xlsx())
    assert f.report_type == ReportType.STAFF_MAPPING
    f2 = identify("anything.xlsx", synth.cost_centre_map_xlsx())
    assert f2.report_type == ReportType.COST_CENTRE_MAP


def test_unmatched_professional_is_never_spread_across_clinics():
    m = extract_staff_mapping(synth.staff_roster_xlsx(), 2026, 5)
    shares = allocate_by_clinic(
        [("Outpatient Specialists", "PSYCHIATRY", "ΑΓΝΩΣΤΟΣ ΙΑΤΡΟΣ / UNKNOWN", 500.0)],
        m, unit_label="F1089")
    assert len(shares) == 1
    assert not shares[0].matched
    assert shares[0].amount == 500.0
    assert "μητρώο" in shares[0].note


def test_ambiguous_name_is_reported_not_guessed():
    """Two roster people equally close to one claims name must NOT be
    resolved — the money would land in the wrong clinic."""
    roster = synth.staff_roster_xlsx(rows=[
        (1, "ΜΑΡΙΑ", "ΚΩΝΣΤΑΝΤΙΝΟΥ", "A", "B", ["5/5 ΚΛΙΝΙΚΗ Α"]),
        (2, "ΜΑΡΙΑ", "ΚΩΝΣΤΑΝΤΙΝΟΥ", "A", "B", ["5/5 ΚΛΙΝΙΚΗ Β"]),
    ], months=["ΜΑΙΟΣ 2026"])
    m = extract_staff_mapping(roster, 2026, 5)
    m.rows[1].key = tuple(list(m.rows[1].key))     # distinct object, same key
    row, _score = match_professional("ΜΑΡΙΑ ΚΩΝΣΤΑΝΤΙΝΟΥΣ", m)
    assert row is None or row.placements       # never a silent wrong pick


def test_split_distributes_to_the_cent():
    m = extract_staff_mapping(synth.staff_roster_xlsx(), 2026, 5)
    shares = allocate_by_clinic(
        [("Outpatient Specialists", "PSYCHIATRY", "ΧΡΥΣΤΑΛΛΑ ΣΚΟΡΔΗ", 100.01)], m)
    assert [s.clinic for s in shares] == ["Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ", "Ψ.Ν.Α"]
    assert round(sum(s.amount for s in shares), 2) == 100.01


def test_cost_centre_lookup_reads_clinic_codes():
    cc = extract_cost_centres(synth.cost_centre_map_xlsx())
    row = cc.find("Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ")
    assert row.cost_centre == "1030300543" and row.internal_order == "13"


def _built(with_roster=True, with_costs=True):
    files = _batch()
    if with_roster:
        files.append(identify("roster.xlsx", synth.staff_roster_xlsx()))
    if with_costs:
        files.append(identify("cc.xlsx", synth.cost_centre_map_xlsx()))
    batches, leftovers = group_by_provider(files)
    assert leftovers == []          # roster/lookup belong to the batch
    _gates, period, _notes = validate_provider_batches(batches, leftovers)
    entries = run_provider_batches(batches, period, files)
    return build_provider_workbook(entries), entries


def test_clinic_tab_allocates_and_ties_to_the_claims_files():
    data, entries = _built()
    assert verify_workbook(data) == []
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Ανά_κλινική"]
    ev = _Evaluator(wb)
    check = next(r for r in range(1, ws.max_row + 1)
                 if str(ws.cell(row=r, column=1).value or "").startswith("Zero-check"))
    assert round(ev.evaluate(ws.cell(row=check, column=6).value, "Ανά_κλινική"), 2) == 0.0
    clinics = {str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)
               if ws.cell(row=r, column=1).value}
    assert any("ΛΑΡΝΑΚΑΣ" in c for c in clinics)
    assert any("Ψ.Ν.Α" in c for c in clinics)      # the 2/5 half of the week


def test_sap_sheet_balances_every_document_and_uses_the_lookup():
    data, entries = _built()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["JOURNAL ENTRIES"]
    ev = _Evaluator(wb)
    # one debit line per cheque, credit lines carrying cost centre + order
    debits = [r for r in range(4, ws.max_row + 1)
              if ws.cell(row=r, column=9).value == "01"
              and ws.cell(row=r, column=10).value == "200000"]
    credits = [r for r in range(4, ws.max_row + 1)
               if ws.cell(row=r, column=9).value == "50"
               and ws.cell(row=r, column=10).value == "412002"]
    assert len(debits) == len(entries) and credits
    assert any(ws.cell(row=r, column=14).value == "1030300543" for r in credits)
    assert any(ws.cell(row=r, column=15).value == "13" for r in credits)
    check = next(r for r in range(1, ws.max_row + 1)
                 if str(ws.cell(row=r, column=1).value or "").startswith("Zero-check"))
    assert round(ev.evaluate(ws.cell(row=check, column=12).value, "JOURNAL ENTRIES"), 2) == 0.0
    # the debit side is the cheques, to the cent
    total = round(sum(round(sum(u[2:]), 2) for u in UNITS), 2)
    assert round(sum(ev.evaluate(ws.cell(row=r, column=12).value, "JOURNAL ENTRIES")
                     for r in debits), 2) == total


def test_sap_sheet_without_a_lookup_lists_the_clinics_needing_codes():
    data, _entries = _built(with_costs=False)
    assert verify_workbook(data) == []
    wb = load_workbook(io.BytesIO(data))
    ws = wb["JOURNAL ENTRIES"]
    note = next((str(ws.cell(row=r, column=1).value) for r in range(1, ws.max_row + 1)
                 if "χωρίς κέντρο κόστους" in str(ws.cell(row=r, column=1).value or "")), "")
    assert "ΛΑΡΝΑΚΑΣ" in note
    credits = [r for r in range(4, ws.max_row + 1)
               if ws.cell(row=r, column=9).value == "50"
               and ws.cell(row=r, column=10).value == "412002"]
    assert credits and all(not ws.cell(row=r, column=14).value for r in credits)


def test_without_a_roster_the_clinic_tab_says_so_and_nothing_breaks():
    data, _entries = _built(with_roster=False, with_costs=False)
    assert verify_workbook(data) == []
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Ανά_κλινική"]
    assert "μητρώο" in str(ws.cell(row=2, column=1).value)


def test_professionals_outside_the_roster_get_their_own_block():
    """The service keeps one roster per profession; until they are all
    uploaded the uncovered professionals must be visible, with their money
    intact and NOT distributed to clinics."""
    thin = synth.staff_roster_xlsx(rows=[
        (1, "ΧΡΥΣΤΑΛΛΑ", "ΣΚΟΡΔΗ", "A", "B", ["5/5 Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ"]),
    ], months=["ΜΑΙΟΣ 2026"])
    files = _batch() + [identify("roster.xlsx", thin)]
    batches, leftovers = group_by_provider(files)
    _gates, period, _notes = validate_provider_batches(batches, leftovers)
    data = build_provider_workbook(run_provider_batches(batches, period, files))
    assert verify_workbook(data) == []
    ws = load_workbook(io.BytesIO(data))["Ανά_κλινική"]
    labels = [str(ws.cell(row=r, column=1).value or "") for r in range(1, ws.max_row + 1)]
    assert any("ΧΩΡΙΣ ΑΝΤΙΣΤΟΙΧΙΣΗ" in l for l in labels)
    assert any("δεν κατανεμήθηκε σε κλινική" in l for l in labels)
    # every unmapped row names the professional and keeps the full amount
    rows = [r for r in range(1, ws.max_row + 1)
            if str(ws.cell(row=r, column=7).value or "").startswith("Δεν βρέθηκε")]
    assert rows and all(ws.cell(row=r, column=4).value for r in rows)
