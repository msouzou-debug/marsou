"""Δοκιμές εγγραφής σε συνθετικό workbook που μιμείται τη δομή του πραγματικού
(φύλλα ΣΥΝΟΠΤΙΚΟ + Ανάλυση, μπλοκ ανά μήνα με κενή γραμμή-διαχωριστικό)."""

from datetime import datetime

import pytest
from openpyxl import Workbook, load_workbook

from daspani.excel_write import (
    ExcelWriteError,
    append_month_block,
    build_history,
    delete_month_block,
    last_data_row,
    month_block,
    write_workbook,
)
from daspani.models import TableRow, hospital_from_remarks

HEADERS = ["ΜΗΝΑΣ", "Α/Α", "ΑΚΑ", "ΑΔΤ", "Ε/ΑΣΤ", "ΟΝΟΜΑΤΕΠΩΝΥΜΟ",
           "ΗΜΕΡΟΜΗΝΙΑ ΤΟΠΟΘΕΤΗΣΗΣ", "ΝΟΣΗΛΕΥΤΗΡΙΟ", "ΒΑΣΙΚΟΣ ΜΙΣΘΟΣ",
           "ΤΙΜΑΡΙΘΜΟΣ\n12,67%", "ΑΥΞΗΣΗ 10/2024", "ΑΥΞΗΣΗ 11/2024",
           "ΑΥΞΗΣΗ 12/2024", "ΑΥΞΗΣΗ 13ΟΥ", "ΑΥΞΗΣΗ ΜΙΣΘΟΥ", "ΕΠΙΔΟΜΑ ΒΑΡΔΙΑΣ",
           "ΕΠΙΔΟΜ ΚΥΡΙΑΚΗΣ", "ΣΥΝΟΛΟ", "ΕΙΣΦΟΡΑ", "ΚΟΣΤΟΣ", "ΔΙΟΙΚΗΤΙΚΑ", "ΣΥΝΟΛΙΚΟ"]


def build_workbook(with_january=True):
    wb = Workbook()
    syn = wb.active
    syn.title = "ΣΥΝΟΠΤΙΚΟ"
    syn["D4"] = '=SUMIFS(Ανάλυση!$V:$V,Ανάλυση!$H:$H,ΣΥΝΟΠΤΙΚΟ!A4,Ανάλυση!$A:$A,ΣΥΝΟΠΤΙΚΟ!D3)'
    ws = wb.create_sheet("Ανάλυση")
    for c, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=c).value = h
    if with_january:
        data = [
            ("ΙΑΝΟΥΑΡΙΟΣ", 1, 426310, 644847, 5435, "ΓΕΩΡΓΙΑΔΗΣ ΜΙΧΑΛΗΣ", 1200),
            ("ΙΑΝΟΥΑΡΙΟΣ", 2, 447126, 674259, 5344, "ΣΤΥΛΙΑΝΟΥ ΑΥΓΗ", 1400),
        ]
        for i, (m, aa, aka, adt, east, name, nosil) in enumerate(data):
            r = 2 + i
            ws.cell(row=r, column=1).value = m
            ws.cell(row=r, column=2).value = aa
            ws.cell(row=r, column=3).value = aka
            ws.cell(row=r, column=4).value = adt
            ws.cell(row=r, column=5).value = east
            ws.cell(row=r, column=6).value = name
            ws.cell(row=r, column=7).value = datetime(2022, 10, 12)
            ws.cell(row=r, column=8).value = nosil
            ws.cell(row=r, column=9).value = 2170.33
            ws.cell(row=r, column=18).value = f"=SUM(I{r}:Q{r})"
            ws.cell(row=r, column=19).value = f"=R{r}*24.25%"
            ws.cell(row=r, column=20).value = f"=R{r}+S{r}"
            ws.cell(row=r, column=21).value = f"=T{r}*10%"
            ws.cell(row=r, column=22).value = f"=T{r}+U{r}"
    return wb


def make_row(aa, aka, name, remarks="ΛΕΥΚΩΣΙΑ-Ε.Ο.Φ. ΜΑΚΑΡΙΟ ΝΟΣΟΚΟΜΕΙΟ", **amounts):
    defaults = dict(basic=2264.67, tim=291.24, auxisi=33.97, bardia=171.58, kyriaki=206.52)
    defaults.update(amounts)
    return TableRow(aa=str(aa), aka=str(aka), east="5435", name=name,
                    date="12/10/2022", remarks=remarks, **defaults)


def test_hospital_mapping():
    assert hospital_from_remarks("ΛΕΥΚΩΣΙΑ-Ε.Ο.Φ. ΜΑΚΑΡΙΟ ΝΟΣΟΚΟΜΕΙΟ") == 1200
    assert hospital_from_remarks("ΛΕΥΚΩΣΙΑ-ΛΑΤΣΙΑ-ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ") == 1400
    assert hospital_from_remarks("ΑΜΜΟΧΩΣΤΟΣ - Κ.Ε.Μ. - ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ") == 3200
    assert hospital_from_remarks("ΛΑΡΝΑΚΑ-ΣΤΑΘΜΟΣ ΠΟΛΕΩΣ-ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ") == 4100
    assert hospital_from_remarks("ΛΕΜΕΣΟΣ-ΣΤΑΘΜΟΣ ΠΟΛΕΜΙΔΙΩΝ-ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ") == 5100
    assert hospital_from_remarks("ΠΑΦΟΣ-ΚΕΝΤΡΙΚΟΣ ΣΤΑΘΜΟΣ-ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ") == 6100
    assert hospital_from_remarks("ΜΙΣΘΟΣ 3/2026") is None


def test_append_month_after_existing_with_separator():
    wb = build_workbook()
    ws = wb["Ανάλυση"]
    rows = [make_row(1, 426310, "ΓΕΩΡΓΙΑΔΗΣ ΜΙΧΑΛΗΣ"),
            make_row(2, 447126, "ΣΤΥΛΙΑΝΟΥ ΑΥΓΗ", remarks="ΛΕΥΚΩΣΙΑ-ΛΑΤΣΙΑ-ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ")]
    result = append_month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ", rows)
    # ΙΑΝ: γραμμές 2-3, κενή η 4, ΦΕΒ: 5-6.
    assert (result.first_row, result.last_row) == (5, 6)
    assert ws.cell(row=4, column=1).value is None
    assert ws.cell(row=5, column=1).value == "ΦΕΒΡΟΥΑΡΙΟΣ"
    assert ws.cell(row=5, column=3).value == 426310
    assert ws.cell(row=5, column=9).value == 2264.67
    assert ws.cell(row=5, column=18).value == "=SUM(I5:Q5)"
    assert ws.cell(row=6, column=22).value == "=T6+U6"
    # Ημερομηνία ως datetime
    assert ws.cell(row=5, column=7).value == datetime(2022, 10, 12)


def test_adt_and_hospital_filled_from_history():
    wb = build_workbook()
    ws = wb["Ανάλυση"]
    row = make_row(1, 426310, "ΓΕΩΡΓΙΑΔΗΣ ΜΙΧΑΛΗΣ", remarks="")  # χωρίς παρατήρηση
    result = append_month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ", [row])
    r = result.first_row
    assert ws.cell(row=r, column=4).value == 644847   # ΑΔΤ από τον Ιανουάριο
    assert ws.cell(row=r, column=8).value == 1200     # νοσηλευτήριο από τον Ιανουάριο
    assert result.filled_from_history == 1


def test_new_hire_without_history_is_flagged():
    wb = build_workbook()
    ws = wb["Ανάλυση"]
    row = make_row(1, 999999, "ΝΕΟΣ ΥΠΑΛΛΗΛΟΣ", remarks="ΠΑΦΟΣ-ΚΕΝΤΡΙΚΟΣ ΣΤΑΘΜΟΣ")
    result = append_month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ", [row])
    r = result.first_row
    assert ws.cell(row=r, column=4).value is None       # χωρίς ΑΔΤ
    assert ws.cell(row=r, column=8).value == 6100       # από τις παρατηρήσεις
    assert any("ΑΔΤ" in w for w in result.warnings)


def test_apo_date_kept_as_text():
    wb = build_workbook()
    ws = wb["Ανάλυση"]
    row = make_row(1, 426310, "ΓΕΩΡΓΙΑΔΗΣ ΜΙΧΑΛΗΣ")
    row.date = "ΑΠΟ 23/3/2026"
    result = append_month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ", [row])
    assert ws.cell(row=result.first_row, column=7).value == "ΑΠΟ 23/3/2026"


def test_unnumbered_correction_row():
    wb = build_workbook()
    ws = wb["Ανάλυση"]
    row = make_row("", 426310, "ΓΕΩΡΓΙΑΔΗΣ ΜΙΧΑΛΗΣ", remarks="ΜΙΣΘΟΣ 3/2026",
                   basic=-262.54, tim=-33.76, auxisi=-3.94, bardia=0, kyriaki=0)
    result = append_month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ", [row])
    r = result.first_row
    assert ws.cell(row=r, column=2).value is None       # χωρίς Α/Α
    assert ws.cell(row=r, column=9).value == -262.54
    assert ws.cell(row=r, column=8).value == 1200       # νοσηλευτήριο από ιστορικό


def test_write_workbook_rejects_existing_month_without_replace(tmp_path):
    wb = build_workbook()
    path = tmp_path / "ΧΡΕΩΣΕΙΣ.xlsx"
    wb.save(path)
    with pytest.raises(ExcelWriteError, match="υπάρχει ήδη"):
        write_workbook(str(path), "ΙΑΝΟΥΑΡΙΟΣ", [make_row(1, 426310, "Χ")])


def test_replace_last_month_is_idempotent(tmp_path):
    wb = build_workbook()
    path = tmp_path / "ΧΡΕΩΣΕΙΣ.xlsx"
    wb.save(path)
    rows = [make_row(1, 426310, "ΓΕΩΡΓΙΑΔΗΣ ΜΙΧΑΛΗΣ"),
            make_row(2, 447126, "ΣΤΥΛΙΑΝΟΥ ΑΥΓΗ")]

    r1 = write_workbook(str(path), "ΦΕΒΡΟΥΑΡΙΟΣ", rows, save_path=str(path))
    assert r1.written == 2 and not r1.replaced_existing

    # Δεύτερο τρέξιμο με το ίδιο PDF: αντικατάσταση, όχι διπλοεγγραφή.
    r2 = write_workbook(str(path), "ΦΕΒΡΟΥΑΡΙΟΣ", rows, replace_existing=True, save_path=str(path))
    assert r2.replaced_existing
    ws = load_workbook(path)["Ανάλυση"]
    feb_rows = [r for r in range(2, ws.max_row + 1)
                if ws.cell(row=r, column=1).value == "ΦΕΒΡΟΥΑΡΙΟΣ"]
    assert len(feb_rows) == 2
    assert last_data_row(ws) == 6  # ΙΑΝ 2-3, κενή 4, ΦΕΒ 5-6 — ίδιο αποτύπωμα


def test_replace_middle_month_refreshes_formulas(tmp_path):
    wb = build_workbook()
    ws = wb["Ανάλυση"]
    append_month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ", [make_row(1, 426310, "Α"), make_row(2, 447126, "Β")])
    append_month_block(ws, "ΜΑΡΤΙΟΣ", [make_row(1, 426310, "Α")])
    path = tmp_path / "ΧΡΕΩΣΕΙΣ.xlsx"
    wb.save(path)

    # Αντικατάσταση του ΦΕΒΡΟΥΑΡΙΟΥ (ενδιάμεσος) με 3 γραμμές πλέον.
    rows = [make_row(1, 426310, "Α"), make_row(2, 447126, "Β"), make_row("", 426310, "Α", remarks="ΜΙΣΘΟΣ 1/2026")]
    result = write_workbook(str(path), "ΦΕΒΡΟΥΑΡΙΟΣ", rows, replace_existing=True, save_path=str(path))
    ws2 = load_workbook(path)["Ανάλυση"]
    # Ο ΜΑΡΤΙΟΣ ανέβηκε· οι φόρμουλές του δείχνουν στη σωστή (νέα) γραμμή του.
    mar = month_block(ws2, "ΜΑΡΤΙΟΣ")
    assert mar is not None
    r = mar[0]
    assert ws2.cell(row=r, column=18).value == f"=SUM(I{r}:Q{r})"
    # Ο ΦΕΒΡΟΥΑΡΙΟΣ ξαναγράφτηκε στο τέλος με 3 γραμμές.
    feb = month_block(ws2, "ΦΕΒΡΟΥΑΡΙΟΣ")
    assert feb[1] - feb[0] + 1 == 3
    assert feb[0] > mar[1]
    assert any("δεν ήταν ο τελευταίος" in w for w in result.warnings)


def test_delete_month_block_removes_separator():
    wb = build_workbook()
    ws = wb["Ανάλυση"]
    append_month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ", [make_row(1, 426310, "Α")])
    assert delete_month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ")
    assert month_block(ws, "ΦΕΒΡΟΥΑΡΙΟΣ") is None
    assert last_data_row(ws) == 3  # μόνο ο Ιανουάριος


def test_build_history_ignores_target_month():
    wb = build_workbook()
    ws = wb["Ανάλυση"]
    hist = build_history(ws, "ΙΑΝΟΥΑΡΙΟΣ")
    assert hist == {}  # ο ίδιος ο μήνας δεν μετρά ως ιστορικό
    hist = build_history(ws, "ΦΕΒΡΟΥΑΡΙΟΣ")
    assert hist["426310"]["adt"] == 644847
