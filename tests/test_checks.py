from recon.checks import (ReconBundle, conditional_requirements,
                          gate4_internal_asserts, run_reconciliation,
                          validate_batch)
from recon.extract import (extract_claims_all, extract_gl,
                           extract_inpatient_summary, extract_is_auditor,
                           extract_pharma_claims, extract_simple_report,
                           extract_xml_activity, parse_pharmacist_fee_text,
                           parse_sra_text)
from recon.identify import identify
from recon.models import Bucket, IdentifiedFile, ReportType

import synth


def _identified_batch():
    files = [
        identify("endo.xlsx", synth.inpatient_summary_xlsx()),
        identify("claims_OS.xlsx", synth.claims_all_xlsx()),
        identify("pharma.xlsx", synth.pharma_claims_xlsx()),
    ]
    sra = IdentifiedFile("sra.pdf", b"", report_type=ReportType.SRA,
                         hospital_code="F1049", year=2026, month=3)
    fee = IdentifiedFile("fee.pdf", b"", report_type=ReportType.PHARMACIST_FEE,
                         hospital_code="F1049", year=2026, month=3)
    return files + [sra, fee]


def full_bundle(with_optional=False):
    b = ReconBundle(hospital_code="F1049", year=2026, month=3)
    b.sra = parse_sra_text(synth.sra_text())
    b.inpatient = extract_inpatient_summary(synth.inpatient_summary_xlsx())
    b.claims = extract_claims_all(synth.claims_all_xlsx())
    b.pharma = extract_pharma_claims(synth.pharma_claims_xlsx())
    b.phfee = parse_pharmacist_fee_text(synth.pharmacist_fee_text())
    b.capitation = extract_simple_report(b"", raw_text=synth.capitation_text())
    if with_optional:
        b.gl = extract_gl(synth.gl_xlsx(), "F1049")
        b.isaud = extract_is_auditor(synth.is_auditor_xlsx(), "F1049")
        b.xml_activity = extract_xml_activity(synth.xml_activity_bytes())
    return b


def test_gates_pass_on_complete_batch():
    gates, hospital, period, notes = validate_batch(_identified_batch())
    assert all(g.passed for g in gates)
    assert hospital == "F1049"
    assert period == (2026, 3)


def test_gate2_rejects_mixed_hospitals():
    files = _identified_batch()
    files[3].hospital_code = "F1054"
    gates, _, _, _ = validate_batch(files)
    assert any(not g.passed and g.number == 2 for g in gates)
    assert "δύο νοσοκομεία" in [g for g in gates if not g.passed][0].message


def test_gate2_rejects_mixed_months():
    files = _identified_batch()
    files[0].month = 4          # a CLAIM report from another month — mixed batch
    gates, _, _, _ = validate_batch(files)
    assert any(not g.passed and g.number == 2 for g in gates)


def test_gate2_sra_month_match_gets_arrears_note():
    # the SRA's stored month is already the derived service month (document
    # date − 1); when it matches the claim reports an informational note
    # explains the arrears dating
    files = _identified_batch()   # SRA month = 3, same as the claim reports
    gates, hospital, period, notes = validate_batch(files)
    assert all(g.passed for g in gates)
    assert period == (2026, 3)
    assert notes and "καθυστέρηση" in notes[0] and "04/2026" in notes[0]


def test_gate2_wrong_month_sra_warns_but_never_blocks():
    # a wrong month's SRA won't tie out — the reconciliation shows the break,
    # so the date mismatch is a warning, not a hard stop
    files = _identified_batch()
    files[3].month = 5          # SRA seems to settle May, reports are March
    gates, _, period, notes = validate_batch(files)
    assert all(g.passed for g in gates)
    assert period == (2026, 3)  # the claim reports' month wins
    warning = next(n for n in notes if n.startswith("Προσοχή"))
    assert "05/2026" in warning and "03/2026" in warning


def test_gate2_sra_only_period_used_when_reports_have_none():
    files = _identified_batch()
    for f in files:
        if f.report_type != ReportType.SRA:
            f.year, f.month = None, None
    gates, _, period, notes = validate_batch(files)
    assert all(g.passed for g in gates)
    assert period == (2026, 3)
    assert notes


def test_gate3_missing_reports():
    files = _identified_batch()[:3]   # no SRA, no pharmacist fee
    gates, _, _, _ = validate_batch(files)
    failed = [g for g in gates if not g.passed]
    assert failed and failed[0].number == 3
    assert "SRA" in failed[0].message


def test_gate3_crosscheck_mode_waives_sra_only():
    files = [f for f in _identified_batch() if f.report_type != ReportType.SRA]
    gates, _, _, _ = validate_batch(files, crosscheck_mode=True)
    assert all(g.passed for g in gates)


def test_unknown_files_ignored_with_warning_not_blocking():
    # a full-month dump may contain report types the app doesn't know yet:
    # they must be excluded with a warning, never block the run
    files = _identified_batch()
    files.append(IdentifiedFile("mystery_report.xlsx", b"", report_type=None,
                                error="Άγνωστος τύπος αναφοράς"))
    gates, hospital, period, notes = validate_batch(files)
    assert all(g.passed for g in gates)
    assert hospital == "F1049" and period == (2026, 3)
    warn = next(n for n in notes if n.startswith("Προσοχή"))
    assert "mystery_report.xlsx" in warn and "ΑΓΝΟΟΥΝΤΑΙ" in warn


def test_gate1_duplicate_type_rejected():
    files = _identified_batch()
    files.append(identify("endo2.xlsx", synth.inpatient_summary_xlsx()))
    gates, _, _, _ = validate_batch(files)
    assert not gates[-1].passed and gates[-1].number == 1


def test_conditional_requirements_from_sra_lines():
    sra = parse_sra_text(synth.sra_text())
    assert conditional_requirements(sra) == [ReportType.CAPITATION]


def test_gate4_ties():
    assert all(g.passed for g in gate4_internal_asserts(full_bundle()))


def test_gate4_catches_claims_vs_endo_break():
    b = full_bundle()
    b.inpatient.synolo = 1_000_000.00
    g = gate4_internal_asserts(b)[0]
    assert not g.passed and "Ενδ" in g.message


def test_buckets_match_brief_f1049():
    res = run_reconciliation(full_bundle())
    assert res.buckets[Bucket.INPATIENT] == 1_061_728.70
    assert res.buckets[Bucket.AE] == 131_284.66
    assert res.buckets[Bucket.OUTPATIENT] == 78_729.74
    assert res.buckets[Bucket.PHARMA] == 664_785.09
    assert res.cheque_total == 1_936_528.19
    assert round(sum(res.buckets.values()), 2) == res.cheque_total


def test_crosschecks_tie_and_annotate_known_variances():
    res = run_reconciliation(full_bundle(with_optional=True))
    by_name = {c.name: c for c in res.crosschecks}
    endo = by_name["Ενδ. Πληρωμένες Απαιτήσεις (inpatient claims file) = SRA IS"]
    assert endo.diff == 0.0 and endo.flag == "ok"
    fee_gl = next(c for c in res.crosschecks if "25501" in c.name)
    assert fee_gl.flag == "amber"                    # known booking issue
    assert "booking" in fee_gl.note
    assert abs(fee_gl.diff - (24_000.00 - 12_921.60)) < 0.01
    isaud = next(c for c in res.crosschecks if "IS Auditor" in c.name)
    assert isaud.diff == 0.0


def test_split_grand_total_equals_cheque():
    res = run_reconciliation(full_bundle())
    grand = round(sum(s.subtotal for s in res.split), 2)
    assert grand == 1_936_528.19


def test_crosscheck_mode_report_vs_report_flags():
    # without an SRA, checks compare report-vs-report so known variances
    # (pharmacist fee flat GL booking) still get flagged — per the F1054 months
    b = full_bundle(with_optional=True)
    b.sra = None
    res = run_reconciliation(b, crosscheck_mode=True)
    fee_gl = next(c for c in res.crosschecks if "25501" in c.name)
    assert fee_gl.sra_side == 12_921.60      # vs the fee report, not the SRA
    assert fee_gl.flag == "amber"
    endo = next(c for c in res.crosschecks if c.name.startswith("Ενδ."))
    assert endo.sra_side == 1_061_728.70     # vs claims-all Inpatient
    assert endo.flag == "ok"
    isaud = next(c for c in res.crosschecks if "IS Auditor" in c.name)
    assert isaud.diff == 0.0


def test_two_sras_pass_gates_and_reconcile():
    from recon.extract import merge_sras
    files = _identified_batch()
    sra2 = IdentifiedFile("sra2.pdf", b"", report_type=ReportType.SRA,
                          hospital_code="F1049", year=2026, month=3)
    files.append(sra2)
    gates, hospital, period, notes = validate_batch(files)
    assert all(g.passed for g in gates)           # duplicates allowed for SRA
    assert period == (2026, 3)

    b = full_bundle()
    from recon.extract import parse_sra_text
    b.sra = merge_sras([parse_sra_text(synth.sra_text()),
                        parse_sra_text(synth.sra_text_second())])
    assert all(g.passed for g in gate4_internal_asserts(b))
    res = run_reconciliation(b)
    assert res.cheque_total == round(1_936_528.19 + 1_000.00, 2)
    assert res.buckets[Bucket.INPATIENT] == round(1_061_728.70 + 1_000.00, 2)
    from recon.build_xlsx import build_workbook, verify_workbook
    data = build_workbook(res)
    assert verify_workbook(data) == []            # merged workbook still ties
    from openpyxl import load_workbook
    import io
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames[0] == "SRA_259434+900001"


def test_gate4_reports_broken_cheque_by_number():
    from recon.extract import merge_sras, parse_sra_text
    b = full_bundle()
    bad = parse_sra_text(synth.sra_text_second())
    bad.lines[0].amount = 999.00                  # break the second cheque only
    b.sra = merge_sras([parse_sra_text(synth.sra_text()), bad])
    g = gate4_internal_asserts(b)[0]
    assert not g.passed and "#900001" in g.message and "#259434" not in g.message


def test_feb_fee_netting_ties():
    # the pharmacy identity: SRA PH (daily lines) = claims gross + fee invoice
    # — both directions must tie to the cent, CRN-Packages apart as PHF
    from recon.checks import ReconBundle
    from recon.extract import parse_pharmacist_fee_text, parse_sra_text
    from recon.models import PharmaClaims

    b = ReconBundle(hospital_code="F1049", year=2026, month=2)
    b.sra = parse_sra_text(synth.sra_text_feb())
    b.phfee = parse_pharmacist_fee_text(synth.pharmacist_fee_text(packages=7422, unit=1.62))
    b.pharma = PharmaClaims(by_type={"Drugs": 42_623.01})
    res = run_reconciliation(b)
    fee = next(c for c in res.crosschecks if c.side_kind == "fee_net")
    assert fee.source_total == 12_023.64          # 7,422 × 1.62
    assert fee.sra_side == 12_023.64 and fee.flag == "ok"
    ph = next(c for c in res.crosschecks if c.side_kind == "ph_minus_fee")
    assert ph.source_total == 42_623.01
    assert ph.sra_side == 42_623.01 and ph.flag == "ok"
    # the IS check ties now that hemodialysis is its own code
    endo_side = sum(l.amount for l in b.sra.lines if l.code == "IS")
    assert round(endo_side, 2) == 840_526.10
    from recon.build_xlsx import build_workbook, verify_workbook
    assert verify_workbook(build_workbook(res)) == []


def test_xml_activity_filtered_by_payment_number():
    # activities paid by OTHER cheques must drop out of the cross-check
    b = full_bundle()
    b.xml_activity = extract_xml_activity(synth.xml_claims_export_bytes(claims=[
        ("1", "259434", [40_000.00, 20_000.00]),   # this month's cheque
        ("2", "259434", [5_000.00]),
        ("3", "990001", [777.77]),                 # another cheque — excluded
    ]))
    res = run_reconciliation(b)
    row = next(c for c in res.crosschecks if "XML activity" in c.name)
    assert "PAYMENT NO." in row.name
    assert row.source_total == 65_000.00
    assert "777,77" in row.note


def test_endo_vs_sra_gap_named_when_claims_tie_sra():
    # SRA IS == claims file but Ενδ. lags by an old-period claim: the row
    # must say so (amber) instead of «unexplained» (red)
    from recon.models import SRALine
    b = full_bundle()
    b.claims.by_segment["Inpatient"] = round(b.claims.by_segment["Inpatient"] + 1297.43, 2)
    b.claims.inpatient_rows.append(("99476712", "2022-10-18", 1297.43))
    b.sra.lines.append(SRALine(code="IS", description="IS - HCP SERVICES old claim",
                               amount=1297.43, bucket=Bucket.INPATIENT,
                               channel="Claims", source_report="—"))
    res = run_reconciliation(b)
    row = next(c for c in res.crosschecks if "= SRA IS" in c.name)
    assert row.flag == "amber"
    assert "παλαιότερων" in row.note and "99476712" in row.note


def test_gl_inpatient_identity_includes_hemo_and_referral_adjustment():
    # Apr-2026 (F1048) identity, verified to the cent on the real month:
    # GL 26xxx = SRA IS + HEMO + IS-ADJ (the A&E-referral deduction)
    from recon.models import SRALine
    b = full_bundle()
    b.sra.lines.append(SRALine(
        code="HEMO", description="ADJ- IS - Adjustment for Hemodialysis",
        amount=46_100.00, bucket=Bucket.INPATIENT, channel="Adjustment",
        source_report="Hemodialysis report"))
    b.sra.lines.append(SRALine(
        code="IS-ADJ", description="ADJ-AE Referral IS - Adjustment",
        amount=-24_639.25, bucket=Bucket.INPATIENT, channel="Adjustment",
        source_report="Πληρωμένες Απαιτήσεις «all»"))
    b.sra.stated_total = round(b.sra.stated_total + 46_100.00 - 24_639.25, 2)
    rows = [
        ("F1049", "26001", "40001001", 561_728.70),
        ("F1049", "26002", "40001002", 400_000.00),
        ("F1049", "26003", "40001003", 60_000.00),
        # per diem centre holds hemodialysis and the referral deduction
        ("F1049", "26007", "40001003", round(40_000.00 + 46_100.00 - 24_639.25, 2)),
        ("F1049", "25801", "40002001", 131_284.66),
    ]
    b.gl = extract_gl(synth.gl_xlsx(rows=rows), "F1049")
    res = run_reconciliation(b)
    gl_ip = next(c for c in res.crosschecks if c.name.startswith("GL: Ενδονοσοκομειακή"))
    assert gl_ip.source_total == 1_083_189.45
    assert gl_ip.sra_side == 1_083_189.45 and gl_ip.flag == "ok"
    ae = next(c for c in res.crosschecks if "GL: ΤΑΕΠ" in c.name)
    assert ae.flag == "ok"                      # referral adj is NOT in AE
    from recon.build_xlsx import build_workbook, verify_workbook
    assert verify_workbook(build_workbook(res)) == []


def test_prior_period_settlements_stay_out_of_monthly_checks():
    # May-2026: year-end DRG cheque (IS-PRIOR) and innovative-antibiotics
    # cheque (PH-PRIOR) are pass-throughs — the monthly ties must stay green
    # and the settlements must appear as their own split rows
    from recon.models import SRALine
    b = full_bundle()
    b.sra.lines.append(SRALine(
        code="IS-PRIOR", description="ADJ-DRG- IS - Year End Adj. JAN_DEC25",
        amount=104_544.07, bucket=Bucket.INPATIENT, channel="Prior-period",
        source_report="—"))
    b.sra.lines.append(SRALine(
        code="PH-PRIOR", description="Innovativeantibiotic01-Apr2023to30-Sept2025",
        amount=451_569.75, bucket=Bucket.PHARMA, channel="Prior-period",
        source_report="—"))
    b.sra.stated_total = round(b.sra.stated_total + 104_544.07 + 451_569.75, 2)
    res = run_reconciliation(b)
    endo = next(c for c in res.crosschecks if "= SRA IS" in c.name)
    assert endo.flag == "ok"                      # IS-PRIOR not counted as IS
    ph = next(c for c in res.crosschecks if "SRA PHD" in c.name)
    assert ph.flag == "ok"                        # PH-PRIOR not counted as PHD
    ip_rows = [r.label for s in res.split for r in s.rows
               if s.bucket == Bucket.INPATIENT]
    assert any("προηγούμενων περιόδων" in l for l in ip_rows)
    from recon.build_xlsx import build_workbook, verify_workbook
    assert verify_workbook(build_workbook(res)) == []


def test_gl_inpatient_gap_annotated_when_it_matches_z_gap():
    # May-2026: GL 26xxx sits 3.941,25 below SRA IS AND the Z row shows the
    # SAME gap — the known Z-tail classification issue, amber on both rows
    b = full_bundle()
    rows = [
        ("F1049", "26001", "40001001", 561_728.70),
        ("F1049", "26002", "40001002", 400_000.00),
        ("F1049", "26003", "40001003", 60_000.00),
        ("F1049", "26007", "40001003", round(40_000.00 - 3_941.25, 2)),
    ]
    b.gl = extract_gl(synth.gl_xlsx(rows=rows), "F1049")
    res = run_reconciliation(b)
    gl_ip = next(c for c in res.crosschecks if c.name.startswith("GL: Ενδονοσοκομειακή"))
    z = next(c for c in res.crosschecks if "Z-catalogue" in c.name and "GL" in c.name)
    assert gl_ip.diff == z.diff == -3_941.25
    assert gl_ip.flag == "amber" and "Ίδια διαφορά" in gl_ip.note


def test_pd_daily_lines_equal_capitation_plus_claims_pd():
    # exact identity verified Apr+May 2026: SRA PD (daily) = capitation
    # report + claims «Personal Doctors» segment; fixed-price items (OOH,
    # vaccinations) classify apart as PD-FP and don't pollute the tie
    from recon.models import SRALine, SimpleReport
    b = full_bundle()
    b.sra.lines = [l for l in b.sra.lines if l.code != "PD-CAP"]
    b.capitation = SimpleReport(total=21_972.81)
    b.claims.by_segment["Personal Doctors"] = 6_430.86
    b.sra.lines.append(SRALine(
        code="PD", description="PD - HCP SERVICES", amount=28_403.67,
        bucket=Bucket.OUTPATIENT, channel="Claims",
        source_report="Πληρωμένες Απαιτήσεις «all»"))
    b.sra.lines.append(SRALine(
        code="PD-FP", description="PD - OOH SERVICES", amount=660.00,
        bucket=Bucket.OUTPATIENT, channel="Fixed price", source_report="—"))
    b.sra.stated_total = round(sum(l.amount for l in b.sra.lines), 2)
    res = run_reconciliation(b)
    pd = next(c for c in res.crosschecks if "Personal Doctors» = SRA PD" in c.name)
    assert pd.source_total == 28_403.67
    assert pd.sra_side == 28_403.67 and pd.flag == "ok"   # OOH not counted


def test_is_auditor_within_rounding_tolerance_is_ok():
    b = full_bundle(with_optional=True)
    b.isaud.inpatient_total = round(b.isaud.inpatient_total + 2.13, 2)
    res = run_reconciliation(b)
    c = next(ch for ch in res.crosschecks if "IS Auditor" in ch.name)
    assert c.flag == "ok" and "στρογγυλοποίησης" in c.note
    assert c.diff == 2.13                       # the live diff stays visible


def test_capitation_by_doctor_and_workbook_doctor_tab():
    # per-doctor rows parsed from the capitation PDF; workbook gains the
    # «Ανά_ιατρό» tab with claims + capitation sections summed live
    import io
    from openpyxl import load_workbook
    from recon.build_xlsx import build_workbook

    b = full_bundle()
    assert b.capitation.by_doctor == [("D1681 ΜΥΡΟΦΟΡΑ ΙΩΑΝΝΟΥ", 3_255.40)]
    assert b.claims.by_doctor          # populated from ASSOCIATED DOCTOR
    res = run_reconciliation(b)
    wb = load_workbook(io.BytesIO(build_workbook(res)))
    ws = wb["Ανά_ιατρό"]
    text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
    assert "CHRYSTALLA SKORDI" in text
    assert "ΜΥΡΟΦΟΡΑ ΙΩΑΝΝΟΥ" in text
    assert "GENERAL SURGERY" in text
    assert "HIO REIMB" in text          # verification block re-ties to sources


def test_gate4_flags_printed_total_vs_column_sum_mismatch():
    # a report whose ΟΑΥ-printed total disagrees with the summation of its
    # own rows is a finding — the summation is used, the mismatch named
    from recon.models import SimpleReport
    b = full_bundle()
    b.quality = SimpleReport(total=2_585.35, stated_total=2_000.00,
                             lines=[("KPI", 2_585.35)])
    g = gate4_internal_asserts(b)[0]
    assert not g.passed
    assert "Ποιοτικά" in g.message and "άθροισμα" in g.message


def test_gl_capitation_ties_report_when_bundled_in_pd():
    # no PD-CAP line on the SRA (capitation bundled in PD): the GL account
    # must tie the capitation REPORT instead — exact on Apr-2026
    b = full_bundle(with_optional=True)
    b.sra.lines = [l for l in b.sra.lines if l.code != "PD-CAP"]
    b.sra.stated_total = round(sum(l.amount for l in b.sra.lines), 2)
    res = run_reconciliation(b)
    cap = next(c for c in res.crosschecks if "GL: Capitation" in c.name)
    assert "Capitation report" in cap.name
    assert cap.sra_side == b.capitation.total and cap.flag == "ok"


def test_old_period_claim_named_in_gate4_finding():
    # a 2022 claim paid in this cheque but absent from the Ενδ. summary must
    # be NAMED in the gate-4 finding (Larnaca Apr-2026 case: 1,297.43)
    b = full_bundle()
    b.claims.by_segment["Inpatient"] = round(b.claims.by_segment["Inpatient"] + 1297.43, 2)
    b.claims.inpatient_rows.append(("99476712", "2022-10-18", 1297.43))
    g = gate4_internal_asserts(b)[0]
    assert not g.passed
    assert "99476712" in g.message and "2022-10-18" in g.message
    # and the report-vs-report crosscheck row carries the same candidate
    res = run_reconciliation(b)
    row = next(c for c in res.crosschecks if "report vs report" in c.name)
    assert row.flag == "red" and "99476712" in row.note


def test_documented_sra_residual_tolerated_in_zero_checks():
    # gate 4 SRA tie broken by a tiny parsing residual: the run proceeds, the
    # residual shows as a red crosscheck row, and gate 5 tolerates EXACTLY it
    from recon.build_xlsx import build_workbook, verify_workbook
    b = full_bundle()
    b.sra.stated_total = round(b.sra.stated_total + 0.20, 2)   # lines − stated = −0.20
    res = run_reconciliation(b)
    assert res.sra_residual == -0.20
    row = next(c for c in res.crosschecks if "lines vs stated" in c.name)
    assert row.flag == "red"
    data = build_workbook(res)
    assert verify_workbook(data) != []                      # strict: fails
    assert verify_workbook(data, documented_residual=-0.20) == []  # documented: ok
    assert verify_workbook(data, documented_residual=-0.50) != []  # wrong residual


def test_pharmacy_adjustment_descriptions_map_to_pharma():
    from recon.extract import classify_sra_line
    for desc in ["Manual Adj. F104 Χειρόγραφες συνταγές ΓΝ Λάρνακας",
                 "ADJ- Adjustment PharmacyLine - Feb",
                 "ISSUANCES ISSUANCES 11.24-10.25 -4,000,000.00",
                 "Issuances of EOAF & Medicines"]:
        code, bucket, _, _ = classify_sra_line("", desc)
        assert bucket == Bucket.PHARMA, (desc, code)


def test_crosscheck_mode_matrix():
    b = full_bundle(with_optional=True)
    b.sra = None
    res = run_reconciliation(b, crosscheck_mode=True)
    assert res.matrix, "matrix should have rows"
    ip = next(r for r in res.matrix if "Inpatient" in r["stream"])
    values = [v for v in ip["values"].values() if v is not None]
    assert len(values) >= 3            # Ενδ, claims, GL, IS auditor
    assert ip["range"] == 0.0          # three-way tie
    fee = next(r for r in res.matrix if "Φαρμακοποιού" in r["stream"])
    assert fee["range"] == round(24_000.00 - 12_921.60, 2)
