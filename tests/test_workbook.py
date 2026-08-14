"""End-to-end: bundle -> workbook -> reopen -> recompute (gate 5)."""
import io

from openpyxl import load_workbook

from recon.build_xlsx import _Evaluator, build_workbook, verify_workbook
from recon.checks import run_reconciliation

from test_checks import full_bundle


def _build(with_optional=False, crosscheck=False):
    b = full_bundle(with_optional=with_optional)
    if crosscheck:
        b.sra = None
    res = run_reconciliation(b, crosscheck_mode=crosscheck)
    return build_workbook(res), res


def test_workbook_has_five_tabs_and_zero_checks_pass():
    data, _ = _build(with_optional=True)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["SRA_259434", "Reconciliation", "GL_Bridge",
                             "Απαιτήσεις_vs_SRA",
                             "Source_crosscheck", "Ανάλυση_ελέγχων",
                             "By_Clinic_Split", "Ανά_ιατρό",
                             "JOURNAL ENTRIES", "Πώς_δένουν", "Legend"]
    assert verify_workbook(data) == []      # gate 5: every zero-check reads 0


def test_sra_tab_total_is_live_formula_and_ties():
    data, res = _build()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["SRA_259434"]
    n = len(res.bundle.sra.lines)
    total_cell = ws.cell(row=n + 2, column=6).value
    assert isinstance(total_cell, str) and total_cell.startswith("=SUM(")
    ev = _Evaluator(wb)
    assert round(ev.evaluate(total_cell, "SRA_259434"), 2) == 1_936_528.19


def test_reconciliation_buckets_are_sumifs_referencing_label_cells():
    data, _ = _build()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Reconciliation"]
    for r in range(4, 8):
        f = ws.cell(row=r, column=3).value
        assert f.startswith("=SUMIFS(")
        assert f'$B{r}' in f            # criteria reference a label cell
        assert '"' not in f             # never a quoted string
    ev = _Evaluator(wb)
    assert round(ev.evaluate(ws["C4"].value, "Reconciliation"), 2) == 1_061_728.70
    assert round(ev.evaluate(ws["C5"].value, "Reconciliation"), 2) == 131_284.66
    assert round(ev.evaluate(ws["C6"].value, "Reconciliation"), 2) == 78_729.74
    assert round(ev.evaluate(ws["C7"].value, "Reconciliation"), 2) == 664_785.09


def test_live_retie_changing_a_blue_cell_breaks_the_check():
    data, res = _build()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["SRA_259434"]
    ws["F2"] = 999_999.99               # reviewer edits a blue input
    buf = io.BytesIO()
    wb.save(buf)
    failures = verify_workbook(buf.getvalue())
    assert failures, "editing an input must surface a broken zero-check"


def test_pharmacist_fee_row_is_live_packages_times_price():
    data, _ = _build()
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Source_crosscheck"]
    row = next(r for r in range(2, ws.max_row + 1)
               if "Φαρμακοποιού (packages" in str(ws.cell(row=r, column=1).value))
    assert ws.cell(row=row, column=6).value == 8076
    assert ws.cell(row=row, column=7).value == 1.60
    f = ws.cell(row=row, column=2).value
    assert f == f"=F{row}*G{row}"
    ev = _Evaluator(wb)
    assert round(ev.evaluate(f, "Source_crosscheck"), 2) == 12_921.60


def test_crosscheck_mode_workbook():
    data, _ = _build(with_optional=True, crosscheck=True)
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Crosscheck_Matrix", "Source_crosscheck",
                             "Ανάλυση_ελέγχων", "By_Clinic_Split", "Ανά_ιατρό",
                             "Πώς_δένουν", "Legend"]
    assert verify_workbook(data) == []
    ws = wb["Crosscheck_Matrix"]
    # Range column is a live MAX-MIN formula
    found = any(isinstance(c.value, str) and c.value.startswith("=MAX(")
                for row in ws.iter_rows() for c in row)
    assert found


# ------------------------------------------- Ανάλυση_ελέγχων (audit trail)

def _audit_rows(ws):
    return {str(ws.cell(row=r, column=1).value or ""): r
            for r in range(1, ws.max_row + 1)}


def test_audit_tab_reconciles_every_check_with_live_subtotals():
    data, res = _build(with_optional=True)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Ανάλυση_ελέγχων"]
    labels = _audit_rows(ws)
    # one block per check that has both sides
    blocks = [k for k in labels if k[:1].isdigit() and ". " in k]
    assert len(blocks) == len([c for c in res.crosschecks if c.sra_side is not None])
    # every subtotal / difference / tie-back is a formula, never a typed number
    for key, row in labels.items():
        if key.startswith(("   Σύνολο —", "Διαφορά Α", "Έλεγχος: Σύνολο")):
            assert str(ws.cell(row=row, column=2).value).startswith("=")


def test_audit_tab_ties_back_to_source_crosscheck():
    """The two tie-back cells per block prove the audit sheet and
    Source_crosscheck print the same figures — recomputed, not asserted."""
    data, _ = _build(with_optional=True)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Ανάλυση_ελέγχων"]
    ev = _Evaluator(wb)
    ties = [r for r in range(1, ws.max_row + 1)
            if str(ws.cell(row=r, column=1).value or "").startswith("Έλεγχος: Σύνολο")]
    assert ties, "no tie-back cells written"
    for r in ties:
        assert round(ev.evaluate(ws.cell(row=r, column=2).value,
                                 "Ανάλυση_ελέγχων"), 2) == 0.0


def test_audit_tab_breaks_out_the_sides_into_components():
    data, _ = _build(with_optional=True)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Ανάλυση_ελέγχων"]
    labels = _audit_rows(ws)
    # the claims row is broken out per DR SEGMENT
    assert "   DR SEGMENT: Inpatient" in labels
    assert "   DR SEGMENT: Outpatient Specialists" in labels
    # the IS Auditor row separates DRG fees from the Z-catalogue
    assert "   DRG / Fixed-fee αμοιβές" in labels
    # the GL fee row names the ledger centre against packages × unit price
    assert any(k.startswith("   Κέντρο κόστους 25501") for k in labels)
    assert any(k.startswith("   Συσκευασίες 8076 × 1,60 €") for k in labels)
    # SRA components are LIVE SUMIFS whose criteria reference helper cells
    for key, row in labels.items():
        if key.startswith("   SRA γραμμές") or key.startswith("   SRA OS"):
            f = str(ws.cell(row=row, column=2).value)
            assert f.startswith("=SUMIFS(") and '"' not in f
            assert ws.cell(row=row, column=4).value      # the code helper cell


def test_audit_tab_shows_an_unexplained_gap_instead_of_hiding_it():
    """A side whose itemisation doesn't add up gets an explicit
    «not itemised» row — the block still ties, and the gap is visible."""
    from recon.checks import CheckPart
    data, res = _build(with_optional=True)
    chk = next(c for c in res.crosschecks if c.parts_a)
    chk.parts_a = [CheckPart(label="μερική ανάλυση", amount=chk.source_total - 100.0)]
    wb = load_workbook(io.BytesIO(build_workbook(res)))
    ws = wb["Ανάλυση_ελέγχων"]
    row = next(r for r in range(1, ws.max_row + 1)
               if "μη αναλυμένα" in str(ws.cell(row=r, column=1).value or ""))
    assert ws.cell(row=row, column=2).value == 100.0
    assert verify_workbook(build_workbook(res)) == []


def test_audit_tab_breaks_out_the_pharmacy_stream_month():
    """On a PH-stream month the pharma block reads A = drugs + consumables,
    B = the SRA PH lines less the pharmacist-fee invoice."""
    from recon.extract import parse_sra_text
    import synth
    from test_checks import full_bundle
    b = full_bundle()
    b.sra = parse_sra_text(synth.sra_text_feb())
    res = run_reconciliation(b)
    wb = load_workbook(io.BytesIO(build_workbook(res)))
    ws = wb["Ανάλυση_ελέγχων"]
    labels = _audit_rows(ws)
    assert "   Φάρμακα (Drugs)" in labels and "   Αναλώσιμα (Consumables)" in labels
    assert "   SRA γραμμές PH (φαρμακείο)" in labels
    assert "   μείον τιμολόγιο αμοιβής φαρμακοποιού" in labels
    ev = _Evaluator(wb)
    row = labels["   SRA γραμμές PH (φαρμακείο)"]
    f = ws.cell(row=row, column=2).value
    assert f.startswith("=SUMIFS(")
    assert round(ev.evaluate(f, "Ανάλυση_ελέγχων"), 2) == 54_646.65


def test_tolerance_ok_check_is_not_claimed_to_be_zero():
    """A check that passes WITHIN a documented tolerance (IS Auditor per-row
    rounding) has a non-zero difference — it must stay visible, not be
    painted as a zero-check that gate 5 then fails on."""
    from test_checks import full_bundle
    b = full_bundle(with_optional=True)
    b.isaud.drg_fees = round(b.isaud.drg_fees + 2.61, 2)   # rounding drift
    b.isaud.inpatient_total = round(b.isaud.inpatient_total + 2.61, 2)
    res = run_reconciliation(b)
    chk = next(c for c in res.crosschecks if "IS Auditor" in c.name)
    assert chk.flag == "ok" and chk.diff == 2.61            # ok, but NOT zero
    data = build_workbook(res)
    assert verify_workbook(data) == []                      # gate 5 still passes
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Ανάλυση_ελέγχων"]
    block = next(r for r in range(1, ws.max_row + 1)
                 if "IS Auditor" in str(ws.cell(row=r, column=1).value or "")
                 and str(ws.cell(row=r, column=1).value or "")[:1].isdigit())
    diff = next(r for r in range(block, block + 30)
                if str(ws.cell(row=r, column=1).value or "").startswith("Διαφορά"))
    cell = ws.cell(row=diff, column=2)
    assert not str(cell.fill.fgColor.rgb).endswith("FFFF00")
    assert round(_Evaluator(wb).evaluate(cell.value, "Ανάλυση_ελέγχων"), 2) == 2.61


# ------------------------------- three-column inpatient split + GL bridge

def test_split_tab_writes_drg_daily_and_zdrugs_with_column_subtotals():
    """The inpatient fee splits three ways. Each column carries its own live
    subtotal, and the grand total still ties to the cheque."""
    import synth
    from recon.checks import ReconBundle
    from test_checks import full_bundle
    b = full_bundle()
    b.inpatient = __import__("recon.extract", fromlist=["x"]).extract_inpatient_summary(
        synth.inpatient_summary_xlsx(with_procedure_detail=True))
    b.claims.inpatient_by_clinic = []          # use the Ενδ. per-clinic detail
    res = run_reconciliation(b)
    data = build_workbook(res)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["By_Clinic_Split"]
    assert [ws.cell(row=3, column=c).value for c in range(1, 6)] == [
        "Κλινική / Γραμμή (Clinic / Line)", "DRG €",
        "Ημερήσιες θεραπείες (Daily treat.) €",
        "Ζ-φάρμακα/πράξεις (Z-drugs) €", "Ποσό (Amount €)"]
    ev = _Evaluator(wb)
    sub = next(r for r in range(4, ws.max_row + 1)
               if str(ws.cell(row=r, column=1).value or "").startswith("Υποσύνολο")
               and "Inpatient" in str(ws.cell(row=r, column=1).value))
    drg, daily, z = (round(ev.evaluate(ws.cell(row=sub, column=c).value,
                                       "By_Clinic_Split"), 2) for c in (2, 3, 4))
    assert (drg, daily, z) == (700_000.00, 336_000.00, 25_728.70)
    assert verify_workbook(data) == []         # still ties to the cheque


def test_gl_bridge_tab_compares_cash_with_booked_and_checks_the_variances():
    data, res = _build(with_optional=True)
    wb = load_workbook(io.BytesIO(data))
    assert "GL_Bridge" in wb.sheetnames
    ws = wb["GL_Bridge"]
    ev = _Evaluator(wb)
    rows = {str(ws.cell(row=r, column=1).value or ""): r
            for r in range(1, ws.max_row + 1)}
    # panel A links to Reconciliation — the same figure the cheque ties to
    ip = next(r for k, r in rows.items() if k.startswith("Ενδονοσοκομειακή"))
    assert str(ws.cell(row=ip, column=2).value).startswith("='Reconciliation'!")
    assert round(ev.evaluate(ws.cell(row=ip, column=2).value, "GL_Bridge"), 2) == \
        1_061_728.70
    # panel B is the ledger, panel C the live variance
    assert ws.cell(row=ip, column=4).value == 1_061_728.70
    assert ws.cell(row=ip, column=5).value == f"=B{ip}-D{ip}"
    ph = next(r for k, r in rows.items() if k.startswith("Φάρμακα"))
    assert round(ev.evaluate(ws.cell(row=ph, column=5).value, "GL_Bridge"), 2) == \
        round(664_785.09 - (24_000.00 + 651_863.49), 2)
    assert ws.cell(row=ph, column=6).value      # the variance is explained
    # both zero-checks recompute to 0 (gate 5 covers them too)
    for key in ("Zero-check = ταμείο ανά καλάθι − επιταγή (must be 0)",
                "Zero-check = άθροισμα διαφορών − (ταμείο − καθολικό) (must be 0)"):
        r = rows[key]
        cell = next(c for c in (2, 5) if isinstance(ws.cell(row=r, column=c).value, str))
        assert round(ev.evaluate(ws.cell(row=r, column=cell).value, "GL_Bridge"), 2) == 0.0
    assert verify_workbook(data) == []


def test_gl_bridge_is_absent_without_a_gl_extract():
    data, _res = _build(with_optional=False)
    assert "GL_Bridge" not in load_workbook(io.BytesIO(data)).sheetnames


def test_hospital_sap_journal_posts_every_revenue_stream_in_one_document():
    """A hospital posts by clinic and stream, not by professional, so its
    journal lines are the By_Clinic_Split rows — one document carrying the
    whole month's revenue and tying to the cheque."""
    data, _res = _build(with_optional=True)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["JOURNAL ENTRIES"]
    ev = _Evaluator(wb)
    debits = [r for r in range(4, ws.max_row + 1)
              if ws.cell(row=r, column=9).value == "01"
              and ws.cell(row=r, column=10).value == "200000"]
    credits = [r for r in range(4, ws.max_row + 1)
               if ws.cell(row=r, column=9).value == "50"
               and ws.cell(row=r, column=10).value == "412002"]
    assert len(debits) == 1                       # the month, in one document
    # all four buckets present, named in the analysis column
    assert {ws.cell(row=r, column=25).value for r in credits} == \
        {"Inpatient", "A&E", "Outpatient", "Pharma"}
    # credits = the cheque, and the debit line is their live SUM
    assert round(sum(ws.cell(row=r, column=12).value for r in credits), 2) == \
        1_936_528.19
    assert round(ev.evaluate(ws.cell(row=debits[0], column=12).value,
                             "JOURNAL ENTRIES"), 2) == 1_936_528.19
    # every bucket adds up to what the Reconciliation tab pays it
    rec = wb["Reconciliation"]
    for i, bucket in enumerate(("Inpatient", "A&E", "Outpatient", "Pharma")):
        posted = round(sum(ws.cell(row=r, column=12).value for r in credits
                           if ws.cell(row=r, column=25).value == bucket), 2)
        assert posted == round(ev.evaluate(rec.cell(row=4 + i, column=3).value,
                                           "Reconciliation"), 2)
    assert verify_workbook(data) == []


def test_hospital_journal_codes_a_whole_bucket_from_one_lookup_row():
    """Eight hospitals, one lookup file: rows carry an F-code so a cost centre
    belongs to one hospital, and a row keyed on the BUCKET codes every line in
    it — four rows per hospital are enough to post at stream level."""
    import synth
    from recon.mapping import extract_cost_centres
    data, res = _build(with_optional=True)
    res.bundle.cost_centres = extract_cost_centres(synth.cost_centre_map_xlsx(
        rows=[("Inpatient", "26001", "11", "INPATIENT", "", "F1049"),
              ("A&E", "25801", "12", "A AND E", "", "F1049"),
              ("Outpatient", "25500", "13", "OUTPATIENT", "", "F1049"),
              ("Pharma", "25501", "14", "PHARMA", "", "F1049"),
              # another hospital's codes must never leak into this one
              ("Inpatient", "99999", "99", "OTHER HOSPITAL", "", "F1054")],
        with_hospital=True))
    wb = load_workbook(io.BytesIO(build_workbook(res)))
    ws = wb["JOURNAL ENTRIES"]
    credits = [r for r in range(4, ws.max_row + 1)
               if ws.cell(row=r, column=9).value == "50"
               and ws.cell(row=r, column=10).value == "412002"]
    assert credits and all(ws.cell(row=r, column=14).value for r in credits)
    assert "99999" not in {ws.cell(row=r, column=14).value for r in credits}
    coded = {ws.cell(row=r, column=25).value: ws.cell(row=r, column=14).value
             for r in credits}
    assert coded == {"Inpatient": "26001", "A&E": "25801",
                     "Outpatient": "25500", "Pharma": "25501"}
