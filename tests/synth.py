"""Synthetic ΟΑΥ report files shaped like the REAL ones (per the diagnostics
of F1049 Mar-2026): invoice-level SRA lines, «Έτος | Μήνας» labeled columns,
Γέννες category, «per clinic» pivot sheet, dotted headers (HIO REIMB.),
ISO claim dates from earlier months, capitation invoice rows.

The month mirrors the F1049 Mar-2026 acceptance numbers from the brief:
cheque 1,936,528.19 / buckets 1,061,728.70 + 131,284.66 + 78,729.74
+ 664,785.09 / pharmacist fee 8,076 × 1.60 = 12,921.60.
"""
from __future__ import annotations

import io

from lxml import etree
from openpyxl import Workbook


def _wb_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _anglo(v: float) -> str:
    return f"{v:,.2f}"


# ---------------------------------------------------------------- Ενδ. summary

def inpatient_summary_xlsx(kanonika=475_000.00, kanonika_parap=61_728.70,
                           exeidikevmena=300_000.00, exeid_parap=100_000.00,
                           gennes=25_000.00, z=100_000.00, hospital="F1049",
                           hospital_name="ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ ΑΜΜΟΧΩΣΤΟΥ (ΟΚΥπΥ)",
                           year=2026, month=3, synolo=None,
                           with_per_clinic=True, detail_extra=0.0,
                           with_detail=True, with_procedure_detail=False) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # real layout: labeled columns, values in the row below
    ws.append(["Κωδικός ΓεΣΥ Παροχέα", "Επωνυμία Παροχέα", "Έτος", "Μήνας"])
    ws.append([hospital, hospital_name, year, month])
    ws.append([])
    ws.append(["ΣΥΝΟΠΤΙΚΟΣ ΠΙΝΑΚΑΣ *"])
    ws.append(["Κατηγορία", "Συμφωνημένος αριθμός μονάδων για τον μήν",
               "Συμφωνημένο Base Rate  (€)", "Πραγματικός αριθμός μονάδων",
               "Αμοιβή νοσηλευτηρίου (€)", "Συνολική αμοιβή (€)"])
    ws.append(["Κανονικά", 342.3012, 3469, 48.974, kanonika, kanonika])
    ws.append(["Εξειδικευμένα", 11.0524, 4187, 8.718, exeidikevmena, exeidikevmena])
    ws.append(["Γέννες", 0, 3400, 6.245, gennes, gennes])
    ws.append(["Κανονικά με παραπεμπτικό από ΤΑΕΠ", 0, 3469, 149.027,
               kanonika_parap, kanonika_parap])
    ws.append(["Εξειδικευμένα με παραπεμπτικό", 0, 4187, 23.9, exeid_parap, exeid_parap])
    ws.append(["Κατάλογος Ζ", None, None, None, z, z])
    total = synolo if synolo is not None else round(
        kanonika + kanonika_parap + exeidikevmena + exeid_parap + gennes + z, 2)
    ws.append(["Σύνολο", None, None, None, total, total])

    # per-claim listing below the summary — the WIDER universe: with
    # detail_extra it also holds an old-period claim absent from the
    # ΣΥΝΟΠΤΙΚΟΣ (the real Apr-2026 F1048 case, claim 99476712)
    if with_detail:
        ws.append([])
        ws.append(["Αρ. Απαίτησης", "Ημ. Εισαγωγής", "DRG", "Συνολική αμοιβή (€)"])
        half = round(total * 0.5, 2)
        ws.append([99_000_001, "2026-03-05", "K60A", half])
        ws.append([99_000_002, "2026-03-14", "F62B", round(total - half, 2)])
        if detail_extra:
            ws.append([99_476_712, "2022-10-18", "G67B", detail_extra])
        ws.append(["Σύνολο", None, None, round(total + detail_extra, 2)])

    if with_procedure_detail:
        # the per-claim DETAIL table: «Procedure Class Id» is the only place
        # daily treatments (FixedFee) and Z-catalogue items (ZDRUG/ZPROC/
        # ZCONSU) can be told apart — ΟΑΥ's per-clinic pivot lumps them
        # — and, as in the real Apr-2026 files, it lists only the FIXED FEE
        # side: the DRG euros stay in the per-clinic pivot alone, so the two
        # tables have to be combined, not chosen between
        det = wb.create_sheet("detail")
        det.append(["Claim Id", "Specialty", "Procedure Class Id",
                    "Hospital amount per claim activity (€)"])
        for clinic, daily, zdrug, zproc in (
                ("INTERNAL MEDICINE", 100_000.00, 10_728.70, 1_000.00),
                ("GENERAL SURGERY", 140_000.00, 9_360.00, 640.00),
                ("CARDIOLOGY", 96_000.00, 4_000.00, 0.0)):
            det.append([1, clinic, "FixedFee", daily])
            det.append([2, clinic, "ZDRUG", zdrug])
            det.append([3, clinic, "ZPROC", zproc])
            det.append([4, clinic, "ZCONSU", 0.0])

    if with_per_clinic:
        # real workbooks carry a «per clinic» pivot (headers duplicated twice)
        pc = wb.create_sheet("per clinic")
        pc.append(["Row Labels", "Sum of FIXED FEE", "Sum of INPATIENTS",
                   "Sum of Grand Total", "Row Labels", "Sum of FIXED FEE",
                   "Sum of INPATIENTS", "Sum of Grand Total"])
        clinics = [("INTERNAL MEDICINE", 111_728.70, 250_000.00),
                   ("GENERAL SURGERY", 150_000.00, 250_000.00),
                   ("CARDIOLOGY", 100_000.00, 200_000.00)]
        for name, ff, ip in clinics:
            gt = round(ff + ip, 2)
            pc.append([name, ff, ip, gt, name, ff, ip, gt])
        pc.append(["Grand Total", 361_728.70, 700_000.00, 1_061_728.70])
    return _wb_bytes(wb)


# --------------------------------------------------------------- claims «all»

DEFAULT_SEGMENTS = {
    "Inpatient": 1_061_728.70,
    "A&E": 131_284.66,
    "Outpatient Specialists": 40_000.00,
    "Nurses-Midwives": 20_000.00,
    "Allied Health": 5_000.00,
}


def claims_all_xlsx(segments=None, cheque="259434", with_doctor=True) -> bytes:
    """Real shape: dotted headers, ISO claim dates from EARLIER months (old
    claims paid in this cheque), PAYMENT NO. = the cheque, DR SEGMENT in the
    trailing columns, NO clinic/specialty columns, NO F-code anywhere.

    with_doctor=False reproduces the thinner export some hospitals get: no
    ASSOCIATED DOCTOR / DR SPECIALITY columns, so no per-doctor detail."""
    segments = segments if segments is not None else DEFAULT_SEGMENTS
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    head = ["CLAIM ID", "STATUS", "BENEFICIARY NAME", "BENEFICIARY ID",
            "INVOICE DATE", "SUBM. DATE", "VISIT ID", "PAYMENT NO.",
            "CO-PAYMENT", "PERS. CONTR. I", "HIO REIMB.", "TOTAL AMT"]
    ws.append(head + (["ASSOCIATED DOCTOR", "DR SPECIALITY", "DR SEGMENT"]
                      if with_doctor else ["DR SEGMENT"]))
    cid = 121_989_416
    dates = ["2026-01-25", "2026-02-09", "2024-03-04"]
    specs = {"Inpatient": "GENERAL SURGERY",
             "Outpatient Specialists": "CARDIOLOGY",
             "A&E": "ACCIDENT & EMERGENCY",
             "Nurses-Midwives": "NURSING",
             "Allied Health": "PHYSIOTHERAPY"}
    doctors = ["ΧΡΥΣΤΑΛΛΑ ΣΚΟΡΔΗ / CHRYSTALLA SKORDI",
               "TZANIS LEONTARIDIS / TZANIS LEONTARIDIS"]
    for seg, amount in segments.items():
        parts = [round(amount * 0.6, 2)]
        parts.append(round(amount - parts[0], 2))
        for k, part in enumerate(parts):
            row = [cid, "Paid", "SYNTHETIC BENEFICIARY", "NID 0000000000",
                   dates[k % len(dates)], "2026-03-02", 73_858_185 + cid % 999,
                   cheque, 0, 0, part, part]
            row += ([doctors[k % 2], specs.get(seg, "GENERAL"), seg]
                    if with_doctor else [seg])
            ws.append(row)
            cid += 1
    return _wb_bytes(wb)


# --------------------------------------------------------------- pharma claims

def pharma_claims_xlsx(drugs=600_000.00, consumables=51_863.49,
                       cheque="259434") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append(["DISPENS. ID", "PRESCRIPTION ID", "STATUS", "TYPE", "CLAIM DATE",
               "PRESCRIBING DOCTOR NAME", "EXECUTED BY", "CO-PAYMENT", "PC II",
               "HIO REIMB.", "TOTAL AMT"])
    ws.append([77_356_519, 48_576_587, "Paid", "Drugs", "2024-03-04",
               "SYNTH DOCTOR", "SYNTH PHARMACIST", 1, 0, round(drugs * 0.6, 2),
               round(drugs * 0.6, 2) + 1])
    ws.append([123_362_618, 69_276_579, "Paid", "Drugs", "2026-02-23",
               "SYNTH DOCTOR", "SYNTH PHARMACIST", 1, 0,
               round(drugs - round(drugs * 0.6, 2), 2), 0])
    ws.append([123_426_107, 74_313_353, "Paid", "Consumables", "2026-02-23",
               "SYNTH DOCTOR", "SYNTH PHARMACIST", 1, 0, consumables, 0])
    return _wb_bytes(wb)


# ----------------------------------------------------------------- PDFs (text)

def sra_text(cheque="259434", hospital="F1049") -> str:
    """Real SRA format: header block with Payment Date (one month AFTER the
    service month), then invoice-level lines
    «date invoice-no description invoice-total EUR amount-paid»."""
    lines = [
        "ΚΑΤΑΣΤΑΣΗ ΠΛΗΡΩΜΗΣ",
        "REMITTANCE ADVICE",
        "Klimentos 17 & 19, 4th floor",
        "1061 Nicosia",
        "tel: 22557200",
        f"STATE HEALTH SERVICES ORGANIZATION INCOME- PAR-{hospital} Payment Date: 20/04/2026",
        "Προδρόμου 1 Payment Currency: EUR",
        f"Strovolos 2063 Payment/Cheque No: {cheque}",
        "Cyprus Supplier No: 1491",
        "Total paid in this batch: 1,936,528.19",
        "Page No: 1 of 2",
        "Ημερομηνία Αρ . Τιμολογίου Περιγραφή Ποσό Τιμολογίου Νόμισμα Ποσό πληρωμής",
        "Invoice Date Invoice No. Description Invoice Total Currency Amount Paid",
    ]
    invoices = [
        ("01/03/2026", 5_636_100, "IS - HCP SERVICES", 530_864.35),
        ("15/03/2026", 5_636_101, "IS - HCP SERVICES", 530_864.35),
        ("01/03/2026", 5_636_247, "AE - HCP SERVICES", 65_642.33),
        ("02/03/2026", 5_640_316, "AE - HCP SERVICES", 65_642.33),
        ("03/03/2026", 5_644_001, "OS - HCP SERVICES", 40_000.00),
        ("04/03/2026", 5_644_002, "NM - HCP SERVICES", 20_000.00),
        ("05/03/2026", 5_644_003, "AP - HCP SERVICES", 5_000.00),
        ("31/03/2026", 5_729_128, "PD - CAPITATION", 13_729.74),
        ("07/03/2026", 5_644_005, "PHD - PHARMACY DRUGS", 600_000.00),
        ("08/03/2026", 5_644_006, "PHC - PHARMACY CONSUMABLES", 51_863.49),
        ("30/03/2026", 5_730_058, "PHF - ΑΜΟΙΒΗ ΦΑΡΜΑΚΟΠΟΙΟΥ", 12_921.60),
    ]
    for date, inv, desc, amt in invoices:
        lines.append(f"{date} {inv} {desc} {_anglo(amt)} EUR {_anglo(amt)}")
    lines.append("Total paid in this batch: 1,936,528.19")
    lines.append("Page No: 2 of 2")
    return "\n".join(lines)


def pharmacist_fee_text(packages=8076, unit=1.60, hospital="F1049") -> str:
    # the unit price varies by month (1.60 €, then 1.62 €) — read, not assumed
    amount = _anglo(round(packages * unit, 2))
    return f"""Αμοιβή Φαρμακοποιού
Μήνας: 3
Έτος: 2026
Κωδικός Παροχέα Χρέωσης: {hospital}
ID Τιμολογίου Τύπος Πληρωμένο Ημερομηνία Παροχέας Υγείας Τιμή Μονάδας Συσκευασίες που Αμοιβή από ΟΑΥ
EBS Τιμολογίου Τιμολογίου εκτελέστηκαν
5730058 STANDARD No 30/03/2026 {hospital} ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ {unit:.2f} € {packages} {amount} €
ΑΜΜΟΧΩΣΤΟΥ (ΟΚΥπΥ)
Σελίδα 1
"""


def sra_text_feb(cheque="256797", hospital="F1049") -> str:
    """February-style SRA: «PH - HCP SERVICES» pharmacy stream (which also
    carries the pharmacist-fee invoice), CRN-Packages fee corrections, a
    hemodialysis adjustment labelled «ADJ- IS -», an ADJ-MRI/CT line, and a
    trailing-minus credit note."""
    return f"""ΚΑΤΑΣΤΑΣΗ ΠΛΗΡΩΜΗΣ
REMITTANCE ADVICE
STATE HEALTH SERVICES ORGANIZATION INCOME- PAR-{hospital} Payment Date: 13/03/2026
Strovolos 2063 Payment/Cheque No: {cheque}
Total paid in this batch: 929,012.04
Ημερομηνία Αρ . Τιμολογίου Περιγραφή Ποσό Τιμολογίου Νόμισμα Ποσό πληρωμής
28/02/2026 ADJ-MRI/CT QC- 501.03 EUR 501.03
Feb
Invoice Date Invoice No. Description Invoice Total Currency Amount Paid
01/02/2026 5548208 AE - HCP SERVICES 24,327.00 EUR 24,327.00
01/02/2026 5548210 IS - HCP SERVICES 840,526.10 EUR 840,526.10
01/02/2026 5548223 PH - HCP SERVICES 42,623.01 EUR 42,623.01
27/02/2026 5632647 PH - HCP SERVICES 12,023.64 EUR 12,023.64
28/02/2026 CRN-Packages PH - CORRECTION-Packages-02-2026 6,388.29- EUR 6,388.29-
28/02/2026 ADJ- IS - Adjustment for Hemodialysis 5,260.00 EUR 5,260.00
05/02/2026 5560001 OS - HCP SERVICES 10,000.00 EUR 10,000.00
06/02/2026 5560002 CREDIT NOTE AE 12.25- EUR 12.25-
se 12.25
28/02/2026 KPIs-02-2026- PD-KPIs-02-2026-CHILD 151.80 EUR 151.80
Total paid in this batch: 929,012.04
"""


def sra_text_second(cheque="900001", hospital="F1049") -> str:
    """A small second cheque for the same month (multi-SRA scenario)."""
    return f"""ΚΑΤΑΣΤΑΣΗ ΠΛΗΡΩΜΗΣ
REMITTANCE ADVICE
STATE HEALTH SERVICES ORGANIZATION INCOME- PAR-{hospital} Payment Date: 22/04/2026
Payment/Cheque No: {cheque}
Total paid in this batch: 1,000.00
Invoice Date Invoice No. Description Invoice Total Currency Amount Paid
20/03/2026 5700001 IS - HCP SERVICES 1,000.00 EUR 1,000.00
Total paid in this batch: 1,000.00
"""


def capitation_text(total=13_729.74, hospital="F1049") -> str:
    return f"""Κατά κεφαλήν αμοιβή
Μήνας: 3
Έτος: 2026
Κωδικός Παροχέα Χρέωσης: {hospital}
ID Τιμολογίου Αμοιβή από
Παροχέας Υγείας Τύπος Τιμολογίου Ημερομηνία Τιμολογίου
EBS ΟΑΥ
5729128 {hospital} ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ STANDARD 31/03/2026 {_anglo(total)} €
ΑΜΜΟΧΩΣΤΟΥ (ΟΚΥπΥ)
D1681 ΜΥΡΟΦΟΡΑ ΙΩΑΝΝΟΥ / Ηλικίες Σχόλια Ημερήσια Κατά Αριθμός Συνολικός 3,255.40 €
0 - 3 years - 0.508 41 1238 628.58 €
4 - 7 years - 0.375 75 2309 865.50 €
"""


def quality_criteria_xlsx(total=0.0) -> bytes:
    """Real quality-criteria export: header row with SINGULAR «QUALITY
    CRITERION», amounts as text like '€ 0.00', often empty."""
    wb = Workbook()
    ws = wb.active
    ws.append(["CLAIM DATE", "CLAIM ID", "QUALITY CRITERION", "AMOUNT",
               "PERSONAL DOCTOR CODE", "PERSONAL DOCTOR", "DR SPECIALITY"])
    ws.append(["SUM AMOUNT"])
    ws.append(["Total", f"€ {total:.2f}"])
    return _wb_bytes(wb)


# -------------------------------------------------------------------- GL / IS

def gl_xlsx(rows=None, sheet_name="ALL OKYPY 03.26", hospital="F1049") -> bytes:
    """(vendor, cost_center, account, amount) rows; defaults tie to the synthetic
    month except the pharmacist fee, which reproduces the known flat-booking gap.

    The GL is org-wide: it always carries a foreign vendor row that the
    extractor must filter out, whichever hospital is being reconciled."""
    other = "F1054" if hospital != "F1054" else "F1050"
    if rows is None:
        rows = [
            (hospital, "26001", "40001001", 561_728.70),   # regular DRG
            (hospital, "26002", "40001002", 400_000.00),   # specialized
            (hospital, "26003", "40001003", 60_000.00),    # Z-catalogue
            (hospital, "26007", "40001003", 40_000.00),    # Z-catalogue
            (hospital, "25801", "40002001", 131_284.66),   # A&E
            (hospital, "25301", "40003001", 65_000.00),    # outpatient clinical
            (hospital, "25501", "40004001", 24_000.00),    # pharmacist fee (flat!)
            (hospital, "25502", "40004002", 651_863.49),   # pharma
            (hospital, "10101", "51001001", 13_729.74),    # capitation account
            (other, "26001", "40001001", 9_999_999.99),    # another hospital, filtered out
        ]
    wb = Workbook()
    # decoy sheet FIRST, like the real Apr-2026 workbook: an A&E-only clinic
    # detail whose rows would zero every other bucket if the extractor
    # stopped at the first sheet carrying the header
    decoy = wb.active
    decoy.title = "A&E detail"
    decoy.append(["VENDOR_CODE", "COST_CENTER", "ACCOUNT", "JHDF", "EURO_AMOUNT"])
    decoy.append([hospital, "25801", "51101099", "A&E income", 141_284.66])
    decoy.append([hospital, "25801", "43010001", "Copayments", -10_000.00])
    decoy.append([f"{hospital} Total", "", "", "", 131_284.66])   # subtotal row
    ws = wb.create_sheet(sheet_name)
    ws.append(["VENDOR_CODE", "COST_CENTER", "ACCOUNT", "EURO_AMOUNT"])
    for r in rows:
        ws.append(list(r))
    return _wb_bytes(wb)


def provider_long_name(hospital="F1049") -> str:
    """The org-wide reports spell the provider out in full, e.g.
    «ΓΕΝΙΚΟ ΝΟΣΟΚΟΜΕΙΟ ΑΜΜΟΧΩΣΤΟΥ (ΟΚΥπΥ)»."""
    from recon.models import HOSPITALS
    return f"{HOSPITALS[hospital][0]} (ΟΚΥπΥ)"


def is_auditor_xlsx(rows=None, hospital="F1049") -> bytes:
    """(provider, drg_id, drg_ff_amount, procedures_amount, invoice_category).
    Invoice dates span YEARS (old claims paid now) like the real report."""
    mine = provider_long_name(hospital)
    other = provider_long_name("F1054" if hospital != "F1054" else "F1050")
    if rows is None:
        rows = [
            (mine, "DRG001", 500_000.00, 30_000.00, "Normal", "31/08/2023"),
            (mine, "DRG002", 400_000.00, 51_728.70, "Specialised", "14/10/2024"),
            (mine, None, None, 50_000.00, "Normal", "07/02/2026"),
            (mine, "", None, 30_000.00, "Normal", "30/04/2025"),
            (other, "DRG003", 8_888_888.88, 0, "Normal", "31/08/2025"),
        ]
    wb = Workbook()
    ws = wb.active
    ws.append(["Case Nbr", "Billing Provider Name", "DRG Id", "DRG/FF Total Amount",
               "Procedures Total Amount", "Invoice Category", "Invoice Date"])
    for i, r in enumerate(rows):
        ws.append([19_825_630 + i] + list(r))
    return _wb_bytes(wb)


# ------------------------------------------------------------------- XML

def xml_activity_bytes(amounts=None, hospital="F1049") -> bytes:
    amounts = amounts if amounts is not None else [40_000.00, 20_000.00, 5_000.00]
    root = etree.Element("Activities")
    etree.SubElement(root, "Provider").text = hospital
    etree.SubElement(root, "Period").text = "03/2026"
    for i, amt in enumerate(amounts, start=1):
        act = etree.SubElement(root, "Activity")
        etree.SubElement(act, "ClaimId").text = f"CL{i:05d}"
        etree.SubElement(act, "ActivityReimbursementAmount").text = f"{amt:.2f}"
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


def xml_claims_export_bytes(claims=None) -> bytes:
    """Real Apr-2026 shape: ClaimsExport > Claims > Claim, each claim carrying
    ClaimPaymentNumber (the SRA cheque) and its Activities.
    claims = [(claim_id, payment_no, [activity amounts]), ...]"""
    if claims is None:
        claims = [("124358528", "263000", [40.75, 20.00]),
                  ("124360429", "263000", [60.00]),
                  ("125462247", "263367", [137.89]),
                  ("120886546", "990001", [500.00])]   # paid by another cheque
    root = etree.Element("ClaimsExport")
    wrap = etree.SubElement(root, "Claims")
    for cid, pay, amounts in claims:
        c = etree.SubElement(wrap, "Claim")
        etree.SubElement(c, "ClaimId").text = cid
        etree.SubElement(c, "ClaimPaymentNumber").text = pay
        acts = etree.SubElement(c, "Activities")
        for amt in amounts:
            a = etree.SubElement(acts, "Activity")
            etree.SubElement(a, "ActivityReimbursementAmount").text = f"{amt:.2f}"
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


# ------------------------------------------- mental-health (non-hospital)

def provider_sra_text(code="F1089", cheque="266458", os_amt=42_358.08,
                      nm_amt=71_252.16, ap_amt=30_917.80, adj=896.21) -> str:
    """Real mental-health SRA shape: the payee wraps across two header lines
    («...INCOME-MENTAL / HEALTH-F1089 Payment Currency: EUR») and only the
    service streams OS / NM / AP are paid — no DRG, no pharmacy."""
    total = round(os_amt + nm_amt + ap_amt + adj, 2)
    lines = [
        "ΚΑΤΑΣΤΑΣΗ ΠΛΗΡΩΜΗΣ",
        "REMITTANCE ADVICE",
        "Klimentos 17 & 19, 4th floor",
        "1061 Nicosia",
        "STATE HEALTH SERVICES ORGANIZATION INCOME-MENTAL Payment Date: 12/06/2026",
        f"HEALTH-{code} Payment Currency: EUR",
        f"Προδρόμου 1 Payment/Cheque No: {cheque}",
        "Strovolos 2063 Supplier No: 2951",
        f"Cyprus Total paid in this batch: {_anglo(total)}",
        "Invoice Date Invoice No. Description Invoice Total Currency Amount Paid",
    ]
    for i, (desc, amt) in enumerate((("OS - HCP SERVICES", os_amt),
                                     ("NM - HCP SERVICES", nm_amt),
                                     ("AP - HCP SERVICES", ap_amt))):
        if amt:
            lines.append(f"31/05/2026 {5800001 + i} {desc} {_anglo(amt)} EUR {_anglo(amt)}")
    if adj:
        lines.append("31/05/2026 ADJ-New Reimb OS - Adj. based on new reimb. method- "
                     f"{_anglo(adj)} EUR {_anglo(adj)}")
    lines.append(f"Total paid in this batch: {_anglo(total)}")
    return "\n".join(lines)


def provider_claims_xlsx(cheque="266458", segments=None) -> bytes:
    """Paid claims for one unit: no F-code anywhere — PAYMENT NO. is the only
    content-based link back to its provider."""
    segments = segments if segments is not None else {
        "Outpatient Specialists": 42_358.08, "Nurses-Midwives": 71_252.16,
        "Allied Health": 30_917.80}
    return claims_all_xlsx(segments=segments, cheque=cheque)


def activity_table_xlsx(provider="F1089", name="ΚΟΙΝΟΤΙΚΑ ΚΕΝΤΡΑ ΓΙΑ ΕΝΗΛΙΚΕΣ "
                        "ΨΥΧΙΚΗΣ ΥΓΕΙΑΣ (ΟΚΥπΥ)", rows=None) -> bytes:
    """The activity export as ΟΑΥ ships it for these units: a SPREADSHEET
    (in a folder called «XMLS»), one row per activity."""
    rows = rows if rows is not None else [
        ("CL1", "266458", 100_000.00), ("CL2", "266458", 44_528.04),
        ("CL3", "263370", 1_492.73)]        # paid by an earlier cheque
    wb = Workbook()
    ws = wb.active
    ws.append(["ExportDate", "ClaimsDateFrom", "ClaimsDateTo", "ProviderId",
               "ProviderName", "ClaimId", "ClaimPaymentNumber",
               "ActivityReimbursementAmount"])
    for cid, pay, amt in rows:
        ws.append(["2026-07-07", "2026-05-01", "2026-06-01", provider, name,
                   cid, pay, amt])
    return _wb_bytes(wb)


def staff_roster_xlsx(sheet="ΨΥΧΙΑΤΡΟΙ", rows=None, months=None) -> bytes:
    """The monthly staff roster: one row per professional, one column per
    month holding the clinic placement — «√» meaning «unchanged since the
    last stated month», which is how the service maintains the file."""
    months = months or ["ΙΑΝΟΥΑΡΙΟΣ 2026", "ΦΕΒΡΟΥΑΡΙΟΣ 2026", "ΜΑΡΤΙΟΣ 2026",
                        "ΑΠΡΙΛΙΟΣ 2026", "ΜΑΙΟΣ 2026"]
    if rows is None:
        rows = [
            # (id, first, last, area, subarea, per-month cells)
            (849068, "ΧΡΥΣΤΑΛΛΑ", "ΣΚΟΡΔΗ", "Larnaca_Old_Hospital_MHS", "ΚΚΨΥΧΙΚΗΣ ΥΓΕΙΑ",
             ["3/5 Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ και 2/5 Ψ.Ν.Α (21)", "√", "√", "√", "√"]),
            (760193, "TZANIS", "LEONTARIDIS", "Limassol_General_Hospital", "ΨΥΧ.ΚΛΙ.ΛΕΜ.",
             ["5/5 Ε.Ι.Ψ.Υ.ΛΕΜΕΣΟΥ", "√", "√", "√", "√"]),
        ]
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(["Personal ID", "First Name", "Last Name", "Date of Birth",
               "Hiring Date", "Personnel Area", "Personnel Subarea",
               "Employee Subgroup"] + months)
    for pid, first, last, area, sub, cells in rows:
        ws.append([pid, first, last, "1980-01-01", "2010-01-01", area, sub,
                   "ΙΑΤΡ ΛΕΙΤΟΥΡ 1 ΤΑΞΗΣ"] + list(cells))
    return _wb_bytes(wb)


def cost_centre_map_xlsx(rows=None, with_hospital=False) -> bytes:
    """The optional SAP lookup finance maintains: clinic -> cost centre,
    internal order, and the Latin text SAP shows on the line.  The optional
    hospital column lets ONE file cover every payee."""
    rows = rows if rows is not None else [
        ("Ε.Ι.Ψ.Υ. ΛΑΡΝΑΚΑΣ", "1030300543", "13", "KOINOTIKI LARNAKAS", "", ""),
        ("Ψ.Ν.Α", "1030300507", "13", "COMM. MH SECTOR A", "", ""),
        ("Ε.Ι.Ψ.Υ.ΛΕΜΕΣΟΥ", "1030300533", "16", "KOINOTIKI LEMESOU", "", ""),
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "ΚΕΝΤΡΑ ΚΟΣΤΟΥΣ"
    head = ["Κλινική (clinic)", "Κέντρο κόστους", "Εσωτερική εντολή",
            "Κείμενο SAP", "Ειδικότητα"]
    if with_hospital:
        head.append("Νοσοκομείο (hospital)")
    ws.append(head)
    for r in rows:
        ws.append(list(r)[:len(head)])
    return _wb_bytes(wb)
