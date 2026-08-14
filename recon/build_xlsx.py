"""Output workbook (openpyxl) + gate-5 verification.

Formatting rules from the brief (non-negotiable):
- Blue font = hardcoded input off a source report.  Black = formula.
  Green = cross-sheet link.  Yellow fill = zero-check cells.
- Every subtotal, total, diff and check is a LIVE formula.
- SUMIFS criteria reference header/label cells, not quoted strings.
- Brand colours for headers: navy #062E5C, blue #0072BC, sky #00AEEF,
  green #8DC63F, gray #595959.

verify_workbook() reopens the built file and recomputes every yellow
zero-check cell with a small formula evaluator (openpyxl stores formulas,
it doesn't compute them) — gate 5.
"""
from __future__ import annotations

import io
import re
from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter, range_boundaries

from .checks import CENT, CheckPart as _Part, ReconResult, sra_sum
from .models import (Bucket, BUCKET_ORDER, HOSPITALS, is_hospital, MONTH_NAMES_EL,
                     norm_label,
                     PHARMACIST_FEE_UNIT_PRICE)
from .mapping import name_key as _name_key
from .numbers import format_eur

NAVY, BLUE, SKY, GREEN_BRAND, GRAY = "062E5C", "0072BC", "00AEEF", "8DC63F", "595959"
GREEN_LINK = "1F7A1F"

F_INPUT = Font(color=BLUE)                       # blue font = hardcoded input
F_FORMULA = Font(color="000000")                 # black = formula
F_LINK = Font(color=GREEN_LINK)                  # green = cross-sheet link
F_HEADER = Font(color="FFFFFF", bold=True)
F_RED = Font(color="C00000", bold=True)
F_AMBER = Font(color="B45F06", bold=True)
FILL_HEADER = PatternFill("solid", fgColor=NAVY)
FILL_SECTION = PatternFill("solid", fgColor=SKY)
FILL_CHECK = PatternFill("solid", fgColor="FFFF00")   # yellow = zero-check
FILL_AMBER = PatternFill("solid", fgColor="FFE599")
THIN = Border(bottom=Side(style="thin", color=GRAY))
EUR_FMT = "#,##0.00"


def _header(ws, row: int, labels: list[str]) -> None:
    for j, label in enumerate(labels, start=1):
        c = ws.cell(row=row, column=j, value=label)
        c.font = F_HEADER
        c.fill = FILL_HEADER
        c.alignment = Alignment(vertical="center")


def _amount(ws, row: int, col: int, value, font: Font):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    c.number_format = EUR_FMT
    return c


class _Section:
    """One provider inside a workbook: its result and its own SRA tab.  A
    hospital month has exactly one; a mental-health month has one per unit."""

    def __init__(self, label: str, result: ReconResult, sra_tab: Optional[str],
                 n_lines: int):
        self.label = label
        self.result = result
        self.sra_tab = sra_tab
        self.n_lines = n_lines


def build_workbook(result: ReconResult) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    bundle = result.bundle

    sra_tab = None
    stated_cell = None
    n_lines = 0
    if not result.crosscheck_mode and bundle.sra:
        sra_tab, total_row, stated_row, n_lines = _tab_sra(wb, result)
        stated_cell = f"'{sra_tab}'!F{stated_row}"
        _tab_reconciliation(wb, result, sra_tab, n_lines, stated_cell)
    else:
        _tab_matrix(wb, result)
    _tab_gl_bridge(wb, result, sra_tab)
    _tab_claims_bridge(wb, result, sra_tab, n_lines)
    sections = [_Section("", result, sra_tab, n_lines)]
    cc_rows = _tab_crosscheck(wb, sections)
    _tab_audit(wb, sections, cc_rows)
    split_total_row = _tab_split(wb, result, stated_cell)
    _tab_by_doctor(wb, result, sra_tab, n_lines, split_total_row)
    if bundle.sra and not result.crosscheck_mode:
        _tab_sap_upload(wb, [_Section("", result, sra_tab, n_lines)])
    _tab_truth_map(wb)
    _tab_legend(wb, result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_provider_workbook(entries: list) -> bytes:
    """Workbook for a NON-hospital month: several ΟΑΥ providers (the mental
    health units), each with its own cheque, reconciled in one run.

    entries: [(code, label, ReconResult), ...] in cheque order.  Tabs are a
    consolidated Σύνοψη, one SRA tab per provider, and the shared
    Source_crosscheck / Ανάλυση_ελέγχων / Legend, sectioned per provider."""
    wb = Workbook()
    wb.remove(wb.active)
    summary = wb.create_sheet("Σύνοψη_παρόχων")     # filled last: needs the tabs
    sections = []
    for code, label, result in entries:
        sra_tab, _total_row, stated_row, n_lines = _tab_sra(wb, result)
        sections.append(_Section(f"{label} ({code})", result, sra_tab, n_lines))
        sections[-1].code = code
        sections[-1].stated_row = stated_row
    cc_rows = _tab_crosscheck(wb, sections)
    _tab_audit(wb, sections, cc_rows)
    _tab_provider_by_doctor(wb, sections)
    _tab_by_clinic(wb, sections)
    _tab_sap_upload(wb, sections)
    _tab_legend(wb, entries[0][2] if entries else None)
    _tab_provider_summary(summary, sections)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# streams a non-hospital provider bills: everything else stays visible as an
# «adjustments» column rather than being dropped
_PROVIDER_STREAMS = [("OS", "Εξωτερικά ιατρεία (OS)"),
                     ("NM", "Νοσηλευτές/Μαίες (NM)"),
                     ("AP", "Επαγγελματίες υγείας (AP)")]


def _tab_provider_summary(ws, sections: list) -> None:
    """One row per provider: the cheque split by stream as live SUMIFS into
    that provider's own SRA tab, its claims and activity figures, and the
    differences — plus a grand total that must equal the sum of the cheques."""
    ws.cell(row=1, column=1,
            value="Σύνοψη παρόχων ΟΑΥ — μία γραμμή ανά πάροχο "
                  "(one row per provider, live off each SRA tab)"
            ).font = Font(bold=True, size=12, color=NAVY)
    heads = ["Πάροχος (Provider)", "Κωδικός", "Επιταγή (Cheque)"]
    heads += [lbl for _c, lbl in _PROVIDER_STREAMS]
    heads += ["Προσαρμογές (Adjustments)", "Σύνολο επιταγής (Cheque total)",
              "Claims «all»", "Διαφορά (Diff)", "Activity export", "Διαφορά (Diff)"]
    _header(ws, 3, heads)
    r = 4
    first = r
    for section in sections:
        res = section.result
        b = res.bundle
        ws.cell(row=r, column=1, value=section.label.rsplit(" (", 1)[0]).font = F_INPUT
        ws.cell(row=r, column=2, value=getattr(section, "code", "")).font = F_INPUT
        ws.cell(row=r, column=3, value=b.sra.cheque_no if b.sra else "").font = F_INPUT
        tab, n = section.sra_tab, section.n_lines
        col = 4
        stream_cols = []
        for code, _lbl in _PROVIDER_STREAMS:
            # criteria reference a helper cell on the SRA tab's own header row
            _amount(ws, r, col,
                    f"=SUMIFS('{tab}'!$F$2:$F${n},'{tab}'!$A$2:$A${n},$D$3)"
                    .replace("$D$3", f"{get_column_letter(col)}$2"), F_LINK)
            ws.cell(row=2, column=col, value=code).font = F_INPUT
            stream_cols.append(get_column_letter(col))
            col += 1
        # adjustments = the cheque minus the three streams: whatever ΟΑΥ paid
        # outside them stays visible instead of silently vanishing
        adj_col, total_col = col, col + 1
        _amount(ws, r, total_col,
                f"='{tab}'!F{getattr(section, 'stated_row', 2)}", F_LINK)
        _amount(ws, r, adj_col,
                f"={get_column_letter(total_col)}{r}-"
                + "-".join(f"{c}{r}" for c in stream_cols), F_FORMULA)
        claims_col, cdiff_col = total_col + 1, total_col + 2
        act_col, adiff_col = total_col + 3, total_col + 4
        claims = b.claims.total if b.claims else None
        if claims is not None:
            _amount(ws, r, claims_col, claims, F_INPUT)
            _amount(ws, r, cdiff_col,
                    f"={get_column_letter(claims_col)}{r}-("
                    + "+".join(f"{c}{r}" for c in stream_cols) + ")", F_FORMULA)
        # the export may span other cheques — use the cheque-gated figure the
        # cross-check already computed, so this Δ means the same thing
        act = None
        if b.xml_activity:
            gated = next((c.source_total for c in res.crosschecks
                          if "XML activity" in c.name), None)
            act = gated if gated is not None else b.xml_activity.total
        if act is not None:
            _amount(ws, r, act_col, act, F_INPUT)
            _amount(ws, r, adiff_col,
                    f"={get_column_letter(act_col)}{r}-("
                    + "+".join(f"{c}{r}" for c in stream_cols) + ")", F_FORMULA)
        r += 1
    total_row = r
    ws.cell(row=total_row, column=1, value="ΣΥΝΟΛΟ (all providers)").font = Font(bold=True)
    for c in range(4, 4 + len(_PROVIDER_STREAMS) + 6):
        letter = get_column_letter(c)
        _amount(ws, total_row, c, f"=SUM({letter}{first}:{letter}{total_row - 1})",
                F_FORMULA)
        ws.cell(row=total_row, column=c).font = Font(bold=True)
    # the streams + adjustments must add back to the cheques, live
    check_row = total_row + 1
    ws.cell(row=check_row, column=1,
            value="Zero-check = σύνολο ροών + προσαρμογές − επιταγές (must be 0)")
    stream_letters = [get_column_letter(4 + i) for i in range(len(_PROVIDER_STREAMS))]
    adj_letter = get_column_letter(4 + len(_PROVIDER_STREAMS))
    total_letter = get_column_letter(5 + len(_PROVIDER_STREAMS))
    _amount(ws, check_row, 5 + len(_PROVIDER_STREAMS),
            "=" + "+".join(f"{c}{total_row}" for c in stream_letters)
            + f"+{adj_letter}{total_row}-{total_letter}{total_row}", F_FORMULA)
    ws.cell(row=check_row, column=5 + len(_PROVIDER_STREAMS)).fill = FILL_CHECK
    _autosize(ws)


def _tab_provider_by_doctor(wb: Workbook, sections: list) -> None:
    """The posting sheet for a mental-health month: each unit's cheque split
    by speciality and by professional, off the paid-claims file's ASSOCIATED
    DOCTOR / DR SPECIALITY columns.

    Every unit block bridges from the per-doctor total to its cheque in live
    formulas: the claims-vs-SRA difference and the SRA lines outside the
    service streams are their OWN rows, so nothing is spread across doctors
    to force a tie."""
    ws = wb.create_sheet("Ανά_μονάδα_ιατρό")
    ws.cell(row=1, column=1,
            value="Κατανομή πληρωμών ανά μονάδα και ιατρό/επαγγελματία "
                  "(by unit and professional)").font = Font(bold=True, size=12,
                                                            color=NAVY)
    _header(ws, 3, ["Μονάδα (Unit)", "Ροή (Stream)", "Ειδικότητα (Speciality)",
                    "Ιατρός / Επαγγελματίας (Professional)", "Ποσό (Amount €)"])
    r = 4
    unit_cheque_cells = []
    unit_total_cells = []
    for section in sections:
        b = section.result.bundle
        tab, n = section.sra_tab, section.n_lines
        head = ws.cell(row=r, column=1, value=section.label)
        head.font = Font(bold=True, color="FFFFFF")
        for col in range(1, 6):
            ws.cell(row=r, column=col).fill = FILL_SECTION
        ws.cell(row=r, column=5,
                value=f"Επιταγή #{b.sra.cheque_no}" if b.sra else "").font =             Font(bold=True, color="FFFFFF")
        r += 1
        rows = list(b.claims.by_doctor) if b.claims else []
        subtotal_cells = []
        if rows:
            # clinic (speciality) first, then the professionals inside it
            groups: dict[tuple[str, str], list[tuple[str, float]]] = {}
            for seg, spec, doc, amt in rows:
                groups.setdefault((seg or "—", spec or "—"), []).append((doc or "—", amt))
            for (seg, spec), docs in sorted(groups.items(),
                                            key=lambda kv: -sum(a for _d, a in kv[1])):
                first = r
                for doc, amt in sorted(docs, key=lambda d: -d[1]):
                    ws.cell(row=r, column=2, value=seg).font = F_INPUT
                    ws.cell(row=r, column=3, value=spec).font = F_INPUT
                    ws.cell(row=r, column=4, value=doc).font = F_INPUT
                    _amount(ws, r, 5, amt, F_INPUT)
                    r += 1
                ws.cell(row=r, column=3, value=f"Υποσύνολο — {spec}").font = Font(bold=True)
                _amount(ws, r, 5, f"=SUM(E{first}:E{r - 1})", F_FORMULA)
                ws.cell(row=r, column=5).font = Font(bold=True)
                subtotal_cells.append(f"E{r}")
                r += 1
        else:
            ws.cell(row=r, column=2,
                    value="Το αρχείο claims δεν έχει στήλη ιατρού (no ASSOCIATED "
                          "DOCTOR column)").font = Font(italic=True, color=GRAY)
            r += 1
        claims_row = r
        ws.cell(row=claims_row, column=1,
                value="Σύνολο ανά ιατρό (claims file)").font = Font(bold=True)
        _amount(ws, claims_row, 5,
                ("=" + "+".join(subtotal_cells)) if subtotal_cells else 0.0,
                F_FORMULA if subtotal_cells else F_INPUT)
        r += 1
        # bridge to the cheque, every step live
        codes = [c for c, _l in _PROVIDER_STREAMS]
        code_cells = []
        for i, code in enumerate(codes):
            ws.cell(row=2, column=6 + i, value=code).font = F_INPUT
            code_cells.append(f"{get_column_letter(6 + i)}$2")
        svc = "+".join(f"SUMIFS('{tab}'!$F$2:$F${n},'{tab}'!$A$2:$A${n},{c})"
                       for c in code_cells)
        diff_row = r
        ws.cell(row=diff_row, column=1,
                value="Διαφορά claims έναντι γραμμών υπηρεσιών SRA (μη κατανεμημένη)")
        _amount(ws, diff_row, 5, f"={svc}-E{claims_row}", F_FORMULA)
        # a real gap between the claims file and the cheque's service lines:
        # shown on its own line, never spread across the professionals
        claims_total = b.claims.total if b.claims else 0.0
        services = round(sum(l.amount for l in b.sra.lines
                             if l.code in codes), 2) if b.sra else 0.0
        if abs(services - claims_total) > CENT:
            ws.cell(row=diff_row, column=5).font = F_AMBER
        r += 1
        adj_row = r
        ws.cell(row=adj_row, column=1,
                value="Λοιπές γραμμές SRA εκτός OS/NM/AP (προσαρμογές)")
        cheque_ref = f"'{tab}'!F{getattr(section, 'stated_row', 2)}"
        _amount(ws, adj_row, 5, f"={cheque_ref}-({svc})", F_FORMULA)
        r += 1
        cheque_row = r
        ws.cell(row=cheque_row, column=1, value="Επιταγή ΟΑΥ (HIO cheque)").font = Font(bold=True)
        _amount(ws, cheque_row, 5, f"={cheque_ref}", F_LINK)
        ws.cell(row=cheque_row, column=5).font = Font(bold=True, color=GREEN_LINK)
        r += 1
        check_row = r
        ws.cell(row=check_row, column=1,
                value="Zero-check = κατανομή + γέφυρα − επιταγή (must be 0)")
        _amount(ws, check_row, 5,
                f"=E{claims_row}+E{diff_row}+E{adj_row}-E{cheque_row}", F_FORMULA)
        ws.cell(row=check_row, column=5).fill = FILL_CHECK
        unit_cheque_cells.append(f"E{cheque_row}")
        unit_total_cells.append(f"E{claims_row}")
        r += 2
    if unit_cheque_cells:
        ws.cell(row=r, column=1, value="ΓΕΝΙΚΟ ΣΥΝΟΛΟ — κατανεμημένο ανά ιατρό"
                ).font = Font(bold=True)
        _amount(ws, r, 5, "=" + "+".join(unit_total_cells), F_FORMULA)
        r += 1
        ws.cell(row=r, column=1, value="ΓΕΝΙΚΟ ΣΥΝΟΛΟ — επιταγές").font = Font(bold=True)
        _amount(ws, r, 5, "=" + "+".join(unit_cheque_cells), F_FORMULA)
    _autosize(ws)


def build_sap_workbook(entries: list) -> bytes:
    """The SAP journal upload as its OWN one-sheet file — what finance
    actually feeds to SAP.  Identical content to the workbook's SAP_Upload
    tab; kept separate so nothing else travels with it."""
    wb = Workbook()
    wb.remove(wb.active)
    sections = []
    for code, label, result in entries:
        sections.append(_Section(f"{label} ({code})", result, None, 0))
        sections[-1].code = code
    info = _tab_sap_upload(wb, sections, inline_checks=False)
    _tab_sap_checks(wb, info)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _clinic_shares(sections: list) -> list:
    """Every professional's clinic share for the whole batch, unit by unit."""
    from .mapping import allocate_by_clinic
    out = []
    for section in sections:
        b = section.result.bundle
        if not b.claims:
            continue
        out += allocate_by_clinic(b.claims.by_doctor, getattr(b, "staff", None),
                                  unit_label=section.label)
    return out


def _tab_by_clinic(wb: Workbook, sections: list) -> None:
    """The clinic split ΟΑΥ's files cannot give: the mental-health service
    posts by CLINIC, but ΟΑΥ pays by unit, so each professional's amount is
    re-split across the clinics the monthly roster puts them in.

    A professional the roster does not cover keeps the whole amount in an
    «unmapped» block — visible, never spread across clinics."""
    from .mapping import clinic_key
    ws = wb.create_sheet("Ανά_κλινική")
    ws.cell(row=1, column=1,
            value="Κατανομή ανά κλινική βάσει μητρώου προσωπικού "
                  "(by clinic, from the monthly staff roster)"
            ).font = Font(bold=True, size=12, color=NAVY)
    shares = _clinic_shares(sections)
    staff = next((getattr(s.result.bundle, "staff", None) for s in sections
                  if getattr(s.result.bundle, "staff", None)), None)
    if not staff:
        ws.cell(row=2, column=1,
                value="Δεν ανέβηκε μητρώο προσωπικού — ανεβάστε το μηνιαίο αρχείο "
                      "«Personal ID / First Name / Last Name / <μήνας>» για να "
                      "γίνει η κατανομή ανά κλινική (no staff roster uploaded)."
                ).font = Font(italic=True, color=GRAY)
        _autosize(ws)
        return
    _header(ws, 3, ["Κλινική (Clinic)", "Μονάδα ΟΑΥ (Unit / cheque)",
                    "Ειδικότητα (Speciality)", "Ιατρός / Επαγγελματίας",
                    "Ποσοστό (Share)", "Ποσό (Amount €)", "Σημείωση (Note)"])
    r = 4
    groups: dict[str, list] = {}
    for sh in shares:
        groups.setdefault(clinic_key(sh.clinic), []).append(sh)
    ordered = sorted(groups.values(),
                     key=lambda g: (not g[0].matched, -sum(x.amount for x in g),
                                    g[0].clinic))
    subtotal_cells = []
    for group in ordered:
        title = ws.cell(row=r, column=1, value=group[0].clinic)
        title.font = Font(bold=True, color="FFFFFF")
        for col in range(1, 8):
            ws.cell(row=r, column=col).fill = FILL_SECTION
        r += 1
        first = r
        # tie-break on the ASCII name key, not the Greek string: locale
        # collation differs between the two ports, the key does not
        for sh in sorted(group, key=lambda x: (-x.amount,
                                               " ".join(_name_key(x.professional)),
                                               x.unit)):
            ws.cell(row=r, column=2, value=sh.unit).font = F_INPUT
            ws.cell(row=r, column=3, value=sh.speciality).font = F_INPUT
            ws.cell(row=r, column=4, value=sh.professional).font = F_INPUT
            pct = ws.cell(row=r, column=5, value=round(sh.weight, 4))
            pct.number_format = "0.0%"
            pct.font = F_INPUT
            _amount(ws, r, 6, sh.amount, F_INPUT)
            if sh.note:
                note = ws.cell(row=r, column=7, value=sh.note)
                note.font = Font(italic=True, color=GRAY)
                if not sh.matched:
                    note.fill = FILL_AMBER
            r += 1
        ws.cell(row=r, column=1, value="Υποσύνολο κλινικής").font = Font(bold=True)
        _amount(ws, r, 6, f"=SUM(F{first}:F{r - 1})", F_FORMULA)
        ws.cell(row=r, column=6).font = Font(bold=True)
        subtotal_cells.append(f"F{r}")
        r += 2
    total_row = r
    ws.cell(row=total_row, column=1,
            value="ΓΕΝΙΚΟ ΣΥΝΟΛΟ κατανομής (all clinics)").font = Font(bold=True)
    _amount(ws, total_row, 6, "=" + "+".join(subtotal_cells) if subtotal_cells else 0.0,
            F_FORMULA)
    r += 1
    ws.cell(row=r, column=1,
            value="Σύνολο claims των μονάδων (claims files)").font = Font(bold=True)
    claims_total = round(sum(s.result.bundle.claims.total for s in sections
                             if s.result.bundle.claims), 2)
    _amount(ws, r, 6, claims_total, F_INPUT)
    r += 1
    ws.cell(row=r, column=1,
            value="Zero-check = κατανομή − claims (must be 0)")
    _amount(ws, r, 6, f"=F{total_row}-F{r - 1}", F_FORMULA)
    ws.cell(row=r, column=6).fill = FILL_CHECK
    r += 2
    unmapped = round(sum(x.amount for x in shares if not x.matched), 2)
    if unmapped:
        ws.cell(row=r, column=1, value=(
            f"Προσοχή: {format_eur(unmapped)} δεν κατανεμήθηκε σε κλινική — "
            "επαγγελματίες εκτός μητρώου. Ανεβάστε και τα μητρώα των υπόλοιπων "
            "ειδικοτήτων (professionals with no roster row; upload the rosters "
            "for the remaining professions).")).font = F_AMBER
    _autosize(ws)


# The SAP journal template, reproduced from the service's own workbook
# («JOURNAL ENTRIES» sheet): three header rows — a group row, the descriptive
# row finance reads, and the technical BKPF/BSEG field names — then the data.
_SAP_COLUMNS = [
    ("BKPF-BLDAT", "Document Date (8)"),
    ("BKPF-BUDAT", "Posting Date (8)"),
    ("BKPF-BLART", "Document type (2)\n(KR invoice, KG-Vendor Credit memo etc)"),
    ("BKPF-BUKRS", "Hospital  (4)\n1000, 1010 etc"),
    ("BKPF-WAERS", "Currency (5)\ne.g. EUR"),
    ("BKPF-MONAT", "Period (2)"),
    ("BKPF-XBLNR", "Reference (16)\neg. Vendor invoice"),
    ("BKPF-BKTXT", "Header Text (25)"),
    ("BSEG-BSCHL", "Posting key (2)\neg. 31 vendor invoice\n40 - debit expense\n"
                   "70 - debit asset"),
    ("BSEG-HKONT/KUNNR/LIFNR/ANLN1/ANLN2",
     "Account (17)\ne.g. 535320 (expense)\n102020 (Vendor account)\n"
     "10300000140 (Asset)"),
    ("BSEG-ANBWA", "Asset Transaction type Pruchase - 100"),
    ("BSEG-DMBTR", "Amount in document currency (16)\ne.g. 100,20"),
    ("BSEG-MWSKZ", "Tax cde (2)\ne.g. 19"),
    ("BSEG-KOSTL", "Cost Center (10)\ne.g. 202016"),
    ("BSEG-AUFNR", "Internal order (12)\ne.g. 13"),
    ("BSEG-GEBER", "Fund (10)  e.g. 100000"),
    ("BSEG-FISTL", "fund center (16)\ne.g. 1.07.00"),
    ("BSEG-FIPOS", "Commitment item (24)\ne.g. 03433"),
    ("BSEG-ZUONR", "Assignment (18)\ne.g. 2000"),
    ("BSEG-SGTXT", "Text (40)\ne.g. ηλεκτρολογικά υλικά covid"),
    ("BSEG-XREF1", "XREF1"), ("BSEG-XREF2", "XREF2"), ("BSEG-XREF3", "XREF3"),
    ("", ""),          # helper: the remittance advice the document posts
    ("", ""),          # helper: whose money the line is
]
SAP_SHEET = "JOURNAL ENTRIES"
SAP_DEFAULTS = {"doc_type": "SA", "company": "1003", "currency": "EUR",
                "debit_key": "01", "debit_account": "200000",
                "credit_key": "50", "credit_account": "412002", "tax": "O0"}


def _sap_header(ws) -> None:
    """The template's three header rows, verbatim — data starts on row 4."""
    ws.cell(row=1, column=1, value="Header Data").font = Font(bold=True)
    ws.cell(row=1, column=9, value="Line item Data").font = Font(bold=True)
    ws.cell(row=1, column=24,
            value="Remittance advice").font = Font(bold=True)
    ws.cell(row=1, column=25,
            value="Ανάλυση (Professional / stream)").font = Font(bold=True)
    for j, (tag, label) in enumerate(_SAP_COLUMNS, start=1):
        c = ws.cell(row=2, column=j, value=label)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        t = ws.cell(row=3, column=j, value=tag)
        t.font = Font(color="FFFFFF", bold=True)
        t.fill = PatternFill("solid", fgColor=NAVY)
    ws.freeze_panes = "A4"


def _journal_lines(section) -> tuple[list[dict], dict]:
    """A month's credit lines for one payee, in the service's own journal
    order.

    A hospital and a mental-health unit post the same cheque differently, so
    the lines come from different places — but the journal LAYOUT is one, and
    finance uploads one kind of file either way."""
    b = section.result.bundle
    if is_hospital(b.hospital_code):
        return _journal_lines_by_stream(section)
    return _journal_lines_by_professional(section)


# A By_Clinic_Split line -> what it IS, which decides both the HIO revenue
# account and which flavour of the clinic's cost centre it posts to.
_LINE_KINDS = [
    ("ΠΟΙΟΤΙΚΑ", "quality", "general"),
    ("ΣΤΑΘΕΡΕΣ ΧΡΕΩΣΕΙΣ", "oncall", "general"),
    ("ΕΜΒΟΛΙΑΣΜ", "vaccines", "general"),
    ("ΚΑΤΑ ΚΕΦΑΛΗΝ", "capitation", "general"),
    ("CAPITATION", "capitation", "general"),
]
_BUCKET_KINDS = {
    "Inpatient": ("inpatient_drg", "ward"),
    "A&E": ("ae", "general"),
    "Outpatient": ("outpatient", "clinic"),
    "Pharma": ("pharma", "general"),
}
# «Ειδικοί Ιατροί — GASTROENTEROLOGY (OS)» -> GASTROENTEROLOGY
_SPEC_IN_LABEL = re.compile(r"[—\-]\s*([A-Z][A-Z &/'\-]{3,})\s*(?:\(|$)")


def _line_kind(label: str, bucket: str) -> tuple[str, str]:
    up = norm_label(label)
    for needle, kind, variant in _LINE_KINDS:
        if needle in up:
            return kind, variant
    return _BUCKET_KINDS.get(bucket, ("outpatient", "general"))


def _specialty_of(label: str) -> str:
    """The ΟΑΥ speciality a split line belongs to: inpatient rows ARE the
    clinic name, outpatient rows carry it after the dash."""
    m = _SPEC_IN_LABEL.search(str(label))
    return (m.group(1) if m else str(label)).strip()


def _row_parts(row, kind: str, variant: str) -> list[tuple[float, str, str]]:
    """(amount, account key, cost-centre flavour) for one split row.

    An inpatient clinic row is three different things at once — DRG, daily
    treatments and Z-catalogue items — and SAP keeps a separate revenue
    account for each, so it becomes up to three credit lines that still add
    back to the row."""
    if kind != "inpatient_drg":
        return [(row.amount, kind, variant)]
    three = [(row.drg or 0.0, "inpatient_drg", "ward"),
             (row.fixed_fee or 0.0, "inpatient_daily", "daycare"),
             (row.z_drugs or 0.0, "inpatient_z", "daycare")]
    if round(sum(a for a, _k, _v in three), 2) != round(row.amount, 2):
        return [(row.amount, kind, variant)]     # no split on this row
    return [(a, k, v) for a, k, v in three if a]


def _journal_lines_by_stream(section) -> tuple[list[dict], dict]:
    """A HOSPITAL month: the credit lines are the By_Clinic_Split rows, so one
    document carries every revenue stream of the month — inpatient by clinic,
    ΤΑΕΠ, outpatient by speciality, personal doctors, pharma and the
    adjustment lines — and its total is the same figure that sheet ties to the
    cheque.

    Codes come from OKYπY's own SAP master when it is uploaded: the revenue
    account per stream (412001 in-patient, 412005 day care, 412007 catalogue
    Z, 412003 ΤΑΕΠ …) and the clinic's cost centre inside this hospital's
    company code, picked by flavour — ward for DRG, ημερήσια φροντίδα for
    daily treatments, εξωτερικά ιατρεία for the outpatient specialists.  A
    hand-kept lookup still wins where it names a line, and anything neither
    can resolve stays blank and is listed."""
    from .sapmaster import company_for
    b = section.result.bundle
    lookup = getattr(b, "cost_centres", None)
    master = getattr(b, "sap", None)
    code = b.hospital_code or ""
    company = company_for(code) if master else ""
    out: list[dict] = []
    missing: dict[str, str] = {}
    for sec in section.result.split:
        stream = sec.bucket.value if sec.bucket else sec.title
        for row in sec.rows:
            if not row.amount:
                continue
            kind, variant = _line_kind(row.label, stream)
            for amount, part_kind, part_variant in _row_parts(row, kind, variant):
                hit = lookup.find(row.label, stream, code) if lookup else None
                if lookup and (hit is None or not hit.cost_centre):
                    # a row keyed on the BUCKET codes every line in that
                    # bucket — four rows per hospital post at stream level
                    hit = (lookup.find(stream, "", code)
                           or lookup.find(sec.title, "", code) or hit)
                kostl = hit.cost_centre if hit else ""
                # a hospital posts no internal order — that column belongs to
                # the mental-health professional categories (11-16)
                aufnr = ""
                text = hit.text if hit and hit.text else ""
                centre = None
                if master and not kostl:
                    # the line's own speciality first, then the stream it
                    # belongs to («Αναλώσιμα» is still pharmacy)
                    centre = (master.find_centre(company, _specialty_of(row.label),
                                                 part_variant)
                              or master.find_centre(company, stream, part_variant))
                    if centre:
                        kostl, text = centre.code, centre.name
                account, _atext = master.account(part_kind) if master else ("", "")
                out.append({"kostl": kostl, "aufnr": aufnr,
                            "text": text or row.label, "account": account,
                            "professional": stream, "amount": round(amount, 2)})
                if not kostl:
                    missing[row.label] = round(
                        missing.get(row.label, 0.0) + amount, 2)
    # the split already ties to the cheque with its own zero-check; anything
    # left is still shown rather than absorbed
    residual = round(b.sra.stated_total - sum(x["amount"] for x in out), 2)
    if abs(residual) > 0.005:
        out.append({"kostl": "", "aufnr": "", "account": "",
                    "text": "TO CLASSIFY (split vs SRA)",
                    "professional": "", "amount": residual})
    return out, missing


def _journal_lines_by_professional(section) -> tuple[list[dict], dict]:
    """A MENTAL-HEALTH unit: credit lines ordered by cost centre, internal
    order, then professional — the order the service's own journal uses.

    The amount column is broken down BY PROFESSIONAL, the same figures the
    «Anά_μονάδα_ιατρό» sheet shows, re-split across the clinics the roster
    puts each professional in.  A professional working two clinics therefore
    appears once per clinic, and their lines add back to that sheet's total.

    Whatever the clinic split does not cover (claims vs SRA, adjustment lines)
    becomes its own TO CLASSIFY line rather than being spread over the
    clinics, so the document still posts the whole remittance advice and the
    unallocated part stays visible."""
    from .mapping import clinic_key, name_key
    b = section.result.bundle
    lookup = getattr(b, "cost_centres", None)
    buckets: dict[tuple, float] = {}
    labels: dict[tuple, tuple] = {}
    missing: dict[str, str] = {}
    for sh in _clinic_shares([section]):
        code = b.hospital_code or ""
        row = lookup.find(sh.clinic, sh.speciality, code) if lookup else None
        kostl = row.cost_centre if row else ""
        aufnr = row.internal_order if row else ""
        if lookup and not aufnr:
            # the internal order belongs to the professional category, so a
            # speciality-only row in the lookup may carry it
            alt = lookup.find_speciality(sh.speciality, code)
            aufnr = alt.internal_order if alt else ""
        text = row.text if row and row.text else sh.clinic
        key = (kostl, aufnr, clinic_key(sh.clinic),
               sh.speciality if not kostl else "", str(name_key(sh.professional)))
        buckets[key] = round(buckets.get(key, 0.0) + sh.amount, 2)
        labels[key] = (text, sh.professional)
        if not kostl:
            missing[sh.clinic] = round(missing.get(sh.clinic, 0.0) + sh.amount, 2)
    residual = round(b.sra.stated_total - sum(buckets.values()), 2)
    out = []
    for key in sorted(buckets, key=lambda k: (k[0] == "", k[0], k[1], k[2], k[4])):
        kostl, aufnr = key[0], key[1]
        text, professional = labels[key]
        out.append({"kostl": kostl, "aufnr": aufnr, "text": text,
                    "professional": professional, "amount": buckets[key]})
    if abs(residual) > 0.005:
        out.append({"kostl": "", "aufnr": "",
                    "text": "TO CLASSIFY (claims vs SRA + adj.)",
                    "professional": "", "amount": residual})
    return out, missing


def _tab_sap_upload(wb: Workbook, sections: list,
                    inline_checks: bool = True) -> dict:
    """The month's postings in the service's own SAP journal layout: one
    document per remittance advice — a debit line (posting key 01, account
    200000) whose amount is a live SUM of the credit lines beneath it, then
    one credit line (50 / 412002 / O0) per cost centre × internal order.

    Cost centre and internal order come from the uploaded lookup. Without it
    the lines are still written, keyed by clinic name, with those two columns
    blank and highlighted, and every clinic that needs a code listed on the
    check sheet — the app never invents an account code."""
    from .mapping import month_label
    ws = wb.create_sheet(SAP_SHEET)
    _sap_header(ws)
    from .sapmaster import company_for
    lookup = next((getattr(s.result.bundle, "cost_centres", None) for s in sections
                   if getattr(s.result.bundle, "cost_centres", None)), None)
    b0 = sections[0].result.bundle if sections else None
    master = getattr(b0, "sap", None) if b0 else None
    company = (lookup.company_code if lookup and lookup.company_code
               else (company_for(b0.hospital_code) if master and b0
                     else SAP_DEFAULTS["company"]))
    year, month = (b0.year, b0.month) if b0 else (None, None)
    doc_date = (f"{_month_end(year, month):02d}.{month:02d}.{year}"
                if year and month else "")
    period_label = month_label(year, month)
    short = f"{month:02d}/{str(year)[-2:]}" if year and month else ""
    r = 4
    missing: dict[str, str] = {}
    docs: list[tuple] = []          # (cheque, unit label, head row, total)
    for section in sections:
        b = section.result.bundle
        if not b.sra:
            continue
        cheque = b.sra.cheque_no
        lines, miss = _journal_lines(section)
        for label, amount in miss.items():
            missing[label] = round(missing.get(label, 0.0) + amount, 2)
        head_row = r
        # header (debit) line — the amount is the live sum of its own credits
        head = [doc_date, f"=A{head_row}", SAP_DEFAULTS["doc_type"], company,
                SAP_DEFAULTS["currency"], "", period_label,
                f'="HIO OUTP. INV."&X{head_row}', SAP_DEFAULTS["debit_key"],
                SAP_DEFAULTS["debit_account"], "",
                f"=SUM(L{head_row + 1}:L{head_row + len(lines)})", "", "",
                "", "", "", "", company,
                f'="HIO OUTP. {short} INV."&X{head_row}',
                "", "", "", cheque, ""]
        for j, v in enumerate(head, start=1):
            c = ws.cell(row=r, column=j, value=v)
            c.font = F_INPUT if not str(v).startswith("=") else F_FORMULA
            if j == 12:
                c.number_format = EUR_FMT
        r += 1
        for ln in lines:
            sgtxt = f'="HIO OUTP. {short} INV."&X{r}&" {_q(ln["text"])}"'
            line = ["", "", "", "", "", "", "", "", SAP_DEFAULTS["credit_key"],
                    ln.get("account") or SAP_DEFAULTS["credit_account"],
                    "", ln["amount"],
                    SAP_DEFAULTS["tax"], ln["kostl"], ln["aufnr"], "", "", "",
                    company, sgtxt, "", "", "", cheque, ln["professional"]]
            for j, v in enumerate(line, start=1):
                c = ws.cell(row=r, column=j, value=v)
                c.font = F_FORMULA if str(v).startswith("=") else F_INPUT
                if j == 12:
                    c.number_format = EUR_FMT
                if j in (14, 15) and not v:
                    c.fill = FILL_AMBER      # code still to be filled in
            r += 1
        docs.append((cheque, section.label, head_row, b.sra.stated_total))
    info = {"last": r - 1, "docs": docs, "missing": missing,
            "master_seen": master is not None}
    if inline_checks:
        _sap_checks(ws, info, r + 1)
    _autosize(ws)
    for j in range(1, len(_SAP_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(j)].width = min(
            ws.column_dimensions[get_column_letter(j)].width or 12, 26)
    return info


def _q(text: str) -> str:
    """A cost-centre name going inside a formula string literal."""
    return str(text).replace('"', "'")[:32]


def _sap_checks(ws, info: dict, row: int) -> int:
    """Credits = debits, and every document = its own remittance advice."""
    last = info["last"]
    r = row
    ws.cell(row=r, column=1,
            value="Σύνολο πιστωτικών γραμμών (credit lines)").font = Font(bold=True)
    ws.cell(row=r, column=9, value=SAP_DEFAULTS["credit_key"]).font = F_INPUT
    _amount(ws, r, 12, f"=SUMIFS(L4:L{last},I4:I{last},I{r})", F_FORMULA)
    ws.cell(row=r + 1, column=1,
            value="Σύνολο χρεωστικών γραμμών (debit lines)").font = Font(bold=True)
    ws.cell(row=r + 1, column=9, value=SAP_DEFAULTS["debit_key"]).font = F_INPUT
    _amount(ws, r + 1, 12, f"=SUMIFS(L4:L{last},I4:I{last},I{r + 1})", F_FORMULA)
    ws.cell(row=r + 2, column=1,
            value="Zero-check = πιστωτικές − χρεωστικές (must be 0)")
    _amount(ws, r + 2, 12, f"=L{r}-L{r + 1}", F_FORMULA)
    ws.cell(row=r + 2, column=12).fill = FILL_CHECK
    r += 4
    _header(ws, r, ["Επιταγή / remittance advice", "Ανάρτηση (posted) €",
                    "Επιταγή ΟΑΥ (advice total) €", "Διαφορά (Diff) €"])
    r += 1
    for cheque, label, head_row, stated in info["docs"]:
        ws.cell(row=r, column=1, value=f"{cheque} — {label}".strip(" —")).font = F_INPUT
        _amount(ws, r, 2, f"=L{head_row}", F_LINK)
        _amount(ws, r, 3, stated, F_INPUT)
        _amount(ws, r, 4, f"=B{r}-C{r}", F_FORMULA)
        ws.cell(row=r, column=4).fill = FILL_CHECK
        r += 1
    text = _missing_note(info, info.get("master_seen", False))
    if text:
        r += 1
        note = ws.cell(row=r, column=1, value=text)
        note.font = F_AMBER
        note.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return r


def _missing_note(info: dict, master_seen: bool) -> Optional[str]:
    """The alert: which lines carry money the app could not code, and what
    each is worth.  A line with nothing allocated to it is not a problem, so
    it is not reported."""
    worth = {k: v for k, v in info["missing"].items() if abs(v) > 0.005}
    if not worth:
        return None
    named = " · ".join(f"{k} — {format_eur(v)}" for k, v in
                       sorted(worth.items(), key=lambda kv: -abs(kv[1])))
    head = ("Γραμμές με ποσό αλλά χωρίς κέντρο κόστους — συμπληρώστε τα στο "
            "αρχείο αντιστοίχισης και ανεβάστε το ξανά (lines carrying an "
            "amount with no cost centre): ")
    if not master_seen:
        head = ("ΔΕΝ ανέβηκαν τα βασικά δεδομένα SAP: ανεβάστε το "
                "Chart_of_Accounts.xlsx μαζί με τα αρχεία του μήνα και οι "
                "περισσότερες από αυτές τις γραμμές θα κωδικοποιηθούν μόνες "
                "τους (the SAP master was not uploaded). Γραμμές με ποσό "
                "χωρίς κέντρο κόστους: ")
    return head + named


def _tab_sap_checks(wb: Workbook, info: dict) -> None:
    """The upload sheet stays clean — finance selects it whole and feeds it to
    SAP — so the checks live on their own sheet, pointing back at it."""
    ws = wb.create_sheet("Έλεγχος_SAP")
    ws.cell(row=1, column=1,
            value="Έλεγχοι της ανάρτησης SAP (checks on the journal above)"
            ).font = Font(bold=True, size=12, color=NAVY)
    last = info["last"]
    r = 3
    ws.cell(row=r, column=1,
            value="Σύνολο πιστωτικών γραμμών (credit lines)").font = Font(bold=True)
    _amount(ws, r, 2,
            f"=SUMIFS('{SAP_SHEET}'!L4:L{last},'{SAP_SHEET}'!I4:I{last},C{r})",
            F_LINK)
    ws.cell(row=r, column=3, value=SAP_DEFAULTS["credit_key"]).font = F_INPUT
    ws.cell(row=r + 1, column=1,
            value="Σύνολο χρεωστικών γραμμών (debit lines)").font = Font(bold=True)
    _amount(ws, r + 1, 2,
            f"=SUMIFS('{SAP_SHEET}'!L4:L{last},'{SAP_SHEET}'!I4:I{last},C{r + 1})",
            F_LINK)
    ws.cell(row=r + 1, column=3, value=SAP_DEFAULTS["debit_key"]).font = F_INPUT
    ws.cell(row=r + 2, column=1,
            value="Zero-check = πιστωτικές − χρεωστικές (must be 0)")
    _amount(ws, r + 2, 2, f"=B{r}-B{r + 1}", F_FORMULA)
    ws.cell(row=r + 2, column=2).fill = FILL_CHECK
    r += 4
    _header(ws, r, ["Επιταγή / remittance advice", "Ανάρτηση (posted) €",
                    "Επιταγή ΟΑΥ (advice total) €", "Διαφορά (Diff) €"])
    r += 1
    for cheque, label, head_row, stated in info["docs"]:
        ws.cell(row=r, column=1, value=f"{cheque} — {label}".strip(" —")).font = F_INPUT
        _amount(ws, r, 2, f"='{SAP_SHEET}'!L{head_row}", F_LINK)
        _amount(ws, r, 3, stated, F_INPUT)
        _amount(ws, r, 4, f"=B{r}-C{r}", F_FORMULA)
        ws.cell(row=r, column=4).fill = FILL_CHECK
        r += 1
    text = _missing_note(info, info.get("master_seen", False))
    if text:
        r += 1
        note = ws.cell(row=r, column=1, value=text)
        note.font = F_AMBER
        note.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws)


def _month_end(year: Optional[int], month: Optional[int]) -> int:
    if not year or not month:
        return 1
    if month == 2:
        return 29 if (year % 4 == 0 and (year % 100 != 0 or year % 400 == 0)) else 28
    return 30 if month in (4, 6, 9, 11) else 31


# ------------------------------------------------------------- tab 1: SRA

def _tab_sra(wb: Workbook, result: ReconResult):
    sra = result.bundle.sra
    name = f"SRA_{sra.cheque_no}"[:31]
    ws = wb.create_sheet(name)
    _header(ws, 1, ["Κωδικός (Code)", "Περιγραφή (Description)", "Κανάλι (Channel)",
                    "Κατηγορία (Bucket)", "Πηγή ΟΑΥ (Source report)", "Ποσό (Amount €)",
                    "Επιταγή (Cheque)"])
    r = 2
    for line in sra.lines:
        ws.cell(row=r, column=1, value=line.code).font = F_INPUT
        ws.cell(row=r, column=2, value=line.description).font = F_INPUT
        ws.cell(row=r, column=3, value=line.channel).font = F_INPUT
        ws.cell(row=r, column=4, value=line.bucket.value).font = F_INPUT
        ws.cell(row=r, column=5, value=line.source_report).font = F_INPUT
        _amount(ws, r, 6, line.amount, F_INPUT)
        # the paying cheque, so a check can be restricted to the same cheques
        # a source file covers (SUMIFS second criteria pair)
        ws.cell(row=r, column=7, value=line.cheque or sra.cheque_no).font = F_INPUT
        r += 1
    last_line = r - 1
    total_row = r
    ws.cell(row=total_row, column=1, value="TOTAL (ΣΥΝΟΛΟ)").font = Font(bold=True)
    _amount(ws, total_row, 6, f"=SUM(F2:F{last_line})", F_FORMULA)
    ws.cell(row=total_row, column=6).font = Font(bold=True)
    r += 1
    # one stated row per cheque; several cheques get a live stated TOTAL
    parts = sra.parts if len(sra.parts) > 1 else \
        [(sra.cheque_no, sra.lines_total, sra.stated_total)]
    first_part_row = r
    for cheque, _lines_total, stated in parts:
        ws.cell(row=r, column=1,
                value=f"Δηλωμένο σύνολο επιταγής (stated cheque total) #{cheque}"
                ).font = F_INPUT
        _amount(ws, r, 6, stated, F_INPUT)
        r += 1
    if len(parts) > 1:
        stated_row = r
        ws.cell(row=stated_row, column=1,
                value="Δηλωμένο σύνολο όλων των επιταγών (all cheques)").font = Font(bold=True)
        _amount(ws, stated_row, 6, f"=SUM(F{first_part_row}:F{r - 1})", F_FORMULA)
        r += 1
    else:
        stated_row = first_part_row
    check_row = r
    ws.cell(row=check_row, column=1, value="Check = TOTAL − stated (must be 0)")
    _amount(ws, check_row, 6, f"=F{total_row}-F{stated_row}", F_FORMULA)
    ws.cell(row=check_row, column=6).fill = FILL_CHECK
    _autosize(ws)
    return name, total_row, stated_row, last_line


# ----------------------------------------------------- tab 2: Reconciliation

def _tab_reconciliation(wb: Workbook, result: ReconResult, sra_tab: str,
                        n_lines: int, stated_cell: str) -> None:
    ws = wb.create_sheet("Reconciliation")
    b = result.bundle
    hosp = HOSPITALS[b.hospital_code]
    ws.cell(row=1, column=1, value=f"{hosp[0]} ({hosp[1]}) — {MONTH_NAMES_EL[b.month]} "
                                   f"{b.year} — Επιταγή #{b.sra.cheque_no}").font = Font(bold=True, color=NAVY)
    _header(ws, 3, ["Κατηγορία (Bucket)", "Bucket key", "Ποσό (Amount €)"])
    labels = {
        Bucket.INPATIENT: "Ενδονοσοκομειακή περίθαλψη (Inpatient)",
        Bucket.AE: "ΤΑΕΠ (A&E)",
        Bucket.OUTPATIENT: "Εξωνοσοκομειακή περίθαλψη (Outpatient)",
        Bucket.PHARMA: "Φάρμακα (Pharma)",
    }
    r = 4
    for bucket in BUCKET_ORDER:
        ws.cell(row=r, column=1, value=labels[bucket]).font = F_INPUT
        ws.cell(row=r, column=2, value=bucket.value).font = F_INPUT
        # live SUMIFS on the SRA tab's Bucket column, criteria = the label cell
        _amount(ws, r, 3,
                f"=SUMIFS('{sra_tab}'!$F$2:$F${n_lines},'{sra_tab}'!$D$2:$D${n_lines},$B{r})",
                F_FORMULA)
        r += 1
    total_row = r
    ws.cell(row=total_row, column=1, value="TOTAL (ΣΥΝΟΛΟ)").font = Font(bold=True)
    _amount(ws, total_row, 3, f"=SUM(C4:C{r - 1})", F_FORMULA)
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    cheque_row = total_row + 1
    ws.cell(row=cheque_row, column=1, value="Επιταγή ΟΑΥ (HIO cheque)")
    _amount(ws, cheque_row, 3, f"={stated_cell}", F_LINK)
    check_row = cheque_row + 1
    ws.cell(row=check_row, column=1, value="Zero-check = TOTAL − cheque (must be 0)")
    _amount(ws, check_row, 3, f"=C{total_row}-C{cheque_row}", F_FORMULA)
    ws.cell(row=check_row, column=3).fill = FILL_CHECK
    _autosize(ws)


# ------------------------------------------- tab 2 (cross-check mode): matrix

def _tab_matrix(wb: Workbook, result: ReconResult) -> None:
    ws = wb.create_sheet("Crosscheck_Matrix")
    b = result.bundle
    hosp = HOSPITALS[b.hospital_code]
    ws.cell(row=1, column=1,
            value=f"{hosp[0]} ({hosp[1]}) — {MONTH_NAMES_EL[b.month]} {b.year} — "
                  "Cross-check mode (χωρίς SRA / no SRA)").font = Font(bold=True, color=NAVY)
    cols = result.matrix_columns
    _header(ws, 3, ["Ροή (Stream)"] + cols + ["Range (max−min)"])
    r = 4
    for row in result.matrix:
        ws.cell(row=r, column=1, value=row["stream"]).font = F_INPUT
        populated = []
        for j, col in enumerate(cols, start=2):
            v = row["values"].get(col)
            if v is not None:
                _amount(ws, r, j, v, F_INPUT)
                populated.append(get_column_letter(j) + str(r))
        rng_col = len(cols) + 2
        if len(populated) > 1:
            first, last = get_column_letter(2), get_column_letter(len(cols) + 1)
            _amount(ws, r, rng_col,
                    f"=MAX({first}{r}:{last}{r})-MIN({first}{r}:{last}{r})", F_FORMULA)
            if row["range"] is not None and abs(row["range"]) > 0.5:
                ws.cell(row=r, column=rng_col).font = F_AMBER
        r += 1
    _autosize(ws)


# ------------------------------------------------ tab 3: Source_crosscheck

def _tab_crosscheck(wb: Workbook, sections: list) -> dict:
    """Returns {(section index, check index): row} so the audit tab can tie
    each of its blocks back to the exact row printed here."""
    ws = wb.create_sheet("Source_crosscheck")
    # column names follow the CHECK NAME order: A = the first thing named,
    # B = the second.  (A is not always "the source report" — on GL rows A is
    # the ΟΑΥ ledger and B the report it is compared with.)
    _header(ws, 1, ["Έλεγχος: Α = Β (Check)", "Α — ποσό πρώτης πηγής (Amount A €)",
                    "Β — ποσό δεύτερης πηγής / SRA (Amount B €)",
                    "Διαφορά Α−Β (Diff €)", "Σημείωση (Note)",
                    "Συσκευασίες (Packages)", "Τιμή μονάδας (Unit €)",
                    "Κωδικοί SRA (codes)"])
    r = 2
    cc_rows: dict = {}
    for si, section in enumerate(sections):
        result, sra_tab, n_lines = section.result, section.sra_tab, section.n_lines
        b = result.bundle
        if section.label:
            sec = ws.cell(row=r, column=1, value=section.label)
            sec.font = Font(bold=True, color="FFFFFF")
            for col in range(1, 6):
                ws.cell(row=r, column=col).fill = FILL_SECTION
            r += 1
        first_check_row = r
        # row numbers of the netted pharma/fee pair (they reference each other)
        fee_net_row = next((first_check_row + i for i, c in enumerate(result.crosschecks)
                            if c.side_kind == "fee_net"), None)
        pharma_row = next((first_check_row + i for i, c in enumerate(result.crosschecks)
                           if c.side_kind == "ph_minus_fee"), None)
        for ci, chk in enumerate(result.crosschecks):
            cc_rows[(si, ci)] = r
            r = _crosscheck_row(ws, r, chk, b, sra_tab, n_lines,
                                fee_net_row, pharma_row)
    _autosize(ws)
    return cc_rows


def _crosscheck_row(ws, r: int, chk, b, sra_tab, n_lines,
                    fee_net_row, pharma_row) -> int:
    """Write ONE cross-check row; returns the next free row."""
    ws.cell(row=r, column=1, value=chk.name).font = F_INPUT
    is_phfee = "Φαρμακοποιού (packages" in chk.name or chk.side_kind == "fee_net"
    if is_phfee and b.phfee:
        # packages × unit price (READ from the report — 1.60/1.62 €)
        # as a LIVE formula off two blue inputs
        ws.cell(row=r, column=6, value=b.phfee.packages).font = F_INPUT
        _amount(ws, r, 7, b.phfee.unit_price, F_INPUT)
    def _sumifs(code_cols, cheques=()):
        """SUMIFS terms over the SRA Code column, criteria referencing
        helper cells (never quoted strings).  With `cheques`, one term per
        (code, cheque) pair adds a second criteria pair on the Cheque
        column — that is how a source file covering ONE cheque is compared
        with that cheque only."""
        terms, j = [], 8
        code_cells = []
        for code in code_cols:
            ws.cell(row=r, column=j, value=code).font = F_INPUT
            code_cells.append(f"{get_column_letter(j)}{r}")
            j += 1
        cheque_cells = []
        for cheque in cheques:
            ws.cell(row=r, column=j, value=cheque).font = F_INPUT
            cheque_cells.append(f"{get_column_letter(j)}{r}")
            j += 1
        base = (f"SUMIFS('{sra_tab}'!$F$2:$F${n_lines},"
                f"'{sra_tab}'!$A$2:$A${n_lines},")
        for cc in code_cells:
            if cheque_cells:
                for qc in cheque_cells:
                    terms.append(base + cc + f",'{sra_tab}'!$G$2:$G${n_lines},{qc})")
            else:
                terms.append(base + cc + ")")
        _sumifs.next_col = j
        return "+".join(terms)
    if chk.side_kind == "fee_net" and sra_tab and b.sra:
        # source = packages × unit (live); side = SRA PH − claims gross
        _amount(ws, r, 2, f"=F{r}*G{r}", F_FORMULA)
        side = "=" + _sumifs(["PH"])
        if pharma_row is not None:
            side += f"-B{pharma_row}"
        _amount(ws, r, 3, side, F_LINK)
    elif chk.side_kind == "ph_minus_fee" and sra_tab and b.sra:
        if is_phfee and b.phfee:
            _amount(ws, r, 2, f"=F{r}*G{r}", F_FORMULA)
        else:
            _amount(ws, r, 2, chk.source_total, F_INPUT)
        side = "=" + _sumifs(["PH"])
        if fee_net_row is not None:
            side += f"-F{fee_net_row}*G{fee_net_row}"
        _amount(ws, r, 3, side, F_LINK)
    elif chk.side_kind == "codes_minus" and sra_tab and b.sra:
        _amount(ws, r, 2, chk.source_total, F_INPUT)
        side = "=" + _sumifs(chk.sra_codes, chk.cheques)
        if abs(chk.minus) > 0.005:
            j = _sumifs.next_col
            ws.cell(row=1, column=j, value=chk.minus_label).font = F_HEADER
            ws.cell(row=1, column=j).fill = FILL_HEADER
            _amount(ws, r, j, chk.minus, F_INPUT)
            side += f"-{get_column_letter(j)}{r}"
        _amount(ws, r, 3, side, F_LINK)
    else:
        if is_phfee and b.phfee:
            _amount(ws, r, 2, f"=F{r}*G{r}", F_FORMULA)
        else:
            _amount(ws, r, 2, chk.source_total, F_INPUT)
        if sra_tab and chk.sra_codes and b.sra:
            # SUMIFS over the SRA Code column, criteria referencing the
            # code helper cells (never quoted strings)
            _amount(ws, r, 3, "=" + _sumifs(chk.sra_codes), F_LINK)
        elif chk.sra_side is not None:
            _amount(ws, r, 3, chk.sra_side, F_INPUT)
    if chk.sra_side is not None:
        _amount(ws, r, 4, f"=B{r}-C{r}", F_FORMULA)
        if chk.flag == "red":
            ws.cell(row=r, column=4).font = F_RED
        elif chk.flag == "amber":
            ws.cell(row=r, column=4).font = F_AMBER
    note = ws.cell(row=r, column=5, value=chk.note)
    if chk.flag == "amber":
        note.fill = FILL_AMBER
    return r + 1


_NAME_SPLIT_RE = re.compile(r"\s+(?:=|≈|vs)\s+")


def _name_side(name: str, first: bool) -> str:
    """The A or B half of a check name («X = Y», «X vs Y», «X ≈ Y») — used as
    the row label when a side has no itemised components."""
    bits = _NAME_SPLIT_RE.split(name)
    return (bits[0] if first else bits[-1]).strip() if len(bits) > 1 else name.strip()


# ------------------------------------------ tab: GL_Bridge (cash vs booked)

# ΟΑΥ's own ledger, bucket by bucket: which cost centres carry each stream
_GL_BRIDGE_ROWS = [
    (Bucket.INPATIENT, "Ενδονοσοκομειακή (Inpatient)",
     "26001 + 26002 + 26003 + 26007", ("regular_drg", "specialized",
                                       "z_catalogue_only", "per_diem")),
    (Bucket.AE, "ΤΑΕΠ (A&E)", "25801", ("ae",)),
    (Bucket.OUTPATIENT, "Εξωνοσοκομειακή & ΠΙ (Outpatient)",
     "25xxx κλινικά + 51001001", ("outpatient", "capitation")),
    (Bucket.PHARMA, "Φάρμακα (Pharma)", "25501 + λοιπά 255xx",
     ("pharmacist_fee", "pharma_other")),
]


def _tab_gl_bridge(wb: Workbook, result: ReconResult,
                   sra_tab: Optional[str] = None) -> None:
    """Cash vs booked, on one page: what the cheque paid per bucket (panel A),
    what the ΟΑΥ ledger booked for the same bucket (panel B), and the variance
    (panel C).

    Panel A links to the Reconciliation tab, so it is the same number the
    cheque ties to; panel B is the GL extract's own cost centres. The variance
    column is live, and the bottom zero-check proves the per-bucket variances
    add up to (SRA total − GL total) — no gap can hide in a rounding."""
    b = result.bundle
    if not b.gl or result.crosscheck_mode or not b.sra:
        return
    ws = wb.create_sheet("GL_Bridge")
    hosp = HOSPITALS.get(b.hospital_code, (b.hospital_code, ""))[0]
    ws.cell(row=1, column=1,
            value=f"Γέφυρα ταμείου ↔ καθολικού ΟΑΥ (SRA cash vs booked GL) — "
                  f"{hosp} — {MONTH_NAMES_EL[b.month] if b.month else ''} "
                  f"{b.year or ''}").font = Font(bold=True, size=12, color=NAVY)
    _header(ws, 3, ["Καλάθι (Bucket)", "Α — Ταμείο SRA (cash) €",
                    "Κέντρα κόστους ΟΑΥ (GL cost centres)",
                    "Β — Καθολικό ΟΑΥ (booked) €", "Διαφορά Α−Β (Variance) €",
                    "Σημείωση (Note)"])
    r = 4
    first = r
    for bucket, label, centres, fields in _GL_BRIDGE_ROWS:
        ws.cell(row=r, column=1, value=label).font = F_INPUT
        # panel A: the cheque's own bucket figure, linked from Reconciliation
        recon_row = 4 + BUCKET_ORDER.index(bucket)
        _amount(ws, r, 2, f"='Reconciliation'!C{recon_row}", F_LINK)
        ws.cell(row=r, column=3, value=centres).font = F_INPUT
        # panel B: what the ΟΑΥ ledger booked — every component a blue input
        booked = round(sum(getattr(b.gl, f, 0.0) or 0.0 for f in fields), 2)
        _amount(ws, r, 4, booked, F_INPUT)
        _amount(ws, r, 5, f"=B{r}-D{r}", F_FORMULA)
        cash = result.buckets.get(bucket, 0.0)
        diff = round(cash - booked, 2)
        if abs(diff) > CENT:
            note, flag = _annotate_bridge(bucket, diff)
            cell = ws.cell(row=r, column=6, value=note)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=r, column=5).font = F_AMBER if flag == "amber" else F_RED
            if flag == "amber":
                cell.fill = FILL_AMBER
        else:
            ws.cell(row=r, column=6, value="OK — ταυτίζεται (ties out).").font = \
                Font(italic=True, color=GRAY)
        r += 1
    total_row = r
    ws.cell(row=total_row, column=1, value="ΣΥΝΟΛΟ (TOTAL)").font = Font(bold=True)
    for col in (2, 4, 5):
        letter = get_column_letter(col)
        _amount(ws, total_row, col,
                f"=SUM({letter}{first}:{letter}{total_row - 1})", F_FORMULA)
        ws.cell(row=total_row, column=col).font = Font(bold=True)
    r += 2
    # the cash side must also equal the cheque itself, and the variances must
    # add up to (cash − booked): two live checks, nothing asserted in prose
    cheque_row = r
    ws.cell(row=cheque_row, column=1, value="Επιταγή ΟΑΥ (HIO cheque)").font = Font(bold=True)
    if sra_tab:
        _amount(ws, cheque_row, 2, f"='Reconciliation'!C{4 + len(BUCKET_ORDER) + 1}",
                F_LINK)
    else:
        _amount(ws, cheque_row, 2, b.sra.stated_total, F_INPUT)
    r += 1
    ws.cell(row=r, column=1,
            value="Zero-check = ταμείο ανά καλάθι − επιταγή (must be 0)")
    _amount(ws, r, 2, f"=B{total_row}-B{cheque_row}", F_FORMULA)
    ws.cell(row=r, column=2).fill = FILL_CHECK
    r += 1
    ws.cell(row=r, column=1,
            value="Zero-check = άθροισμα διαφορών − (ταμείο − καθολικό) (must be 0)")
    _amount(ws, r, 5, f"=E{total_row}-(B{total_row}-D{total_row})", F_FORMULA)
    ws.cell(row=r, column=5).fill = FILL_CHECK
    r += 2
    ws.cell(row=r, column=1, value=(
        "Η διαφορά ΔΕΝ κλείνει με προσαρμογή: κάθε καλάθι δείχνει τις δύο "
        "πλευρές και το άνοιγμα, με τη σημείωση που το εξηγεί. Αναλυτικά ανά "
        "λογαριασμό: φύλλο Source_crosscheck (the gap is never plugged — see "
        "Source_crosscheck for the account-level detail).")
    ).font = Font(italic=True, color=GRAY)
    _autosize(ws)


def _annotate_bridge(bucket: Bucket, diff: float) -> tuple[str, str]:
    """Why a bucket's cash and booked figures differ — the known ΟΑΥ ledger
    classifications, stated, never absorbed."""
    if bucket == Bucket.INPATIENT:
        return ("Z-procedures/tail χρεωμένα σε κλινικούς λογαριασμούς στο "
                "καθολικό της ΟΑΥ — ταξινόμηση, όχι ταμείο (HIO-ledger "
                "classification, not cash)."), "amber"
    if bucket == Bucket.PHARMA:
        return ("Το καθολικό ΟΑΥ κρατά τα φάρμακα και την αμοιβή "
                "φαρμακοποιού ΚΑΘΑΡΑ από τις διορθώσεις του μήνα (CRN/OTC, "
                "CRN-Packages)· οι τακτοποιήσεις EOAF πάνε στον 11202192."), "amber"
    if bucket == Bucket.OUTPATIENT:
        return ("Επιταγές δορυφορικών παροχέων και προσαρμογές μεθόδου "
                "αποζημίωσης μένουν εκτός του GL αυτού του παρόχου."), "amber"
    return ("Ανεξήγητη διαφορά (unexplained difference) — δείτε το "
            "Source_crosscheck."), "red"


# ------------------------------ tab: Απαιτήσεις_vs_SRA (claims file → SRA)

# DR SEGMENT in the «Πληρωμένες Απαιτήσεις all» file → the SRA's daily code
_SEGMENT_CODES = [
    ("Inpatient", ("IS",), "Ενδονοσοκομειακή (Inpatient)"),
    ("A&E", ("AE", "A&E"), "ΤΑΕΠ (A&E)"),
    ("Outpatient Specialists", ("OS",), "Ειδικοί Ιατροί (Outpatient Specialists)"),
    ("Nurses-Midwives", ("NM",), "Νοσηλευτές/Μαίες (Nurses-Midwives)"),
    ("Allied Health", ("AP",), "Άλλοι Επαγγελματίες Υγείας (Allied Health)"),
    ("Personal Doctors", ("PD",), "Προσωπικοί Ιατροί (Personal Doctors)"),
]


def _segment_code(sra, candidates: tuple) -> str:
    """ΟΑΥ writes the A&E line as «AE» in some months and «A&E» in others —
    use whichever the cheque actually carries."""
    present = {ln.code for ln in sra.lines}
    return next((c for c in candidates if c in present), candidates[0])

# every other service code the SRA can pay — things the claims export does
# not contain, named so the two sides close without a residual line
_RECON_LABELS = {
    "HEMO": "Αιμοκάθαρση — μηνιαία αναφορά (hemodialysis report)",
    "IS-ADJ": "Ενδονοσοκομειακή — προσαρμογή παραπομπών ΤΑΕΠ (A&E-referral adj.)",
    "AE-ADJ": "ΤΑΕΠ — προσαρμογή (A&E adjustment)",
    "OS-ADJ": "Εξωνοσοκομειακή — προσαρμογή μεθόδου αποζημίωσης (reimb.-method adj.)",
    "PD-CAP": "Κατά κεφαλήν ΠΙ (capitation — δεν τιμολογείται ανά πράξη)",
    "PD-FP": "Σταθερές χρεώσεις ΠΙ: OOH, εμβολιασμοί (PD fixed price)",
    "PD-KPI": "Ποιοτικά κριτήρια ΠΙ (PD quality criteria)",
    "KPI": "Ποιοτικά κριτήρια (quality criteria)",
    "MRI": "Ποιοτικά κριτήρια MRI (MRI)",
    "CT": "Ποιοτικά κριτήρια CT (CT)",
    "MRI/CT": "Ποιοτικά κριτήρια MRI/CT",
    "SAT": "Επιταγές δορυφορικών παροχέων (satellite suppliers)",
    "IS-PRIOR": "Τακτοποίηση παλαιότερης περιόδου (prior-period settlement)",
}


def _tab_claims_bridge(wb: Workbook, result: ReconResult,
                       sra_tab: Optional[str], n_lines: int) -> None:
    """The «Πληρωμένες Απαιτήσεις all» export (A&E included) reconciled to the
    SRA, segment by segment.

    Panel A is the claims file's own DR SEGMENT totals against the SRA's daily
    line for that stream.  Panel B is every OTHER service code the cheque
    pays — built from the codes actually present on the SRA tab, so the two
    panels together are the whole non-pharma cheque by construction and the
    zero-check underneath is a real identity, not a residual line.  Panel C
    then names what explains the panel-A gap and leaves the rest visible."""
    b = result.bundle
    if result.crosscheck_mode or not b.sra or not sra_tab or not b.claims:
        return
    ws = wb.create_sheet("Απαιτήσεις_vs_SRA")
    hosp = HOSPITALS.get(b.hospital_code, (b.hospital_code, ""))[0]
    ws.cell(row=1, column=1,
            value=f"Συμφωνία «Πληρωμένες Απαιτήσεις all» (+ΤΑΕΠ) με το SRA "
                  f"(claims export → SRA) — {hosp} — "
                  f"{MONTH_NAMES_EL[b.month] if b.month else ''} "
                  f"{b.year or ''}").font = Font(bold=True, size=12, color=NAVY)
    _header(ws, 3, ["Ροή / γραμμή (Stream / line)",
                    "Α — Αρχείο ΟΑΥ (claims export) €", "Κωδικός SRA (code)",
                    "Β — SRA €", "Διαφορά Α−Β (Diff) €", "Σημείωση (Note)"])

    def sumifs(code_cell: str) -> str:
        return (f"=SUMIFS('{sra_tab}'!$F$2:$F${n_lines},"
                f"'{sra_tab}'!$A$2:$A${n_lines},{code_cell})")

    r = 4
    ws.cell(row=r, column=1, value="Α. Ανά DR SEGMENT (per DR SEGMENT)").font = \
        Font(bold=True, color=BLUE)
    r += 1
    seg_first = r
    codes_seen = {ln.code for ln in b.sra.lines}
    for segment, candidates, label in _SEGMENT_CODES:
        code = _segment_code(b.sra, candidates)
        amount = b.claims.by_segment.get(segment)
        if amount is None and code not in codes_seen:
            continue
        ws.cell(row=r, column=1, value=label).font = F_INPUT
        _amount(ws, r, 2, round(amount or 0.0, 2), F_INPUT)
        ws.cell(row=r, column=3, value=code).font = F_INPUT
        _amount(ws, r, 4, sumifs(f"$C{r}"), F_LINK)
        _amount(ws, r, 5, f"=B{r}-D{r}", F_FORMULA)
        diff = round((amount or 0.0) - sra_sum(b.sra, [code]), 2)
        if abs(diff) > CENT:
            ws.cell(row=r, column=5).font = F_RED
            ws.cell(row=r, column=6, value=_annotate_segment(segment, result)
                    ).alignment = Alignment(wrap_text=True, vertical="top")
        else:
            ws.cell(row=r, column=6, value="OK — ταυτίζεται (ties out).").font = \
                Font(italic=True, color=GRAY)
        r += 1
    seg_total = r
    ws.cell(row=seg_total, column=1,
            value="Σύνολο ημερησίων γραμμών (daily service lines)").font = Font(bold=True)
    for col in (2, 4, 5):
        letter = get_column_letter(col)
        _amount(ws, seg_total, col,
                f"=SUM({letter}{seg_first}:{letter}{seg_total - 1})", F_FORMULA)
        ws.cell(row=seg_total, column=col).font = Font(bold=True)
    r += 2

    # panel B — the rest of the cheque's service lines, straight off the SRA
    ws.cell(row=r, column=1,
            value="Β. Γραμμές SRA εκτός του αρχείου claims (SRA lines the claims "
                  "export does not carry)").font = Font(bold=True, color=BLUE)
    r += 1
    other_first = r
    daily = {_segment_code(b.sra, c) for _, c, _ in _SEGMENT_CODES}
    rest_codes: list[str] = []
    for line in b.sra.lines:
        if line.bucket == Bucket.PHARMA or line.code in daily:
            continue
        if line.code not in rest_codes:
            rest_codes.append(line.code)
    for code in rest_codes:
        label = _RECON_LABELS.get(code)
        if not label:
            desc = next((ln.description for ln in b.sra.lines
                         if ln.code == code and ln.description), "")
            label = f"{code} — {desc}" if desc else code
        ws.cell(row=r, column=1, value=label).font = F_INPUT
        ws.cell(row=r, column=3, value=code).font = F_INPUT
        _amount(ws, r, 4, sumifs(f"$C{r}"), F_LINK)
        r += 1
    other_total = r
    ws.cell(row=other_total, column=1,
            value="Σύνολο λοιπών γραμμών (other service lines)").font = Font(bold=True)
    if rest_codes:
        _amount(ws, other_total, 4,
                f"=SUM(D{other_first}:D{other_total - 1})", F_FORMULA)
    else:
        _amount(ws, other_total, 4, 0.0, F_FORMULA)
    ws.cell(row=other_total, column=4).font = Font(bold=True)
    r += 2

    # panel C — completeness, then what explains the gap
    ws.cell(row=r, column=1,
            value="Γ. Έλεγχος πληρότητας και εξήγηση της διαφοράς (completeness "
                  "and explanation)").font = Font(bold=True, color=BLUE)
    r += 1
    svc_row = r
    ws.cell(row=svc_row, column=1,
            value="Σύνολο υπηρεσιών SRA — καλάθια Ενδονοσοκ. + ΤΑΕΠ + Εξωνοσοκ. "
                  "(SRA service buckets, pharma excluded)").font = Font(bold=True)
    _amount(ws, svc_row, 4,
            "='Reconciliation'!C4+'Reconciliation'!C5+'Reconciliation'!C6", F_LINK)
    r += 1
    ws.cell(row=r, column=1,
            value="Zero-check = ημερήσιες + λοιπές − σύνολο υπηρεσιών SRA (must be 0)")
    _amount(ws, r, 4, f"=D{seg_total}+D{other_total}-D{svc_row}", F_FORMULA)
    ws.cell(row=r, column=4).fill = FILL_CHECK
    r += 2
    gap_row = r
    ws.cell(row=gap_row, column=1,
            value="Διαφορά αρχείου προς SRA (claims export vs SRA daily lines)")
    _amount(ws, gap_row, 5, f"=E{seg_total}", F_FORMULA)
    r += 1
    first_expl = r
    cap_bundled = (b.capitation is not None
                   and "PD-CAP" not in {ln.code for ln in b.sra.lines})
    if cap_bundled:
        ws.cell(row=r, column=1,
                value="Πλέον: κατά κεφαλήν ΠΙ μέσα στις γραμμές PD — αναφορά capitation "
                      "(capitation paid inside the PD lines, not claimed per activity)"
                ).font = F_INPUT
        _amount(ws, r, 5, b.capitation.total, F_INPUT)
        r += 1
    ws.cell(row=r, column=1,
            value="Ανεξήγητη διαφορά (unexplained — never plugged)").font = Font(bold=True)
    _amount(ws, r, 5, f"=SUM(E{gap_row}:E{r - 1})", F_FORMULA)
    unexplained = round(_sum_segment_gap(result) + (b.capitation.total if cap_bundled
                                                    else 0.0), 2)
    ws.cell(row=r, column=5).font = Font(bold=True, color=("C00000" if abs(unexplained) > CENT
                                              else GREEN_LINK))
    ws.cell(row=r, column=6, value=(
        "Το άνοιγμα μένει ορατό: καμία γραμμή δεν το απορροφά (the gap is shown, "
        "never absorbed). Αναλυτικά ανά απαίτηση: φύλλο Source_crosscheck.")
    ).alignment = Alignment(wrap_text=True, vertical="top")
    if abs(unexplained) <= CENT:
        ws.cell(row=r, column=5).fill = FILL_CHECK
    _autosize(ws)


def _sum_segment_gap(result: ReconResult) -> float:
    """Σ (claims file − SRA daily line) across the DR SEGMENTs — the panel-A
    gap, recomputed in Python so the note can be coloured."""
    b = result.bundle
    total = 0.0
    for segment, candidates, _ in _SEGMENT_CODES:
        code = _segment_code(b.sra, candidates)
        amount = b.claims.by_segment.get(segment)
        if amount is None and not any(ln.code == code for ln in b.sra.lines):
            continue
        total += (amount or 0.0) - sra_sum(b.sra, [code])
    return round(total, 2)


def _annotate_segment(segment: str, result: ReconResult) -> str:
    b = result.bundle
    if segment == "Personal Doctors" and b.capitation:
        return ("Οι γραμμές PD του SRA περιέχουν και το κατά κεφαλήν "
                f"({format_eur(b.capitation.total)}), που δεν τιμολογείται ανά πράξη "
                "(the SRA PD lines also carry capitation, absent from the claims "
                "export).")
    if segment == "Inpatient":
        return ("Απαιτήσεις παλαιότερων περιόδων που πληρώθηκαν με αυτή την "
                "επιταγή λείπουν από τον μηνιαίο πίνακα (old-period claims "
                "sit outside the monthly table) — δείτε Source_crosscheck.")
    return ("Διαφορά προς διερεύνηση (difference to investigate) — δείτε "
            "Source_crosscheck.")


# ------------------------------------------- tab: Ανάλυση_ελέγχων (audit)

def _tab_audit(wb: Workbook, sections: list, cc_rows: dict) -> None:
    """Every Source_crosscheck row written out as a full reconciliation:
    each side broken into its components, live subtotals, the difference,
    and a tie-back cell proving this sheet agrees with Source_crosscheck.

    An auditor reads one block top to bottom and sees exactly which report
    figure, which SRA lines and which reconciling items make up each side —
    nothing is asserted in prose only."""
    ws = wb.create_sheet("Ανάλυση_ελέγχων")
    ws.cell(row=1, column=1,
            value="Ανάλυση ελέγχων — κάθε συμφωνία βήμα προς βήμα "
                  "(audit trail: every check, both sides, live)"
            ).font = Font(bold=True, color=NAVY)
    _header(ws, 3, ["Στοιχείο (Item)", "Ποσό (Amount €)", "Πηγή (Source)",
                    "Κωδικός SRA (code)", "Επιταγή (cheque)"])
    r = 5
    n = 0
    for si, section in enumerate(sections):
      result, sra_tab, n_lines = section.result, section.sra_tab, section.n_lines
      for i, chk in enumerate(result.crosschecks):
        if chk.sra_side is None:
            continue                      # nothing to reconcile against
        cc_row = cc_rows[(si, i)]         # its row on Source_crosscheck
        n += 1
        prefix = f"{section.label} — " if section.label else ""
        title = ws.cell(row=r, column=1, value=f"{n}. {prefix}{chk.name}")
        title.font = Font(bold=True, color="FFFFFF")
        title.fill = FILL_SECTION
        for col in range(2, 6):
            ws.cell(row=r, column=col).fill = FILL_SECTION
        r += 1

        def _side(parts, label, fallback_amount, fallback_label, use_codes):
            """Write one side's components; returns (first_row, last_row)."""
            nonlocal r
            ws.cell(row=r, column=1, value=label).font = Font(bold=True)
            r += 1
            first = r
            rows = list(parts)
            if not rows and use_codes and chk.sra_codes and sra_tab:
                rows = [_Part(f"SRA γραμμές {code}", 0.0, code, chk.cheques)
                        for code in chk.sra_codes]
            if not rows:
                rows = [_Part(fallback_label, fallback_amount, "", [])]
            else:
                # never let an itemisation silently miss part of its side
                itemised = round(sum(p.amount for p in rows), 2)
                if not (use_codes and any(p.code for p in rows)):
                    gap = round(fallback_amount - itemised, 2)
                    if abs(gap) > CENT:
                        rows = rows + [_Part("Λοιπά μη αναλυμένα (not itemised)",
                                             gap, "", [])]
            for part in rows:
                ws.cell(row=r, column=1, value="   " + part.label).font = F_INPUT
                if part.code and sra_tab:
                    ws.cell(row=r, column=4, value=part.code).font = F_INPUT
                    crit = f"'{sra_tab}'!$A$2:$A${n_lines},D{r}"
                    if part.cheques:
                        # criteria always reference helper CELLS, never
                        # quoted strings — one cell per cheque, from col E
                        terms = []
                        for k, q in enumerate(part.cheques):
                            ws.cell(row=r, column=5 + k, value=q).font = F_INPUT
                            qc = f"{get_column_letter(5 + k)}{r}"
                            terms.append(
                                f"SUMIFS('{sra_tab}'!$F$2:$F${n_lines},{crit},"
                                f"'{sra_tab}'!$G$2:$G${n_lines},{qc})")
                        _amount(ws, r, 2, "=" + "+".join(terms), F_LINK)
                    else:
                        _amount(ws, r, 2,
                                f"=SUMIFS('{sra_tab}'!$F$2:$F${n_lines},{crit})",
                                F_LINK)
                    ws.cell(row=r, column=3, value=sra_tab).font = F_INPUT
                else:
                    _amount(ws, r, 2, part.amount, F_INPUT)
                    ws.cell(row=r, column=3, value="Αναφορά ΟΑΥ").font = F_INPUT
                r += 1
            last = r - 1
            ws.cell(row=r, column=1, value=f"   Σύνολο — {label}").font = Font(bold=True)
            _amount(ws, r, 2, f"=SUM(B{first}:B{last})", F_FORMULA)
            ws.cell(row=r, column=2).font = Font(bold=True)
            total_row = r
            r += 1
            return total_row

        a_total = _side(chk.parts_a, f"Α — {chk.label_a or 'Πηγή (source report)'}",
                        chk.source_total, _name_side(chk.name, True), False)
        default_b = "SRA" if sra_tab else "σύγκριση αναφοράς με αναφορά (report vs report)"
        b_total = _side(chk.parts_b, f"Β — {chk.label_b or default_b}",
                        chk.sra_side, _name_side(chk.name, False), True)

        ws.cell(row=r, column=1, value="Διαφορά Α − Β (difference)").font = Font(bold=True)
        _amount(ws, r, 2, f"=B{a_total}-B{b_total}", F_FORMULA)
        cell = ws.cell(row=r, column=2)
        cell.font = Font(bold=True)
        if chk.flag == "red":
            cell.font = F_RED
        elif chk.flag == "amber":
            cell.font = F_AMBER
        elif abs(chk.diff or 0) <= CENT:
            # a REAL zero: gate 5 recomputes it.  A check that is "ok" within
            # a documented tolerance (e.g. IS Auditor per-row rounding) keeps
            # its live difference visible and is NOT claimed to be zero.
            cell.fill = FILL_CHECK
        r += 1
        # provable consistency with Source_crosscheck: both sides must equal
        # the figures printed there
        for label, this_row, cc_col in (
                ("Έλεγχος: Σύνολο Α = Source_crosscheck (must be 0)", a_total, "B"),
                ("Έλεγχος: Σύνολο Β = Source_crosscheck (must be 0)", b_total, "C")):
            ws.cell(row=r, column=1, value=label)
            _amount(ws, r, 2,
                    f"=B{this_row}-'Source_crosscheck'!{cc_col}{cc_row}",
                    F_FORMULA)
            ws.cell(row=r, column=2).fill = FILL_CHECK
            r += 1
        if chk.note:
            note = ws.cell(row=r, column=1, value="Σημείωση (note): " + chk.note)
            note.font = Font(italic=True, color=GRAY)
            note.alignment = Alignment(wrap_text=True, vertical="top")
            r += 1
        r += 1
    _autosize(ws)


# --------------------------------------------------- tab 4: By_Clinic_Split

def _tab_split(wb: Workbook, result: ReconResult, stated_cell: Optional[str]) -> int:
    ws = wb.create_sheet("By_Clinic_Split")
    b = result.bundle
    hosp = HOSPITALS[b.hospital_code]
    ws.cell(row=1, column=1, value=f"Κατανομή ανά κλινική για SAP (By-clinic split) — "
                                   f"{hosp[0]} — {MONTH_NAMES_EL[b.month]} {b.year}"
            ).font = Font(bold=True, color=NAVY)
    # the inpatient fee splits three ways: DRG, daily treatments and the
    # Z-catalogue drugs/procedures — ΟΑΥ's own pivot lumps the last two
    # together under «FIXED FEE», the per-claim detail table tells them apart
    _header(ws, 3, ["Κλινική / Γραμμή (Clinic / Line)", "DRG €",
                    "Ημερήσιες θεραπείες (Daily treat.) €",
                    "Ζ-φάρμακα/πράξεις (Z-drugs) €", "Ποσό (Amount €)"])
    r = 4
    subtotal_cells = []
    for section in result.split:
        sec = ws.cell(row=r, column=1, value=section.title)
        sec.font = Font(bold=True, color="FFFFFF")
        sec.fill = FILL_SECTION
        r += 1
        first = r
        for row in section.rows:
            ws.cell(row=r, column=1, value=row.label).font = F_INPUT
            if row.drg is not None:
                _amount(ws, r, 2, row.drg, F_INPUT)
            if row.fixed_fee is not None:
                _amount(ws, r, 3, row.fixed_fee, F_INPUT)
            if row.z_drugs is not None:
                _amount(ws, r, 4, row.z_drugs, F_INPUT)
            _amount(ws, r, 5, row.amount, F_INPUT)
            r += 1
        ws.cell(row=r, column=1, value=f"Υποσύνολο (Subtotal) — {section.title}"
                ).font = Font(bold=True)
        # every column carries its own live subtotal, so the three inpatient
        # streams add up on the page as well as across
        for col in (2, 3, 4, 5):
            if r > first:
                letter = get_column_letter(col)
                _amount(ws, r, col, f"=SUM({letter}{first}:{letter}{r - 1})", F_FORMULA)
            else:
                _amount(ws, r, col, 0.0, F_FORMULA)
            ws.cell(row=r, column=col).font = Font(bold=True)
        for col in range(1, 6):
            ws.cell(row=r, column=col).border = THIN
        subtotal_cells.append(f"E{r}")
        r += 2
    total_row = r
    ws.cell(row=total_row, column=1, value="ΓΕΝΙΚΟ ΣΥΝΟΛΟ (GRAND TOTAL)").font = Font(bold=True, color=NAVY)
    _amount(ws, total_row, 5, "=" + "+".join(subtotal_cells), F_FORMULA)
    ws.cell(row=total_row, column=5).font = Font(bold=True)
    if stated_cell:
        cheque_row = total_row + 1
        ws.cell(row=cheque_row, column=1, value="Επιταγή ΟΑΥ (HIO cheque)")
        _amount(ws, cheque_row, 5, f"={stated_cell}", F_LINK)
        check_row = cheque_row + 1
        ws.cell(row=check_row, column=1, value="Zero-check = ΓΕΝΙΚΟ ΣΥΝΟΛΟ − επιταγή (must be 0)")
        _amount(ws, check_row, 5, f"=E{total_row}-E{cheque_row}", F_FORMULA)
        ws.cell(row=check_row, column=5).fill = FILL_CHECK
    else:
        ws.cell(row=total_row + 1, column=1,
                value="Cross-check mode: χωρίς επιταγή — no cash tie-out (δεν υπάρχει SRA).")
    _autosize(ws)
    return total_row


# ------------------------------------------- tab 5: by doctor & speciality

def _tab_by_doctor(wb: Workbook, result: ReconResult,
                   sra_tab: Optional[str] = None, n_lines: int = 0,
                   split_total_row: Optional[int] = None) -> None:
    """The SRA payment split by clinic/speciality AND doctor, summed from the
    ROW-LEVEL claims detail (never from ΟΑΥ-printed totals), plus the
    capitation per-doctor breakdown.  Live SUM subtotals per stream; bottom
    block re-ties the tab against the source-report column sums."""
    b = result.bundle
    docs = b.claims.by_doctor if b.claims else []
    cap_docs = b.capitation.by_doctor if b.capitation else []
    if not docs and not cap_docs:
        return
    ws = wb.create_sheet("Ανά_ιατρό")
    ws.cell(row=1, column=1,
            value="Ανάλυση πληρωμής ΟΑΥ ανά ειδικότητα/κλινική και ιατρό "
                  "(SRA payment by speciality & doctor) — αθροισμένη από τις "
                  "αναλυτικές γραμμές των αρχείων ΟΑΥ").font = \
        Font(bold=True, size=14, color=NAVY)
    _header(ws, 3, ["Ροή (Stream)", "Ειδικότητα (Speciality)",
                    "Ιατρός (Doctor)", "Ποσό (Amount €)"])
    r = 4
    subtotal_cells: list[str] = []
    segments: list[str] = []
    for seg, _sp, _d, _v in docs:
        if seg not in segments:
            segments.append(seg)
    for seg in segments:
        head = ws.cell(row=r, column=1, value=f"{seg} — Claims «all»")
        head.font = Font(bold=True)
        head.fill = FILL_SECTION
        r += 1
        # BY CLINIC FIRST, THEN BY DOCTOR: specialities ordered by size,
        # each with a live subtotal over its doctor rows beneath
        seg_rows = [(sp, d, v) for s, sp, d, v in docs if s == seg]
        spec_totals: dict[str, float] = {}
        for sp, _d, v in seg_rows:
            spec_totals[sp] = round(spec_totals.get(sp, 0.0) + v, 2)
        spec_cells: list[str] = []
        for sp in sorted(spec_totals, key=lambda k: -spec_totals[k]):
            drs = [(d, v) for s2, d, v in seg_rows if s2 == sp]
            ws.cell(row=r, column=2, value=sp).font = Font(bold=True)
            _amount(ws, r, 4, f"=SUM(D{r + 1}:D{r + len(drs)})", F_FORMULA)
            spec_cells.append(f"D{r}")
            r += 1
            for d, v in drs:
                ws.cell(row=r, column=3, value=d).font = F_INPUT
                _amount(ws, r, 4, v, F_INPUT)
                r += 1
        ws.cell(row=r, column=1, value=f"Υποσύνολο {seg}").font = Font(bold=True)
        _amount(ws, r, 4, "=" + "+".join(spec_cells), F_FORMULA)
        subtotal_cells.append(f"D{r}")
        r += 1
    if cap_docs:
        head = ws.cell(row=r, column=1,
                       value="Personal Doctors — Capitation report (κατά κεφαλήν)")
        head.font = Font(bold=True)
        head.fill = FILL_SECTION
        r += 1
        first = r
        for label, v in cap_docs:
            ws.cell(row=r, column=2, value="Capitation").font = F_INPUT
            ws.cell(row=r, column=3, value=label).font = F_INPUT
            _amount(ws, r, 4, v, F_INPUT)
            r += 1
        ws.cell(row=r, column=1, value="Υποσύνολο Capitation").font = Font(bold=True)
        _amount(ws, r, 4, f"=SUM(D{first}:D{r - 1})", F_FORMULA)
        subtotal_cells.append(f"D{r}")
        r += 1
    total_row = r
    ws.cell(row=total_row, column=1, value="ΣΥΝΟΛΟ καρτέλας (tab total)").font = \
        Font(bold=True)
    _amount(ws, total_row, 4, "=" + "+".join(subtotal_cells), F_FORMULA)
    r += 2
    # verification block: the tab re-ties against the source-report column
    # sums — a gap here means incomplete row-level detail, shown, never hidden
    src_rows = []
    if b.claims:
        ws.cell(row=r, column=1,
                value="Claims «all» — άθροιση στήλης HIO REIMB. (column sum)"
                ).font = F_INPUT
        _amount(ws, r, 4, b.claims.total, F_INPUT)
        src_rows.append(r)
        r += 1
    if b.capitation:
        ws.cell(row=r, column=1,
                value="Capitation report — άθροιση τιμολογίων EBS (invoice sum)"
                ).font = F_INPUT
        _amount(ws, r, 4, b.capitation.total, F_INPUT)
        src_rows.append(r)
        r += 1
    diff_row = r
    ws.cell(row=diff_row, column=1,
            value="Διαφορά καρτέλας − πηγών (πληρότητα αναλυτικών γραμμών / "
                  "detail completeness)")
    diff_cell = _amount(ws, diff_row, 4,
                        f"=D{total_row}-" + "-".join(f"D{x}" for x in src_rows),
                        F_FORMULA)
    tab_total = round(sum(v for *_x, v in docs)
                      + sum(v for _l, v in cap_docs), 2)
    src_total = round((b.claims.total if b.claims else 0.0)
                      + (b.capitation.total if b.capitation else 0.0), 2)
    if abs(tab_total - src_total) > 0.005:
        diff_cell.font = F_AMBER
        ws.cell(row=diff_row + 1, column=1,
                value="Μερική ανάλυση ανά ιατρό στην πηγή (η αναφορά ΟΑΥ δεν "
                      "αναλύει όλο το ποσό ανά ιατρό) — η διαφορά φαίνεται, "
                      "δεν κρύβεται.").font = F_AMBER
        r += 1
    r = diff_row + 2

    # ------- bridge: from the by-doctor universe to By_Clinic_Split / cheque
    # (only when an SRA exists — cross-check mode has no cash side)
    if sra_tab and src_rows:
        head = ws.cell(row=r, column=1,
                       value="ΓΕΦΥΡΑ ΠΡΟΣ ΤΟ BY_CLINIC_SPLIT / ΤΗΝ ΕΠΙΤΑΓΗ "
                             "(bridge: doctors → cheque)")
        head.font = Font(bold=True, color="FFFFFF")
        head.fill = FILL_SECTION
        r += 1
        bridge_rows: list[int] = []
        first_src = src_rows[0]
        ws.cell(row=r, column=1,
                value="Αποδιδόμενα σε ιατρούς: Claims «all» + Capitation "
                      "(οι πηγές της καρτέλας)").font = F_FORMULA
        _amount(ws, r, 4, "=" + "+".join(f"D{x}" for x in src_rows), F_FORMULA)
        bridge_rows.append(r)
        r += 1

        def sumifs_codes(row: int, codes: list[str], col_letter_ref: str = "A") -> str:
            terms = []
            for k, code in enumerate(codes):
                col = get_column_letter(6 + k)
                ws.cell(row=row, column=6 + k, value=code).font = F_INPUT
                terms.append(
                    f"SUMIFS('{sra_tab}'!$F$2:$F${n_lines},"
                    f"'{sra_tab}'!${col_letter_ref}$2:${col_letter_ref}${n_lines},"
                    f"{col}{row})")
            return "=" + "+".join(terms)

        for label, codes, by_bucket in [
            ("+ Φάρμακα — μη αποδιδόμενα σε ιατρούς (SRA bucket Pharma)",
             ["Pharma"], True),
            ("+ Αιμοκάθαρση (HEMO)", ["HEMO"], False),
            ("+ Προσαρμογές & τακτοποιήσεις (OS-ADJ, IS-ADJ, AE-ADJ, IS-PRIOR)",
             ["OS-ADJ", "IS-ADJ", "AE-ADJ", "IS-PRIOR"], False),
            ("+ Σταθερές χρεώσεις ΠΙ & Ποιοτικά (PD-FP, KPI, MRI/CT)",
             ["PD-FP", "PD-KPI", "KPI", "MRI", "CT", "MRI/CT"], False),
            ("+ Επιταγές δορυφορικών παροχέων (SAT — π.χ. κέντρα υγείας F1085)",
             ["SAT"], False),
        ]:
            ws.cell(row=r, column=1, value=label).font = F_LINK
            _amount(ws, r, 4,
                    sumifs_codes(r, codes, "D" if by_bucket else "A"), F_LINK)
            bridge_rows.append(r)
            r += 1
        bridge_total_row = r
        ws.cell(row=r, column=1, value="Σύνολο γέφυρας (bridge total)"
                ).font = Font(bold=True)
        _amount(ws, r, 4, "=" + "+".join(f"D{x}" for x in bridge_rows), F_FORMULA)
        r += 1
        split_row = r
        ws.cell(row=r, column=1,
                value="ΓΕΝΙΚΟ ΣΥΝΟΛΟ By_Clinic_Split (= επιταγή ΟΑΥ)")
        _amount(ws, r, 4, f"='By_Clinic_Split'!E{split_total_row}", F_LINK)
        r += 1
        ws.cell(row=r, column=1,
                value="Διαφορά γέφυρας — γραμμές SRA χωρίς αναλυτικό ανά ιατρό "
                      "(προσαρμογές OS/NM/AP/PD, επιταγές δορυφορικών παροχέων, "
                      "υπόλοιπο ανάλυσης)")
        d = _amount(ws, r, 4, f"=D{bridge_total_row}-D{split_row}", F_FORMULA)
        d.font = F_AMBER
    _autosize(ws)


# ------------------------------------------------- tab 6: how reports tie

# One universe, many projections: every document in the batch is issued by
# the ΟΑΥ (HIO) about the SAME paid population.  The rows below are the
# identities verified to the cent on real months (Feb/Apr/May 2026); the
# join keys are PAYMENT NO. (the cheque) and the EBS invoice IDs.
TRUTH_MAP_ROWS = [
    ("Ροή (stream)", "Ταυτότητα (identity)", "Κλειδί / σημείωση (key / note)"),
    ("Επιταγή (cheque)",
     "Άθροισμα γραμμών SRA = δηλωμένο σύνολο επιταγής",
     "Το SRA είναι η σπονδυλική στήλη του χρήματος — κάθε γραμμή του είναι "
     "τιμολόγιο EBS της ΟΑΥ."),
    ("Ενδονοσοκομειακή (IS)",
     "SRA IS (ημερήσιες) = Claims «all»·Inpatient = Ενδ. Σύνολο = "
     "IS Auditor DRG+Z (± στρογγυλοποίηση)",
     "Τετραπλό δέσιμο σε έναν αριθμό. Απαιτήσεις παλαιών περιόδων που "
     "πληρώνονται τώρα λείπουν από την Ενδ. — κατονομάζονται."),
    ("ΤΑΕΠ (AE)",
     "SRA AE (ημερήσιες) = Claims «all»·A&E = GL ΟΑΥ 25801 "
     "(51101099 − 43010001 co-pays)",
     "Οι προσαρμογές παραπομπών (AE-ADJ/IS-ADJ) μένουν εκτός των ημερησίων."),
    ("Εξωνοσοκομειακή (OS/NM/AP)",
     "SRA ημερήσιες = Claims «all» segments = XML activities",
     "Το XML δένει σε επίπεδο πράξης μέσω ClaimPaymentNumber (PAYMENT NO.)."),
    ("Προσωπικοί Ιατροί (PD)",
     "SRA PD (ημερήσιες) = Capitation report + Claims «Personal Doctors»",
     "Επαληθευμένο στο σεντ Απρ+Μάι 2026. Σταθερές χρεώσεις (OOH, "
     "εμβολιασμοί) χωριστά ως PD-FP."),
    ("Ποιοτικά κριτήρια (KPI/MRI)",
     "SRA γραμμές KPI/MRI-CT = εξαγωγή Ποιοτικών Κριτηρίων",
     "Κενή εξαγωγή = εύρημα, όχι μηδενισμός."),
    ("Φάρμακα (PH)",
     "SRA PH (ημερήσιες) = Πληρωμένες ΦΑΡΜΑΚΑ (Drugs+Consumables) + "
     "Αμοιβή Φαρμακοποιού (packages × τιμή μονάδας)",
     "Επαληθευμένο στο σεντ Φεβ+Απρ+Μάι 2026. CRN/OTC/ISSUANCES χωριστά "
     "ως PH-ADJ· CRN-Packages ως PHF."),
    ("Αιμοκάθαρση (HEMO)",
     "SRA HEMO = μηνιαία αναφορά αιμοκάθαρσης",
     "Ενδονοσοκομειακή ή εξωνοσοκομειακή ανά ασθενή — μπλε κελί Bucket."),
    ("GL ΟΑΥ (καθολικό)",
     "26xxx = SRA IS + HEMO + IS-ADJ · 25801 = AE · 51001001 = capitation "
     "· 255xx ≈ φάρμακα · λοιπά 25xxx + capitation = εξωνοσοκομειακά",
     "Η λογιστική όψη της ΟΑΥ για τα ίδια ποσά. Γνωστές ταξινομήσεις: "
     "Z-tail σε κλινικούς λογαριασμούς, αμοιβή φαρμακοποιού flat."),
    ("Προσαρμογές (ADJ/CRN)",
     "PH-ADJ / AE-ADJ / IS-ADJ — το στρώμα διορθώσεων",
     "Δένουν με contra λογαριασμούς GL (π.χ. ISSUANCES ↔ 11202192 "
     "Unearned Revenue EOAF)."),
    ("Τακτοποιήσεις (PRIOR)",
     "Μονογραμμικές επιταγές παλαιών περιόδων (year-end DRG, "
     "innovative antibiotics)",
     "Pass-through: εκτός όλων των μηνιαίων ελέγχων, δικές τους γραμμές "
     "στο By_Clinic_Split."),
    ("Δορυφορικοί παροχείς",
     "Δικός τους κωδικός F στην κεφαλίδα SRA (π.χ. F1085) και δικός τους "
     "GL vendor",
     "Οι επιταγές τους μετρούν στο ταμείο του μήνα αλλά όχι στα αρχεία "
     "claims/GL του νοσοκομείου."),
]


def _tab_truth_map(wb: Workbook) -> None:
    ws = wb.create_sheet("Πώς_δένουν")
    ws.cell(row=1, column=1,
            value="Πώς δένουν οι αναφορές ΟΑΥ μεταξύ τους "
                  "(how the HIO reports tie together)").font = \
        Font(bold=True, size=14, color=NAVY)
    ws.cell(row=2, column=1,
            value="Όλα τα έγγραφα είναι εκδόσεις της ΟΑΥ για τον ίδιο πληρωμένο "
                  "πληθυσμό — κάθε αναφορά είναι διαφορετική προβολή του. "
                  "Κλειδιά σύνδεσης: PAYMENT NO. (αρ. επιταγής) και EBS invoice "
                  "IDs. Οι ταυτότητες επαληθεύτηκαν στο σεντ σε πραγματικούς "
                  "μήνες (Φεβ/Απρ/Μάι 2026).").font = Font(italic=True, color=GRAY)
    r = 4
    for i, (stream, identity, note) in enumerate(TRUTH_MAP_ROWS):
        if i == 0:
            _header(ws, r, list(TRUTH_MAP_ROWS[0]))
        else:
            ws.cell(row=r, column=1, value=stream).font = Font(bold=True, color=BLUE)
            ws.cell(row=r, column=2, value=identity)
            ws.cell(row=r, column=3, value=note).font = Font(color=GRAY)
        r += 1
    _autosize(ws)


# ----------------------------------------------------------- tab 6: Legend

def _tab_legend(wb: Workbook, result: ReconResult) -> None:
    ws = wb.create_sheet("Legend")
    ws.cell(row=1, column=1, value="Υπόμνημα (Legend)").font = Font(bold=True, size=14, color=NAVY)
    rows = [
        ("Μπλε γραμματοσειρά (blue font)", "Hardcoded input από αναφορά ΟΑΥ (off a source report)", F_INPUT, None),
        ("Μαύρη γραμματοσειρά (black font)", "Ζωντανός τύπος (live formula)", F_FORMULA, None),
        ("Πράσινη γραμματοσειρά (green font)", "Σύνδεσμος μεταξύ φύλλων (cross-sheet link)", F_LINK, None),
        ("Κίτρινο γέμισμα (yellow fill)", "Zero-check — πρέπει να είναι 0 (must read 0)", None, FILL_CHECK),
        ("Πορτοκαλί (amber)", "Γνωστή απόκλιση με σημείωση (known variance, noted)", F_AMBER, FILL_AMBER),
        ("Κόκκινο (red)", "Ανεξήγητη διαφορά — εύρημα (unexplained diff, a finding)", F_RED, None),
    ]
    r = 3
    for label, meaning, font, fill in rows:
        c = ws.cell(row=r, column=1, value=label)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        ws.cell(row=r, column=2, value=meaning)
        r += 1
    r += 1
    notes = [
        "Source_crosscheck: οι στήλες Α και Β ακολουθούν τη σειρά του ονόματος "
        "του ελέγχου. Π.χ. «GL ΟΑΥ 25501 vs Αναφορά Αμοιβής» → Α = το ποσό του "
        "καθολικού ΟΑΥ, Β = το ποσό της αναφοράς αμοιβής (packages × τιμή).",
        "Ανάλυση_ελέγχων: κάθε έλεγχος του Source_crosscheck γραμμένος αναλυτικά — "
        "τα συστατικά κάθε πλευράς, ζωντανά υποσύνολα, η διαφορά, και δύο κελιά "
        "που αποδεικνύουν ότι το φύλλο συμφωνεί με το Source_crosscheck "
        "(audit trail: every check, both sides, component by component).",
        "Κάθε υποσύνολο/σύνολο/διαφορά είναι ζωντανός τύπος — αλλάζοντας ένα μπλε κελί, "
        "το βιβλίο ξανα-δένει ή δείχνει το σπάσιμο.",
        "Never plug a difference: κάθε ανεξήγητη διαφορά εμφανίζεται με τις δύο πλευρές και το άνοιγμα.",
        "Stateless: τίποτα δεν αποθηκεύεται μετά το κλείσιμο του browser session.",
    ]
    for n in notes:
        ws.cell(row=r, column=1, value=n)
        r += 1
    _autosize(ws)


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 2, 70)


# =================================================== gate 5: verification

_TOKEN_RE = re.compile(r"""
    (?P<func>SUMIFS|SUM|MAX|MIN|ROUND)\(
  | (?P<ref>(?:'[^']+'!)?\$?[A-Z]{1,3}\$?\d+(?::\$?[A-Z]{1,3}\$?\d+)?)
  | (?P<num>\d+(?:\.\d+)?)
  | (?P<op>[+\-*/(),])
""", re.VERBOSE)


class _Evaluator:
    """Just enough of a formula engine to recompute the formulas THIS app
    writes: SUM/SUMIFS/MAX/MIN over (cross-sheet) ranges, cell refs, + - * /."""

    def __init__(self, wb):
        self.wb = wb

    def cell_value(self, sheet: str, coord: str) -> float:
        v = self.wb[sheet][coord.replace("$", "")].value
        if v is None:
            return 0.0
        if isinstance(v, str) and v.startswith("="):
            # fresh evaluator: evaluate() is stateful, recursing on self would
            # clobber the caller's token position
            return _Evaluator(self.wb).evaluate(v, sheet)
        if isinstance(v, (int, float)):
            return float(v)
        return 0.0

    def cell_raw(self, sheet: str, coord: str):
        return self.wb[sheet][coord.replace("$", "")].value

    def _range_cells(self, sheet: str, ref: str) -> list[tuple[str, str]]:
        if "!" in ref:
            sheet_part, ref = ref.split("!")
            sheet = sheet_part.strip("'")
        ref = ref.replace("$", "")
        if ":" in ref:
            min_c, min_r, max_c, max_r = range_boundaries(ref)
            return [(sheet, f"{get_column_letter(c)}{r}")
                    for r in range(min_r, max_r + 1) for c in range(min_c, max_c + 1)]
        return [(sheet, ref)]

    def evaluate(self, formula: str, sheet: str) -> float:
        self.tokens = list(_TOKEN_RE.finditer(formula.lstrip("=")))
        self.pos = 0
        self.sheet = sheet
        return self._expr()

    def _peek(self):
        return self.tokens[self.pos] if self.pos < len(self.tokens) else None

    def _next(self):
        t = self._peek()
        self.pos += 1
        return t

    def _expr(self) -> float:
        v = self._term()
        while (t := self._peek()) and t.group("op") in ("+", "-"):
            op = self._next().group("op")
            rhs = self._term()
            v = v + rhs if op == "+" else v - rhs
        return v

    def _term(self) -> float:
        v = self._factor()
        while (t := self._peek()) and t.group("op") in ("*", "/"):
            op = self._next().group("op")
            rhs = self._factor()
            v = v * rhs if op == "*" else v / rhs
        return v

    def _factor(self) -> float:
        t = self._next()
        if t is None:
            return 0.0
        if t.group("num"):
            return float(t.group("num"))
        if t.group("op") == "-":
            return -self._factor()
        if t.group("op") == "(":
            v = self._expr()
            self._next()  # ')'
            return v
        if t.group("ref"):
            cells = self._range_cells(self.sheet, t.group("ref"))
            if len(cells) == 1:
                return self.cell_value(*cells[0])
            return sum(self.cell_value(s, c) for s, c in cells)
        if t.group("func"):
            return self._call(t.group("func"))
        return 0.0

    def _args(self) -> list:
        """Argument list; each arg is either a float or ('RANGE', ref)."""
        args = []
        depth = 1
        current_start = self.pos
        while self.pos < len(self.tokens):
            t = self.tokens[self.pos]
            op = t.group("op")
            if op == "(":
                depth += 1
            elif op == ")":
                depth -= 1
                if depth == 0:
                    if self.pos > current_start:
                        args.append(self._arg_slice(current_start, self.pos))
                    self.pos += 1
                    return args
            elif op == "," and depth == 1:
                args.append(self._arg_slice(current_start, self.pos))
                current_start = self.pos + 1
            self.pos += 1
        return args

    def _arg_slice(self, start: int, end: int):
        toks = self.tokens[start:end]
        if len(toks) == 1 and toks[0].group("ref"):
            return ("RANGE", toks[0].group("ref"))
        sub = _Evaluator(self.wb)
        sub.tokens = toks
        sub.pos = 0
        sub.sheet = self.sheet
        return sub._expr()

    def _call(self, name: str) -> float:
        args = self._args()

        def cells_of(arg):
            assert isinstance(arg, tuple) and arg[0] == "RANGE"
            return self._range_cells(self.sheet, arg[1])

        def vals(arg):
            if isinstance(arg, tuple):
                return [self.cell_value(s, c) for s, c in cells_of(arg)]
            return [arg]

        if name == "SUM":
            return sum(v for a in args for v in vals(a))
        if name in ("MAX", "MIN"):
            pool = []
            for a in args:
                if isinstance(a, tuple):
                    for s, c in cells_of(a):
                        raw = self.cell_raw(s, c)
                        if raw is not None:
                            pool.append(self.cell_value(s, c))
                else:
                    pool.append(a)
            return (max if name == "MAX" else min)(pool) if pool else 0.0
        if name == "ROUND":
            return round(args[0], int(args[1]))
        if name == "SUMIFS":
            # any number of (criteria range, criteria) pairs — a row is
            # counted only when EVERY pair matches
            sum_cells = cells_of(args[0])
            pairs = []
            for k in range(1, len(args) - 1, 2):
                crit = args[k + 1]
                if isinstance(crit, tuple):
                    s, c = cells_of(crit)[0]
                    crit_val = self.cell_raw(s, c)
                else:
                    crit_val = crit
                pairs.append((cells_of(args[k]), crit_val))
            total = 0.0
            for i, (ss, sc) in enumerate(sum_cells):
                if all(self.cell_raw(*cells[i]) == val for cells, val in pairs):
                    total += self.cell_value(ss, sc)
            return total
        raise ValueError(f"unsupported function {name}")


def verify_workbook(data: bytes,
                    documented_residual: float = 0.0) -> list[tuple[str, str, float]]:
    """Reopen the built workbook and recompute every yellow zero-check cell.
    Returns [(sheet, cell, recomputed value)] for cells NOT reading 0.

    documented_residual: a known SRA parsing difference (lines − stated) that
    is documented as a red row in Source_crosscheck — zero-checks reading
    exactly that value are accepted, per the brief's documented-variances
    clause.  Never silently absorbed: it stays visible."""
    wb = load_workbook(io.BytesIO(data))
    ev = _Evaluator(wb)
    failures = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.fill is not None and cell.fill.fgColor is not None \
                        and str(cell.fill.fgColor.rgb).endswith("FFFF00"):
                    v = cell.value
                    if isinstance(v, str) and v.startswith("="):
                        val = ev.evaluate(v, ws.title)
                    elif isinstance(v, (int, float)):
                        val = float(v)
                    else:
                        continue  # legend colour swatches carry no check value
                    if abs(val) > CENT and abs(val - documented_residual) > CENT:
                        failures.append((ws.title, cell.coordinate, round(val, 2)))
    return failures
